import os
import sys
import json

import time
import shutil
import sqlite3
import getpass
import platform
from datetime import datetime, timedelta
from typing import Optional, List, Tuple, Dict, Any, Callable
import subprocess 
import ctypes
import hashlib

from PySide6.QtWidgets import QCheckBox, QScrollArea
from PySide6.QtCore import Qt, QUrl, QPoint, QSize, QRect, QEvent, QEasingCurve, QPropertyAnimation, QTimer, QStandardPaths, QObject, Signal, QRunnable, QThreadPool, Property
from PySide6.QtGui import QAction, QDesktopServices, QIcon, QPixmap, QPainter, QColor, QPalette, QBrush, QPen, QLinearGradient
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, 
    QListWidget, QListWidgetItem, QToolButton, QLabel, QLineEdit,
    QFileDialog, QAbstractItemView, QDialog, QFormLayout, QStylePainter,
    QDialogButtonBox, QMenu, QStackedWidget, QFrame, QStyle, QStyleOptionComboBox,
    QTableWidget, QTableWidgetItem, QGraphicsOpacityEffect, QHeaderView,
    QSizePolicy, QPlainTextEdit, QComboBox, QGridLayout, QListView,
    QProgressBar
)

APP_TITLE = "Система управления документами"
APP_VERSION = "1.0.0"
SETTINGS_FILE = "settings.json"
REGISTRY_FILE = "registry.json"

WINDOW_W = 1160
WINDOW_H = 690
LEFT_W = 320
KEEP_SELECTION = object()

DOC_STATUS_DEFAULT = "Черновик"
DOC_STATUSES = [
    "Черновик",
    "На согласовании",
    "Действует",
    "На пересмотре",
    "Архив",
]
DOC_STATUS_COLORS = {
    "Черновик": ("#e5e7eb", "rgba(148, 163, 184, 0.18)"),
    "На согласовании": ("#fef3c7", "rgba(245, 158, 11, 0.22)"),
    "Действует": ("#dcfce7", "rgba(34, 197, 94, 0.22)"),
    "На пересмотре": ("#ffedd5", "rgba(249, 115, 22, 0.24)"),
    "Архив": ("#dbeafe", "rgba(59, 130, 246, 0.20)"),
}

THEME_SURFACE_APP = "#0B0713"
THEME_SURFACE_CARD = "#120D1C"
THEME_SURFACE_ELEVATED = "#171224"
THEME_ACCENT = "#C8AEFF"
THEME_ACCENT_HOVER = "#B592FF"
THEME_TEXT = "#F8F4FF"
THEME_TEXT_SOFT = "#F3EEFF"
THEME_MUTED = "#B8B0C9"
THEME_MUTED_DISABLED = "#6E6680"
THEME_STROKE = "rgba(200, 174, 255, 0.14)"

COMBO_POPUP_STYLESHEET = f"""
    QWidget#comboPopupWindow, QFrame#comboPopupWindow,
    QWidget#toolbarFilterPopupWindow, QFrame#toolbarFilterPopupWindow {{
        background: transparent;
        border: none;
    }}
    QFrame#toolbarFilterPopupSurface {{
        background: {THEME_SURFACE_ELEVATED};
        border: 1px solid rgba(200, 174, 255, 0.16);
        border-radius: 16px;
    }}
    QListView#comboPopup {{
        background: {THEME_SURFACE_ELEVATED};
        color: {THEME_TEXT_SOFT};
        border: 1px solid rgba(200, 174, 255, 0.16);
        border-radius: 16px;
        padding: 8px;
        outline: 0;
    }}
    QListWidget#toolbarFilterPopupList {{
        background: {THEME_SURFACE_ELEVATED};
        color: {THEME_TEXT_SOFT};
        border: none;
        border-radius: 16px;
        padding: 8px;
        outline: 0;
    }}
    QListView#comboPopup::viewport, QListWidget#toolbarFilterPopupList::viewport {{
        background: {THEME_SURFACE_ELEVATED};
        border-radius: 12px;
    }}
    QListView#comboPopup::item, QListWidget#toolbarFilterPopupList::item {{
        padding: 10px 12px;
        border-radius: 12px;
        min-height: 26px;
    }}
    QListView#comboPopup::item:hover, QListWidget#toolbarFilterPopupList::item:hover {{
        background: rgba(200, 174, 255, 0.10);
    }}
    QListView#comboPopup::item:selected, QListWidget#toolbarFilterPopupList::item:selected {{
        background: rgba(200, 174, 255, 0.18);
        color: {THEME_TEXT};
    }}
    QScrollBar:vertical {{
        background: transparent;
        width: 10px;
        margin: 10px 2px 10px 0px;
    }}
    QScrollBar::handle:vertical {{
        background: rgba(200, 174, 255, 0.30);
        border-radius: 5px;
        min-height: 24px;
    }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical,
    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
        background: transparent;
        height: 0px;
    }}
"""

PRETTY_DIALOG_STYLESHEET = f"""
    QDialog#prettyDialog {{
        background: {THEME_SURFACE_CARD};
        color: {THEME_TEXT_SOFT};
        border: 1px solid {THEME_STROKE};
        border-radius: 24px;
    }}
    QLabel {{
        color: {THEME_TEXT_SOFT};
    }}
    QLabel#dlgTitle {{
        font-size: 16px;
        font-weight: 900;
        color: {THEME_TEXT};
    }}
    QLabel#dlgText {{
        color: {THEME_MUTED};
    }}
    QLabel#dlgIcon {{
        background: rgba(200, 174, 255, 0.12);
        border: 1px solid rgba(200, 174, 255, 0.24);
        border-radius: 12px;
    }}
    QLineEdit, QComboBox, QPlainTextEdit {{
        background: rgba(255, 255, 255, 0.04);
        border: 1px solid {THEME_STROKE};
        border-radius: 14px;
        padding: 0 14px;
        min-height: 44px;
        color: {THEME_TEXT_SOFT};
        selection-background-color: rgba(200, 174, 255, 0.30);
    }}
    QLineEdit:focus, QComboBox:focus, QPlainTextEdit:focus {{
        background: rgba(200, 174, 255, 0.08);
        border: 1px solid rgba(200, 174, 255, 0.28);
    }}
    QComboBox {{
        padding: 0 14px 0 14px;
        padding-right: 44px;
    }}
    QComboBox::drop-down {{
        border: none;
        background: transparent;
        width: 0px;
    }}
    QPlainTextEdit {{
        min-height: 92px;
        padding: 10px 14px;
    }}
    QDialogButtonBox {{
        border-top: 1px solid rgba(200, 174, 255, 0.10);
        padding-top: 12px;
    }}
    QPushButton {{
        border: none;
        border-radius: 14px;
        padding: 0 16px;
        min-height: 44px;
        max-height: 44px;
        font-weight: 750;
        color: {THEME_TEXT_SOFT};
        background: rgba(255, 255, 255, 0.05);
    }}
    QPushButton:hover {{
        background: rgba(200, 174, 255, 0.14);
    }}
    QPushButton:pressed {{
        background: rgba(200, 174, 255, 0.20);
    }}
    QPushButton:disabled {{
        background: rgba(255, 255, 255, 0.03);
        color: {THEME_MUTED_DISABLED};
    }}
    QPushButton[kind="primary"] {{
        background: {THEME_ACCENT};
        color: {THEME_SURFACE_CARD};
    }}
    QPushButton[kind="primary"]:hover {{
        background: {THEME_ACCENT_HOVER};
    }}
    QPushButton[kind="secondary"] {{
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid {THEME_STROKE};
    }}
    QPushButton[kind="secondary"]:hover {{
        background: rgba(200, 174, 255, 0.14);
    }}
    QCheckBox {{
        color: {THEME_TEXT_SOFT};
        spacing: 8px;
    }}
    QCheckBox::indicator {{
        width: 18px;
        height: 18px;
        border-radius: 6px;
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid {THEME_STROKE};
    }}
    QCheckBox::indicator:checked {{
        background: {THEME_ACCENT};
        border: 1px solid {THEME_ACCENT};
    }}
    QScrollArea {{
        background: transparent;
        border: none;
    }}
"""

DIRECTORY_DIALOG_STYLESHEET = PRETTY_DIALOG_STYLESHEET + f"""
    QFileDialog {{
        background: {THEME_SURFACE_CARD};
        color: {THEME_TEXT_SOFT};
    }}
    QFileDialog QLabel {{
        color: {THEME_TEXT_SOFT};
    }}
    QFileDialog QTreeView,
    QFileDialog QListView {{
        background: rgba(255, 255, 255, 0.04);
        color: {THEME_TEXT_SOFT};
        border: 1px solid {THEME_STROKE};
        border-radius: 16px;
        padding: 8px;
        outline: 0;
        selection-background-color: rgba(200, 174, 255, 0.20);
    }}
    QFileDialog QTreeView::item,
    QFileDialog QListView::item {{
        padding: 6px 8px;
        border-radius: 10px;
    }}
    QFileDialog QTreeView::item:hover,
    QFileDialog QListView::item:hover {{
        background: rgba(200, 174, 255, 0.10);
    }}
    QFileDialog QTreeView::item:selected,
    QFileDialog QListView::item:selected {{
        background: rgba(200, 174, 255, 0.18);
        color: {THEME_TEXT};
    }}
    QFileDialog QHeaderView::section {{
        background: rgba(255, 255, 255, 0.04);
        color: {THEME_MUTED};
        border: none;
        padding: 8px 10px;
    }}
"""

def settings_path() -> str:
    base = QStandardPaths.writableLocation(QStandardPaths.AppDataLocation)
    ensure_dir(base)
    return os.path.join(base, SETTINGS_FILE)

def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def fix_dialog_size(dlg: QDialog):
    dlg.setSizeGripEnabled(False)
    dlg.adjustSize()
    dlg.setFixedSize(dlg.sizeHint())


def style_dialog_buttons(button_box: QDialogButtonBox):
    primary_roles = {
        QDialogButtonBox.AcceptRole,
        QDialogButtonBox.YesRole,
        QDialogButtonBox.ApplyRole,
    }
    for button in button_box.buttons():
        role = button_box.buttonRole(button)
        button.setProperty("kind", "primary" if role in primary_roles else "secondary")
        button.setCursor(Qt.PointingHandCursor)


class ThemedComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        popup = QListView(self)
        popup.setObjectName("comboPopup")
        popup.setFrameShape(QFrame.NoFrame)
        popup.setUniformItemSizes(True)
        popup.setSpacing(4)
        popup.setEditTriggers(QAbstractItemView.NoEditTriggers)
        popup.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        popup.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        popup.setTextElideMode(Qt.ElideRight)
        popup.setStyleSheet(COMBO_POPUP_STYLESHEET)
        self.setView(popup)
        self.setCursor(Qt.PointingHandCursor)
        self.setMinimumHeight(44)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setMinimumContentsLength(12)
        self.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
        self.setMaxVisibleItems(10)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.currentIndexChanged.connect(self._sync_current_tooltip)
        self._popup_progress = 0.0
        self._popup_anim = QPropertyAnimation(self, b"popupProgress", self)
        self._popup_anim.setDuration(150)
        self._popup_anim.setEasingCurve(QEasingCurve.OutCubic)
        self._popup_geometry_anim: Optional[QPropertyAnimation] = None
        self._popup_opacity_anim: Optional[QPropertyAnimation] = None
        self._popup_effect: Optional[QGraphicsOpacityEffect] = None
        self._popup_container: Optional[QWidget] = None
        self._sync_current_tooltip()

    def _sync_current_tooltip(self):
        self.setToolTip(self.currentText().strip())

    def getPopupProgress(self) -> float:
        return self._popup_progress

    def setPopupProgress(self, value: float):
        self._popup_progress = max(0.0, min(1.0, float(value)))
        self.update()

    popupProgress = Property(float, getPopupProgress, setPopupProgress)

    def _animate_popup_indicator(self, target: float):
        self._popup_anim.stop()
        self._popup_anim.setStartValue(self._popup_progress)
        self._popup_anim.setEndValue(target)
        self._popup_anim.start()

    def _ensure_popup_animation(self, container: QWidget):
        if self._popup_container is container and self._popup_effect is not None:
            return
        self._popup_container = container
        self._popup_effect = QGraphicsOpacityEffect(container)
        self._popup_effect.setOpacity(1.0)
        container.setGraphicsEffect(self._popup_effect)
        self._popup_geometry_anim = QPropertyAnimation(container, b"geometry", container)
        self._popup_geometry_anim.setDuration(150)
        self._popup_geometry_anim.setEasingCurve(QEasingCurve.OutCubic)
        self._popup_opacity_anim = QPropertyAnimation(self._popup_effect, b"opacity", container)
        self._popup_opacity_anim.setDuration(150)
        self._popup_opacity_anim.setEasingCurve(QEasingCurve.OutCubic)

    def showPopup(self):
        popup_width = max(0, self.width())
        self.view().setMinimumWidth(popup_width)
        self.view().setMaximumWidth(popup_width)
        super().showPopup()
        container = self.view().window()
        container.setObjectName("comboPopupWindow")
        container.setAttribute(Qt.WA_TranslucentBackground, True)
        container.setContentsMargins(0, 0, 0, 0)
        container.setStyleSheet(COMBO_POPUP_STYLESHEET)
        container.setMinimumWidth(popup_width)
        container.setMaximumWidth(popup_width)
        container.resize(popup_width, container.height())
        self._ensure_popup_animation(container)
        end_rect = container.geometry()
        start_rect = QRect(end_rect.x(), end_rect.y() - 8, end_rect.width(), end_rect.height())
        if self._popup_effect and self._popup_geometry_anim and self._popup_opacity_anim:
            self._popup_geometry_anim.stop()
            self._popup_opacity_anim.stop()
            self._popup_effect.setOpacity(0.0)
            container.setGeometry(start_rect)
            self._popup_geometry_anim.setStartValue(start_rect)
            self._popup_geometry_anim.setEndValue(end_rect)
            self._popup_opacity_anim.setStartValue(0.0)
            self._popup_opacity_anim.setEndValue(1.0)
            self._popup_geometry_anim.start()
            self._popup_opacity_anim.start()
        self._animate_popup_indicator(1.0)

    def hidePopup(self):
        self._animate_popup_indicator(0.0)
        super().hidePopup()

    def paintEvent(self, event):
        painter = QStylePainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        option = QStyleOptionComboBox()
        self.initStyleOption(option)
        option.subControls = QStyle.SC_ComboBoxFrame | QStyle.SC_ComboBoxEditField
        painter.drawComplexControl(QStyle.CC_ComboBox, option)
        painter.drawControl(QStyle.CE_ComboBoxLabel, option)
        color = QColor("#F3EEFF" if self.isEnabled() else "#6E6680")
        pen = QPen(color, 1.8, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        painter.setPen(pen)
        arrow_rect = QRect(max(0, self.width() - 24), max(0, (self.height() - 12) // 2), 12, 12)
        painter.save()
        painter.translate(arrow_rect.center().x(), arrow_rect.center().y())
        painter.rotate(180.0 * self._popup_progress)
        painter.drawLine(QPoint(-4, -2), QPoint(0, 2))
        painter.drawLine(QPoint(0, 2), QPoint(4, -2))
        painter.restore()



class ToolbarFilterPopup(QFrame):
    indexSelected = Signal(int)
    dismissed = Signal()

    def __init__(self, owner=None):
        super().__init__(None, Qt.Popup | Qt.FramelessWindowHint | Qt.NoDropShadowWindowHint)
        self._owner = owner
        self.setObjectName("toolbarFilterPopupWindow")
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setStyleSheet(COMBO_POPUP_STYLESHEET)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self.surface = QFrame()
        self.surface.setObjectName("toolbarFilterPopupSurface")
        outer.addWidget(self.surface)

        surface_l = QVBoxLayout(self.surface)
        surface_l.setContentsMargins(0, 0, 0, 0)
        surface_l.setSpacing(0)

        self.list = QListWidget()
        self.list.setObjectName("toolbarFilterPopupList")
        self.list.setFrameShape(QFrame.NoFrame)
        self.list.setUniformItemSizes(True)
        self.list.setSpacing(4)
        self.list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.list.setSelectionMode(QAbstractItemView.SingleSelection)
        self.list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.list.setTextElideMode(Qt.ElideRight)
        self.list.setStyleSheet(COMBO_POPUP_STYLESHEET)
        self.list.itemClicked.connect(self._handle_item_triggered)
        self.list.itemActivated.connect(self._handle_item_triggered)
        surface_l.addWidget(self.list)

    def sync_items(self, items: List[Dict[str, Any]], current_index: int):
        self.list.clear()
        for index, entry in enumerate(items):
            item = QListWidgetItem(str(entry["text"]))
            item.setData(Qt.UserRole, index)
            tooltip = entry["roles"].get(Qt.ToolTipRole, entry["text"])
            if tooltip:
                item.setToolTip(str(tooltip))
            item.setSizeHint(QSize(0, 40))
            self.list.addItem(item)
        if 0 <= current_index < self.list.count():
            self.list.setCurrentRow(current_index)
            current_item = self.list.item(current_index)
            if current_item is not None:
                self.list.scrollToItem(current_item, QAbstractItemView.PositionAtCenter)

    def show_for(self, anchor: QWidget, width: int, item_count: int, max_visible_items: int = 10):
        row_height = self.list.sizeHintForRow(0) if item_count else 40
        row_height = max(40, row_height)
        visible_rows = max(1, min(max_visible_items, max(1, item_count)))
        height = visible_rows * row_height + 18

        anchor_center = anchor.mapToGlobal(anchor.rect().center())
        screen = QApplication.screenAt(anchor_center) or anchor.screen() or QApplication.primaryScreen()
        available = screen.availableGeometry() if screen else QRect(0, 0, width, height)

        x = anchor.mapToGlobal(QPoint(0, 0)).x()
        max_x = max(available.left() + 8, available.x() + available.width() - width - 8)
        x = min(max(available.left() + 8, x), max_x)

        below_y = anchor.mapToGlobal(QPoint(0, anchor.height() + 6)).y()
        above_y = anchor.mapToGlobal(QPoint(0, -height - 6)).y()
        min_y = available.top() + 8
        max_y = available.y() + available.height() - height - 8
        if below_y + height <= available.y() + available.height() - 8 or above_y < min_y:
            y = below_y
        else:
            y = above_y
        y = min(max(min_y, y), max(min_y, max_y))

        self.setGeometry(x, y, width, height)
        self.show()
        self.raise_()
        self.list.setFocus(Qt.PopupFocusReason)

    def hideEvent(self, event):
        super().hideEvent(event)
        self.dismissed.emit()

    def _handle_item_triggered(self, item: QListWidgetItem):
        index = item.data(Qt.UserRole)
        if index is None:
            return
        self.indexSelected.emit(int(index))
        self.hide()


class ToolbarFilterCombo(QWidget):
    currentIndexChanged = Signal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._items: List[Dict[str, Any]] = []
        self._current_index = -1
        self._hovered = False
        self._pressed = False
        self._popup_progress = 0.0
        self._popup = ToolbarFilterPopup(self)
        self.destroyed.connect(self._popup.deleteLater)
        self._popup.indexSelected.connect(self._on_popup_index_selected)
        self._popup.dismissed.connect(self._on_popup_dismissed)
        self._popup_anim = QPropertyAnimation(self, b"popupProgress", self)
        self._popup_anim.setDuration(140)
        self._popup_anim.setEasingCurve(QEasingCurve.OutCubic)
        self.setCursor(Qt.PointingHandCursor)
        self.setFocusPolicy(Qt.StrongFocus)
        self.setMouseTracking(True)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setMinimumSize(148, 44)
        self.setMaximumHeight(44)

    def getPopupProgress(self) -> float:
        return self._popup_progress

    def setPopupProgress(self, value: float):
        self._popup_progress = max(0.0, min(1.0, float(value)))
        self.update()

    popupProgress = Property(float, getPopupProgress, setPopupProgress)

    def count(self) -> int:
        return len(self._items)

    def clear(self):
        self.hidePopup()
        self._items.clear()
        self._current_index = -1
        self.setToolTip("")
        self.update()

    def addItem(self, text: str, userData: Any = None):
        entry = {
            "text": str(text),
            "roles": {
                Qt.DisplayRole: str(text),
                Qt.ToolTipRole: str(text),
                Qt.UserRole: userData,
            },
        }
        self._items.append(entry)
        if self._current_index < 0:
            self.setCurrentIndex(0)
        else:
            self.update()

    def itemText(self, index: int) -> str:
        if 0 <= index < len(self._items):
            return str(self._items[index]["text"])
        return ""

    def itemData(self, index: int, role: int = Qt.UserRole):
        if not (0 <= index < len(self._items)):
            return None
        return self._items[index]["roles"].get(role)

    def setItemData(self, index: int, value: Any, role: int = Qt.UserRole):
        if not (0 <= index < len(self._items)):
            return
        self._items[index]["roles"][role] = value
        if role in (Qt.DisplayRole, Qt.EditRole):
            self._items[index]["text"] = str(value)
        if self._popup.isVisible():
            self._popup.sync_items(self._items, self._current_index)
        self._sync_tooltip()
        self.update()

    def currentIndex(self) -> int:
        return self._current_index

    def currentText(self) -> str:
        return self.itemText(self._current_index)

    def currentData(self, role: int = Qt.UserRole):
        return self.itemData(self._current_index, role)

    def setCurrentIndex(self, index: int):
        if not self._items:
            index = -1
        elif index < 0:
            index = -1
        else:
            index = min(index, len(self._items) - 1)

        if index == self._current_index:
            self._sync_tooltip()
            self.update()
            return

        self._current_index = index
        if self._popup.isVisible():
            self._popup.sync_items(self._items, self._current_index)
        self._sync_tooltip()
        self.update()
        if not self.signalsBlocked():
            self.currentIndexChanged.emit(index)

    def minimumSizeHint(self) -> QSize:
        return QSize(148, 44)

    def sizeHint(self) -> QSize:
        return QSize(max(148, self.minimumWidth()), 44)

    def showPopup(self):
        if not self.isEnabled() or not self._items:
            return
        self._popup.sync_items(self._items, self._current_index)
        self._popup.show_for(self, self.width(), len(self._items))
        self._animate_popup_indicator(1.0)
        self.update()

    def hidePopup(self):
        if self._popup.isVisible():
            self._popup.hide()
        else:
            self._on_popup_dismissed()

    def _animate_popup_indicator(self, target: float):
        self._popup_anim.stop()
        self._popup_anim.setStartValue(self._popup_progress)
        self._popup_anim.setEndValue(target)
        self._popup_anim.start()

    def _sync_tooltip(self):
        self.setToolTip(self.currentText().strip())

    def _on_popup_index_selected(self, index: int):
        self.setCurrentIndex(index)

    def _on_popup_dismissed(self):
        self._pressed = False
        self._animate_popup_indicator(0.0)
        self.update()

    def enterEvent(self, event):
        self._hovered = True
        self.update()
        super().enterEvent(event)

    def leaveEvent(self, event):
        self._hovered = False
        self._pressed = False
        self.update()
        super().leaveEvent(event)

    def hideEvent(self, event):
        self.hidePopup()
        super().hideEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self.isEnabled():
            self._pressed = True
            self.setFocus(Qt.MouseFocusReason)
            self.update()
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton and self.isEnabled():
            was_pressed = self._pressed
            self._pressed = False
            self.update()
            if was_pressed and self.rect().contains(event.position().toPoint()):
                if self._popup.isVisible():
                    self.hidePopup()
                else:
                    self.showPopup()
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter, Qt.Key_Space, Qt.Key_Down):
            self.showPopup()
            event.accept()
            return
        if event.key() == Qt.Key_Escape and self._popup.isVisible():
            self.hidePopup()
            event.accept()
            return
        super().keyPressEvent(event)

    def wheelEvent(self, event):
        event.ignore()

    def focusInEvent(self, event):
        self.update()
        super().focusInEvent(event)

    def focusOutEvent(self, event):
        self.update()
        super().focusOutEvent(event)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        rect = self.rect().adjusted(0, 0, -1, -1)
        radius = 14.0

        if not self.isEnabled():
            bg = QColor(255, 255, 255, 8)
            border = QColor(200, 174, 255, 18)
            text_color = QColor(THEME_MUTED_DISABLED)
        elif self._popup.isVisible() or self.hasFocus():
            bg = QColor(200, 174, 255, 20)
            border = QColor(200, 174, 255, 62)
            text_color = QColor(THEME_TEXT)
        elif self._hovered or self._pressed:
            bg = QColor(255, 255, 255, 18)
            border = QColor(200, 174, 255, 34)
            text_color = QColor(THEME_TEXT_SOFT)
        else:
            bg = QColor(255, 255, 255, 13)
            border = QColor(200, 174, 255, 26)
            text_color = QColor(THEME_TEXT_SOFT)

        painter.setPen(QPen(border, 1.0))
        painter.setBrush(bg)
        painter.drawRoundedRect(rect, radius, radius)

        text_rect = rect.adjusted(14, 0, -36, 0)
        text = painter.fontMetrics().elidedText(self.currentText(), Qt.ElideRight, max(0, text_rect.width()))
        painter.setPen(text_color)
        painter.drawText(text_rect, Qt.AlignVCenter | Qt.AlignLeft, text)

        arrow_color = QColor(THEME_TEXT_SOFT if self.isEnabled() else THEME_MUTED_DISABLED)
        painter.setPen(QPen(arrow_color, 1.8, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
        arrow_center = QPoint(rect.right() - 18, rect.center().y())
        painter.save()
        painter.translate(arrow_center)
        painter.rotate(180.0 * self._popup_progress)
        painter.drawLine(QPoint(-4, -2), QPoint(0, 2))
        painter.drawLine(QPoint(0, 2), QPoint(4, -2))
        painter.restore()


class AnimatedToolButton(QToolButton):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._hover_progress = 0.0
        self._hover_anim = QPropertyAnimation(self, b"hoverProgress", self)
        self._hover_anim.setDuration(140)
        self._hover_anim.setEasingCurve(QEasingCurve.OutCubic)
        self.setCursor(Qt.PointingHandCursor)

    def getHoverProgress(self) -> float:
        return self._hover_progress

    def setHoverProgress(self, value: float):
        self._hover_progress = max(0.0, min(1.0, float(value)))
        self.update()

    hoverProgress = Property(float, getHoverProgress, setHoverProgress)

    def _animate_hover(self, target: float):
        self._hover_anim.stop()
        self._hover_anim.setStartValue(self._hover_progress)
        self._hover_anim.setEndValue(target)
        self._hover_anim.start()

    def enterEvent(self, event):
        self._animate_hover(1.0)
        super().enterEvent(event)

    def leaveEvent(self, event):
        self._animate_hover(0.0)
        super().leaveEvent(event)

    def paintEvent(self, event):
        super().paintEvent(event)
        if self._hover_progress <= 0.0 or not self.isEnabled():
            return

        kind = str(self.property("kind") or "").strip().lower()
        if kind == "danger":
            glow_base = QColor(255, 118, 150)
            border_base = QColor(255, 146, 170)
        elif kind in {"primary", "solid"}:
            glow_base = QColor(200, 174, 255)
            border_base = QColor(222, 204, 255)
        else:
            glow_base = QColor(200, 174, 255)
            border_base = QColor(214, 196, 255)

        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        rect = self.rect().adjusted(1, 1, -2, -2)
        radius = min(14.0, rect.height() / 2)
        glow = QColor(glow_base)
        glow.setAlpha(int(24 * self._hover_progress))
        border = QColor(border_base)
        border.setAlpha(int(82 * self._hover_progress))
        painter.setBrush(glow)
        painter.setPen(QPen(border, 1))
        painter.drawRoundedRect(rect, radius, radius)

class DashboardBarChart(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._series: List[Tuple[str, int, str]] = []
        self.setMinimumHeight(176)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def set_series(self, series: List[Tuple[str, int, str]]):
        self._series = list(series)
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.fillRect(self.rect(), Qt.transparent)

        rect = self.rect().adjusted(18, 18, -18, -18)
        if not self._series:
            painter.setPen(QColor(THEME_MUTED))
            painter.drawText(rect, Qt.AlignCenter, "Нет данных")
            return

        max_value = max(1, max(value for _, value, _ in self._series))
        label_height = 40
        value_height = 24
        top_pad = 12
        chart_height = max(80, rect.height() - label_height - value_height - top_pad)
        count = len(self._series)
        gap = 14
        slot_width = max(52, int((rect.width() - gap * max(0, count - 1)) / max(count, 1)))
        bar_width = max(26, min(58, slot_width - 16))

        chart_top = rect.top() + value_height + top_pad
        chart_bottom = chart_top + chart_height
        baseline = chart_bottom

        guide_pen = QPen(QColor(255, 255, 255, 20), 1)
        for step in range(5):
            guide_y = chart_top + int((chart_height / 4) * step)
            painter.setPen(guide_pen)
            painter.drawLine(rect.left(), guide_y, rect.right(), guide_y)

        for index, (label, value, color_hex) in enumerate(self._series):
            slot_x = rect.left() + index * (slot_width + gap)
            bar_height = 0 if max_value == 0 else int((value / max_value) * max(24, chart_height - 10))
            bar_x = slot_x + max(0, (slot_width - bar_width) // 2)
            bar_y = baseline - bar_height

            painter.setPen(Qt.NoPen)
            painter.setBrush(QColor(255, 255, 255, 16))
            painter.drawRoundedRect(bar_x, chart_top, bar_width, chart_height, 14, 14)

            if value > 0:
                gradient = QLinearGradient(bar_x, bar_y, bar_x, baseline)
                base = QColor(color_hex)
                gradient.setColorAt(0.0, base.lighter(130))
                gradient.setColorAt(1.0, base)
                painter.setBrush(gradient)
                painter.drawRoundedRect(bar_x, bar_y, bar_width, max(12, bar_height), 14, 14)

            painter.setPen(QColor(THEME_TEXT_SOFT))
            painter.drawText(slot_x, rect.top(), slot_width, value_height, Qt.AlignCenter, str(value))

            painter.setPen(QColor(THEME_MUTED))
            painter.drawText(
                slot_x - 2,
                baseline + 10,
                slot_width + 4,
                label_height,
                Qt.AlignHCenter | Qt.AlignTop | Qt.TextWordWrap,
                label,
            )


def fmt_date_ddmmyyyy(iso_str: str) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%d.%m.%Y")
    except Exception:
        try:
            y, m, d = iso_str.split("T")[0].split("-")
            return f"{d}.{m}.{y}"
        except Exception:
            return iso_str

def today_ymd() -> str:
    return datetime.now().date().isoformat()

def parse_date_to_ymd(s: str) -> Optional[str]:
    s = (s or "").strip()
    if not s:
        return None

    try:
        if "." in s:
            d, m, y = s.split(".")
            dt = datetime(int(y), int(m), int(d)).date()
            return dt.isoformat()
    except Exception:
        pass

    try:
        if "-" in s and len(s) >= 10:
            dt = datetime.fromisoformat(s[:10]).date()
            return dt.isoformat()
    except Exception:
        pass

    return None

def compose_badged_icon(base: QIcon, badge: QIcon, size: int = 16) -> QIcon:
    pm = base.pixmap(size, size)
    out = QPixmap(pm.size())
    out.fill(Qt.transparent)
    p = QPainter(out)
    p.drawPixmap(0, 0, pm)

    badge_pm = badge.pixmap(int(size * 0.75), int(size * 0.75))
    x = size - badge_pm.width() + 1
    y = size - badge_pm.height() + 1
    p.drawPixmap(x, y, badge_pm)
    p.end()
    return QIcon(out)

def atomic_copy(src: str, dst: str):
    ensure_dir(os.path.dirname(dst))
    tmp = dst + f".tmp_{os.getpid()}_{int(time.time() * 1000)}"
    shutil.copy2(src, tmp)
    os.replace(tmp, dst)

def parse_ver(v: str) -> Tuple[int, int, int]:
    try:
        a, b, c = v.strip().split(".")
        return int(a), int(b), int(c)
    except Exception:
        return (0, 0, 0)

def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def local_app_dir(app_name: str = "PDDocsRegistry") -> str:
    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    p = os.path.join(base, app_name)
    ensure_dir(p)
    return p

def updater_exe_path() -> str:
    if getattr(sys, "frozen", False):
        return os.path.join(os.path.dirname(sys.executable), "updater.exe")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "updater.py")

class FileLock:
    def __init__(self, lock_path: str, owner: str, stale_seconds: int = 120):
        self.lock_path = lock_path
        self.owner = owner
        self.stale_seconds = stale_seconds
        self._acquired = False

    def acquire(self, timeout_seconds: float = 10.0) -> bool:
        start = time.time()
        while True:
            try:
                if os.path.exists(self.lock_path):
                    age = time.time() - os.path.getmtime(self.lock_path)
                    if age > self.stale_seconds:
                        try:
                            os.remove(self.lock_path)
                        except Exception:
                            pass

                fd = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                payload = {"owner": self.owner, "created_at": now_iso(), "pid": os.getpid()}
                os.write(fd, json.dumps(payload, ensure_ascii=False).encode("utf-8"))
                os.close(fd)
                self._acquired = True
                return True
            except FileExistsError:
                if time.time() - start >= timeout_seconds:
                    return False
                time.sleep(0.12)
            except Exception:
                if time.time() - start >= timeout_seconds:
                    return False
                time.sleep(0.12)

    def release(self):
        if not self._acquired:
            return
        try:
            if os.path.exists(self.lock_path):
                os.remove(self.lock_path)
        except Exception:
            pass
        self._acquired = False

    def __enter__(self):
        ok = self.acquire()
        if not ok:
            raise TimeoutError("Не удалось получить блокировку.")
        return self

    def __exit__(self, exc_type, exc, tb):
        self.release()


class BackgroundTaskSignals(QObject):
    finished = Signal(object)
    failed = Signal(str)


class BackgroundTask(QRunnable):
    def __init__(self, fn: Callable[[], Any]):
        super().__init__()
        self._fn = fn
        self.signals = BackgroundTaskSignals()

    def run(self):
        try:
            result = self._fn()
        except Exception as e:
            self.signals.failed.emit(str(e))
        else:
            self.signals.finished.emit(result)


class PrettyDialog(QDialog):
    def __init__(self, parent, title: str, icon: QIcon):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setObjectName("prettyDialog")
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setStyleSheet(PRETTY_DIALOG_STYLESHEET)

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 16, 18, 14)
        root.setSpacing(12)

        header = QHBoxLayout()
        header.setSpacing(12)

        self.icon_label = QLabel()
        self.icon_label.setObjectName("dlgIcon")
        self.icon_label.setFixedSize(34, 34)
        self.icon_label.setAlignment(Qt.AlignCenter)
        self.icon_label.setPixmap(icon.pixmap(22, 22))

        self.title_label = QLabel(title)
        self.title_label.setObjectName("dlgTitle")
        self.title_label.setWordWrap(True)

        header.addWidget(self.icon_label, 0, Qt.AlignTop)
        header.addWidget(self.title_label, 1, Qt.AlignVCenter)

        root.addLayout(header)
        self._content_holder = QVBoxLayout()
        self._content_holder.setSpacing(10)
        root.addLayout(self._content_holder)

    def content_layout(self) -> QVBoxLayout:
        return self._content_holder


class ThemedMessageDialog(PrettyDialog):
    def __init__(self, parent, title: str, message: str, kind: str = "info", buttons=("ok",)):
        if kind == "info":
            ico = parent.style().standardIcon(QStyle.SP_MessageBoxInformation)
        elif kind == "warn":
            ico = parent.style().standardIcon(QStyle.SP_MessageBoxWarning)
        else:
            ico = parent.style().standardIcon(QStyle.SP_MessageBoxQuestion)

        super().__init__(parent, title, ico)
        self._result_yes = False

        msg = QLabel(message)
        msg.setObjectName("dlgText")
        msg.setWordWrap(True)
        msg.setTextFormat(Qt.PlainText)
        msg.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.content_layout().addWidget(msg)

        bb = QDialogButtonBox()
        if "yes" in buttons:
            bb.addButton("Да", QDialogButtonBox.AcceptRole)
        if "no" in buttons:
            bb.addButton("Нет", QDialogButtonBox.RejectRole)
        if buttons == ("ok",):
            bb.addButton("Ок", QDialogButtonBox.AcceptRole)

        style_dialog_buttons(bb)
        bb.accepted.connect(self._accept_yes)
        bb.rejected.connect(self.reject)
        self.content_layout().addWidget(bb)

        fix_dialog_size(self)

    def _accept_yes(self):
        self._result_yes = True
        self.accept()

    def yes(self) -> bool:
        return self._result_yes


class WorkspaceSetupDialog(PrettyDialog):
    def __init__(self, parent=None):
        icon = parent.style().standardIcon(QStyle.SP_DirOpenIcon) if parent is not None else QIcon()
        super().__init__(parent, "Рабочая папка", icon)

        intro = QLabel(
            "Выберите папку, в которой приложение будет хранить базу документов и вложения."
        )
        intro.setObjectName("dlgText")
        intro.setWordWrap(True)
        self.content_layout().addWidget(intro)

        details = QLabel(
            "В этой папке автоматически появятся:\n"
            "- `pd_docs.db` с данными\n"
            "- `registry.json` с настройкой рабочего пространства\n"
            "- скрытая папка `.pd_docs` с загруженными файлами"
        )
        details.setObjectName("dlgText")
        details.setTextFormat(Qt.MarkdownText)
        details.setWordWrap(True)
        self.content_layout().addWidget(details)

        note = QLabel(
            "Подойдет локальная папка или общая сетевая папка, если у пользователя есть права на запись."
        )
        note.setObjectName("dlgText")
        note.setWordWrap(True)
        self.content_layout().addWidget(note)

        bb = QDialogButtonBox()
        choose_btn = bb.addButton("Выбрать папку", QDialogButtonBox.AcceptRole)
        cancel_btn = bb.addButton("Выход", QDialogButtonBox.RejectRole)
        choose_btn.setDefault(True)
        style_dialog_buttons(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        self.content_layout().addWidget(bb)

        self.setMinimumWidth(620)
        fix_dialog_size(self)


class AddDocDialog(PrettyDialog):
    def __init__(self, parent=None):
        super().__init__(parent, "Добавить документ", parent.style().standardIcon(QStyle.SP_FileDialogNewFolder))

        self.title_edit = QLineEdit()
        self.title_edit.setPlaceholderText("Например: Политика обработки ПДн v1")
        self.title_edit.setMinimumWidth(460)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignLeft)
        form.setFormAlignment(Qt.AlignTop)
        form.addRow("Название:", self.title_edit)
        self.content_layout().addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("Добавить")
        buttons.button(QDialogButtonBox.Cancel).setText("Отмена")
        style_dialog_buttons(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        self.content_layout().addWidget(buttons)

        fix_dialog_size(self)

    def get_value(self):
        return self.title_edit.text().strip()


class InputDialog(PrettyDialog):
    def __init__(self, parent, title: str, label: str, default: str = ""):
        super().__init__(parent, title, parent.style().standardIcon(QStyle.SP_FileDialogContentsView))

        self.edit = QLineEdit()
        self.edit.setText(default)
        self.edit.setPlaceholderText(label)
        self.edit.setMinimumWidth(460)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignLeft)
        form.setFormAlignment(Qt.AlignTop)
        form.addRow(label, self.edit)
        self.content_layout().addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("Ок")
        buttons.button(QDialogButtonBox.Cancel).setText("Отмена")
        style_dialog_buttons(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        self.content_layout().addWidget(buttons)

        fix_dialog_size(self)

    def value(self) -> str:
        return self.edit.text().strip()


class SopdDialog(PrettyDialog):
    _TRANSFER_VALUES = ["Не указано", "Да", "Нет"]

    def __init__(self, parent, title: str = "Карточка СОПД", data: Optional[Dict[str, str]] = None):
        super().__init__(parent, title, parent.style().standardIcon(QStyle.SP_FileDialogDetailedView))

        self.consent_type_edit = QLineEdit()
        self.consent_type_edit.setPlaceholderText("Например: Согласие клиента на маркетинговые коммуникации")

        self.purpose_edit = QPlainTextEdit()
        self.purpose_edit.setPlaceholderText("Цель обработки")
        self.purpose_edit.setFixedHeight(78)

        self.categories_edit = QLineEdit()
        self.categories_edit.setPlaceholderText("Общие, специальные, биометрические ...")

        self.pd_list_edit = QPlainTextEdit()
        self.pd_list_edit.setPlaceholderText("Перечень персональных данных")
        self.pd_list_edit.setFixedHeight(86)

        self.processing_method_edit = QLineEdit()
        self.processing_method_edit.setPlaceholderText("Автоматизированная / неавтоматизированная / смешанная")

        self.third_party_combo = ThemedComboBox()
        self.third_party_combo.addItems(self._TRANSFER_VALUES)

        self.transfer_to_edit = QLineEdit()
        self.transfer_to_edit.setPlaceholderText("Укажи получателей (компании/категории получателей)")

        self.sopd_description_edit = QPlainTextEdit()
        self.sopd_description_edit.setPlaceholderText("Дополнительное описание СОПД")
        self.sopd_description_edit.setFixedHeight(82)

        self.validity_edit = QLineEdit()
        self.validity_edit.setPlaceholderText("До отзыва / 3 года / до достижения цели ...")

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignLeft)
        form.setFormAlignment(Qt.AlignTop)
        form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        form.addRow("Вид согласия:", self.consent_type_edit)
        form.addRow("Цель:", self.purpose_edit)
        form.addRow("Категории ПД:", self.categories_edit)
        form.addRow("Перечень ПД:", self.pd_list_edit)
        form.addRow("Способ обработки:", self.processing_method_edit)
        form.addRow("Передача 3-им лицам:", self.third_party_combo)
        form.addRow("Кому передаётся:", self.transfer_to_edit)
        form.addRow("Описание СОПД:", self.sopd_description_edit)
        form.addRow("Срок действия:", self.validity_edit)
        self.content_layout().addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("Сохранить")
        buttons.button(QDialogButtonBox.Cancel).setText("Отмена")
        style_dialog_buttons(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        self.content_layout().addWidget(buttons)

        if data:
            self.set_data(data)

        self.third_party_combo.currentTextChanged.connect(self._update_transfer_to_state)
        self._update_transfer_to_state(self.third_party_combo.currentText())

        self.setMinimumWidth(720)
        fix_dialog_size(self)

    def _update_transfer_to_state(self, transfer_value: str):
        need_transfer_to = ((transfer_value or "").strip().lower() == "да")
        self.transfer_to_edit.setEnabled(need_transfer_to)
        if need_transfer_to:
            self.transfer_to_edit.setPlaceholderText("Укажи получателей (обязательно при передаче)")
        else:
            self.transfer_to_edit.setPlaceholderText("Не требуется, если передачи нет")
            if not self.transfer_to_edit.text().strip():
                self.transfer_to_edit.clear()

    def set_data(self, data: Dict[str, str]):
        self.consent_type_edit.setText((data.get("consent_type") or "").strip())
        self.purpose_edit.setPlainText((data.get("purpose") or "").strip())
        self.categories_edit.setText((data.get("pd_categories") or "").strip())
        self.pd_list_edit.setPlainText((data.get("pd_list") or "").strip())
        self.processing_method_edit.setText((data.get("processing_method") or "").strip())
        transfer = (data.get("third_party_transfer") or "").strip()
        idx = self.third_party_combo.findText(transfer) if transfer else -1
        self.third_party_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.transfer_to_edit.setText((data.get("transfer_to") or "").strip())
        self.sopd_description_edit.setPlainText((data.get("sopd_description") or "").strip())
        self.validity_edit.setText((data.get("validity_period") or "").strip())
        self._update_transfer_to_state(self.third_party_combo.currentText())

    def values(self) -> Dict[str, str]:
        return {
            "consent_type": self.consent_type_edit.text().strip(),
            "purpose": self.purpose_edit.toPlainText().strip(),
            "pd_categories": self.categories_edit.text().strip(),
            "pd_list": self.pd_list_edit.toPlainText().strip(),
            "processing_method": self.processing_method_edit.text().strip(),
            "third_party_transfer": self.third_party_combo.currentText().strip(),
            "transfer_to": self.transfer_to_edit.text().strip(),
            "sopd_description": self.sopd_description_edit.toPlainText().strip(),
            "validity_period": self.validity_edit.text().strip(),
        }


class Db:
    def __init__(self, db_path: str, current_user: str):
        self.db_path = db_path
        self.user = current_user
        ensure_dir(os.path.dirname(db_path))

        self.conn = sqlite3.connect(self.db_path, timeout=5.0, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._apply_pragmas()
        self._init_schema()
        self._migrate()

    def close(self):
        try:
            self.conn.close()
        except Exception:
            pass

    def _apply_pragmas(self):
        cur = self.conn.cursor()
        cur.execute("PRAGMA foreign_keys = ON;")
        journal_mode = "DELETE"
        try:
            row = cur.execute("PRAGMA journal_mode = WAL;").fetchone()
            if row:
                journal_mode = str(row[0] or journal_mode).upper()
        except sqlite3.DatabaseError:
            journal_mode = "DELETE"
        cur.execute("PRAGMA synchronous = NORMAL;")
        cur.execute("PRAGMA busy_timeout = 8000;")
        cur.execute("PRAGMA temp_store = MEMORY;")
        cur.execute("PRAGMA cache_size = -20000;")
        if journal_mode == "WAL":
            cur.execute("PRAGMA wal_autocheckpoint = 1000;")

        self.conn.commit()

    def rename_document(self, doc_id: int, new_title: str):
        new_title = new_title.strip()
        if not new_title:
            raise ValueError("Название документа не может быть пустым.")

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            row = cur.execute("SELECT doc_title FROM documents WHERE id=?;", (doc_id,)).fetchone()
            cur.execute("""
                UPDATE documents
                SET doc_title=?, updated_at=?, updated_by=?
                WHERE id=?;
            """, (new_title, now_iso(), self.user, doc_id))
            old_title = str(row["doc_title"] or "").strip() if row else ""
            if old_title != new_title:
                self._append_document_history(
                    cur,
                    doc_id,
                    "rename",
                    f"Переименован документ: «{old_title or 'без названия'}» -> «{new_title}».",
                )
            self.conn.commit()

        self._retry_write(work)

    def set_accept_date(self, doc_id: int, accept_date: Optional[str]):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            row = cur.execute("SELECT accept_date FROM documents WHERE id=?;", (doc_id,)).fetchone()
            cur.execute(
                "UPDATE documents SET accept_date=?, updated_at=?, updated_by=? WHERE id=?;",
                (accept_date, now_iso(), self.user, doc_id),
            )
            old_value = (row["accept_date"] or "").strip() if row else ""
            new_value = (accept_date or "").strip()
            if old_value != new_value:
                if new_value:
                    text = f"Изменена дата утверждения: {fmt_date_ddmmyyyy(old_value) or old_value or 'не была указана'} -> {fmt_date_ddmmyyyy(new_value) or new_value}."
                else:
                    text = "Дата утверждения очищена."
                self._append_document_history(cur, doc_id, "accept_date", text)
            self.conn.commit()
        self._retry_write(work)

    def _normalize_doc_status(self, status: Optional[str]) -> str:
        st = (status or "").strip()
        if st in DOC_STATUSES:
            return st
        return DOC_STATUS_DEFAULT

    def set_document_status(self, doc_id: int, status: str):
        status = self._normalize_doc_status(status)

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            row = cur.execute("SELECT status FROM documents WHERE id=?;", (doc_id,)).fetchone()
            cur.execute(
                "UPDATE documents SET status=?, updated_at=?, updated_by=? WHERE id=?;",
                (status, now_iso(), self.user, doc_id),
            )
            old_status = self._normalize_doc_status(row["status"] if row else None)
            if old_status != status:
                self._append_document_history(
                    cur,
                    doc_id,
                    "status",
                    f"Изменён статус: {old_status} -> {status}.",
                )
            self.conn.commit()

        self._retry_write(work)

    def _retry_write(self, fn, attempts: int = 8):
        delay = 0.05
        for _ in range(attempts):
            try:
                return fn()
            except sqlite3.OperationalError as e:
                msg = str(e).lower()
                if "locked" in msg or "busy" in msg:
                    time.sleep(min(delay, 0.8))
                    delay *= 2
                    continue
                raise
        raise sqlite3.OperationalError("База занята слишком долго (database is locked). Попробуй ещё раз.")

    def _add_column_if_missing(self, table: str, col: str, ddl: str):
        cur = self.conn.cursor()
        cur.execute(f"PRAGMA table_info({table});")
        cols = [r["name"] for r in cur.fetchall()]
        if col not in cols:
            cur.execute(f"ALTER TABLE {table} ADD COLUMN {col} {ddl};")
            self.conn.commit()

    def _init_schema(self):
        cur = self.conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL,
            created_by TEXT
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            doc_title TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Черновик',
            pdf_path TEXT,
            office_path TEXT,
            comment TEXT,
            needs_office INTEGER NOT NULL DEFAULT 0,
            review_due TEXT,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            created_by TEXT,
            updated_at TEXT NOT NULL,
            updated_by TEXT,
            FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS sections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            created_by TEXT,
            FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE,
            UNIQUE(company_id, name)
        );
        """)

        cur.execute("""
        CREATE TABLE IF NOT EXISTS document_sections (
            doc_id INTEGER NOT NULL,
            section_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            created_by TEXT,
            PRIMARY KEY (doc_id, section_id),
            FOREIGN KEY(doc_id) REFERENCES documents(id) ON DELETE CASCADE,
            FOREIGN KEY(section_id) REFERENCES sections(id) ON DELETE CASCADE
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS document_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id INTEGER NOT NULL,
            event_type TEXT NOT NULL,
            event_text TEXT NOT NULL,
            created_at TEXT NOT NULL,
            created_by TEXT,
            FOREIGN KEY(doc_id) REFERENCES documents(id) ON DELETE CASCADE
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS sopd_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            consent_type TEXT NOT NULL,
            purpose TEXT NOT NULL,
            legal_basis TEXT,
            pd_categories TEXT,
            data_subjects TEXT,
            pd_list TEXT,
            processing_operations TEXT,
            processing_method TEXT,
            third_party_transfer TEXT,
            transfer_to TEXT,
            sopd_description TEXT,
            attachment_path TEXT,
            validity_period TEXT,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            created_by TEXT,
            updated_at TEXT NOT NULL,
            updated_by TEXT,
            FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
        );
        """)
        self.conn.commit()

    def _migrate(self):
        self._add_column_if_missing("companies", "created_by", "TEXT")
        self._add_column_if_missing("companies", "sopd_file_path", "TEXT")
        self._add_column_if_missing("documents", "sort_order", "INTEGER NOT NULL DEFAULT 0")
        self._add_column_if_missing("documents", "created_by", "TEXT")
        self._add_column_if_missing("documents", "accept_date", "TEXT")
        self._add_column_if_missing("documents", "updated_by", "TEXT")
        self._add_column_if_missing("documents", "review_due", "TEXT")
        self._add_column_if_missing("documents", "comment", "TEXT")
        self._add_column_if_missing("documents", "needs_office", "INTEGER NOT NULL DEFAULT 0")
        self._add_column_if_missing("documents", "status", f"TEXT NOT NULL DEFAULT '{DOC_STATUS_DEFAULT}'")

        self._add_column_if_missing("documents", "updated_at", "TEXT")
        self._add_column_if_missing("documents", "created_at", "TEXT")
        self._add_column_if_missing("sopd_records", "company_id", "INTEGER NOT NULL DEFAULT 0")
        self._add_column_if_missing("sopd_records", "consent_type", "TEXT NOT NULL DEFAULT ''")
        self._add_column_if_missing("sopd_records", "purpose", "TEXT NOT NULL DEFAULT ''")
        self._add_column_if_missing("sopd_records", "legal_basis", "TEXT")
        self._add_column_if_missing("sopd_records", "pd_categories", "TEXT")
        self._add_column_if_missing("sopd_records", "data_subjects", "TEXT")
        self._add_column_if_missing("sopd_records", "pd_list", "TEXT")
        self._add_column_if_missing("sopd_records", "processing_operations", "TEXT")
        self._add_column_if_missing("sopd_records", "processing_method", "TEXT")
        self._add_column_if_missing("sopd_records", "third_party_transfer", "TEXT")
        self._add_column_if_missing("sopd_records", "transfer_to", "TEXT")
        self._add_column_if_missing("sopd_records", "sopd_description", "TEXT")
        self._add_column_if_missing("sopd_records", "attachment_path", "TEXT")
        self._add_column_if_missing("sopd_records", "validity_period", "TEXT")
        self._add_column_if_missing("sopd_records", "sort_order", "INTEGER NOT NULL DEFAULT 0")
        self._add_column_if_missing("sopd_records", "created_at", "TEXT")
        self._add_column_if_missing("sopd_records", "created_by", "TEXT")
        self._add_column_if_missing("sopd_records", "updated_at", "TEXT")
        self._add_column_if_missing("sopd_records", "updated_by", "TEXT")

        cur = self.conn.cursor()

        ts = now_iso()
        cur.execute("UPDATE documents SET created_at=? WHERE created_at IS NULL OR created_at='';", (ts,))
        cur.execute("UPDATE documents SET updated_at=created_at WHERE updated_at IS NULL OR updated_at='';")
        cur.execute("UPDATE documents SET comment='' WHERE comment IS NULL;")
        cur.execute("UPDATE documents SET needs_office=0 WHERE needs_office IS NULL;")

        cur.execute("""
            UPDATE documents
            SET review_due = NULL
            WHERE review_due IS NULL
            OR TRIM(review_due) = ''
            OR date(TRIM(review_due)) IS NULL;
        """)

        cur.execute("UPDATE documents SET sort_order = id WHERE sort_order = 0;")
        cur.execute("""
            UPDATE documents
            SET status = ?
            WHERE status IS NULL OR TRIM(status) = '';
        """, (DOC_STATUS_DEFAULT,))

        # Лёгкая авто-миграция статусов для старых записей без статуса.
        cur.execute("""
            UPDATE documents
            SET status='На пересмотре'
            WHERE (status IS NULL OR TRIM(status)='' OR status=?)
              AND review_due IS NOT NULL
              AND TRIM(review_due) <> ''
              AND date(TRIM(review_due)) IS NOT NULL
              AND date(TRIM(review_due)) <= date('now');
        """, (DOC_STATUS_DEFAULT,))
        cur.execute("""
            UPDATE documents
            SET status='Действует'
            WHERE (status IS NULL OR TRIM(status)='' OR status=?)
              AND accept_date IS NOT NULL
              AND TRIM(accept_date) <> ''
              AND date(TRIM(accept_date)) IS NOT NULL;
        """, (DOC_STATUS_DEFAULT,))

        allowed = tuple(DOC_STATUSES)
        placeholders = ",".join("?" for _ in allowed)
        cur.execute(
            f"UPDATE documents SET status=? WHERE status NOT IN ({placeholders});",
            (DOC_STATUS_DEFAULT, *allowed),
        )
        self.conn.commit()

        cur.execute("CREATE INDEX IF NOT EXISTS idx_companies_name ON companies(name);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_documents_company_order ON documents(company_id, sort_order, id);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_documents_company_title ON documents(company_id, doc_title);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_document_history_doc ON document_history(doc_id, id DESC);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sections_company_order ON sections(company_id, sort_order, id);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sections_company_name ON sections(company_id, name);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_doc_sections_section ON document_sections(section_id, doc_id);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_doc_sections_doc ON document_sections(doc_id, section_id);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sopd_company_order ON sopd_records(company_id, sort_order, id);")
        self.conn.commit()

    def _append_document_history(
        self,
        cur: sqlite3.Cursor,
        doc_id: int,
        event_type: str,
        event_text: str,
        created_at: Optional[str] = None,
    ):
        text = (event_text or "").strip()
        if not text:
            return
        cur.execute(
            """
            INSERT INTO document_history(doc_id, event_type, event_text, created_at, created_by)
            VALUES (?, ?, ?, ?, ?);
            """,
            (doc_id, event_type, text, created_at or now_iso(), self.user),
        )

    def _company_name_for_history(self, cur: sqlite3.Cursor, company_id: Optional[int]) -> str:
        if company_id in (None, ""):
            return "не указана"
        row = cur.execute("SELECT name FROM companies WHERE id=?;", (int(company_id),)).fetchone()
        return str(row["name"] or "").strip() if row else f"#{company_id}"

    def _section_names_for_history(self, cur: sqlite3.Cursor, section_ids: List[int]) -> str:
        normalized = sorted({int(section_id) for section_id in (section_ids or [])})
        if not normalized:
            return ""
        placeholders = ",".join("?" for _ in normalized)
        rows = cur.execute(
            f"SELECT name FROM sections WHERE id IN ({placeholders}) ORDER BY sort_order ASC, id ASC;",
            tuple(normalized),
        ).fetchall()
        return ", ".join(str(row["name"] or "").strip() for row in rows if str(row["name"] or "").strip())

    def list_document_history(self, doc_id: int, limit: int = 12):
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT event_type, event_text, created_at, created_by
            FROM document_history
            WHERE doc_id=?
            ORDER BY id DESC
            LIMIT ?;
            """,
            (doc_id, max(1, int(limit))),
        )
        return cur.fetchall()

    def list_companies(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, name FROM companies ORDER BY name;")
        return cur.fetchall()

    def list_companies_with_stats(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                c.id,
                c.name,
                COUNT(DISTINCT d.id) AS docs_count,
                COUNT(DISTINCT s.id) AS sections_count
            FROM companies c
            LEFT JOIN documents d ON d.company_id = c.id
            LEFT JOIN sections s ON s.company_id = c.id
            GROUP BY c.id, c.name
            ORDER BY c.name COLLATE NOCASE ASC;
        """)
        return cur.fetchall()

    def add_company(self, name: str) -> int:
        name = name.strip()
        if not name:
            raise ValueError("Название компании не может быть пустым.")

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute(
                "INSERT INTO companies(name, created_at, created_by) VALUES (?, ?, ?);",
                (name, now_iso(), self.user),
            )
            self.conn.commit()
            return cur.lastrowid

        return self._retry_write(work)

    def rename_company(self, company_id: int, new_name: str):
        new_name = new_name.strip()
        if not new_name:
            raise ValueError("Название компании не может быть пустым.")

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("UPDATE companies SET name=? WHERE id=?;", (new_name, company_id))
            self.conn.commit()

        self._retry_write(work)

    def delete_company(self, company_id: int):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("DELETE FROM companies WHERE id=?;", (company_id,))
            self.conn.commit()

        self._retry_write(work)

    def list_documents_filtered(self, company_id: int, section_id: Optional[int]):
        cur = self.conn.cursor()

        if section_id is None:
            # Все документы
            cur.execute("""
                SELECT id, doc_title, status, pdf_path, office_path, review_due, accept_date,
                    sort_order, updated_at, updated_by
                FROM documents
                WHERE company_id=?
                ORDER BY sort_order ASC, id ASC;
            """, (company_id,))
            return cur.fetchall()

        if section_id == -1:
            # Без раздела
            cur.execute("""
                SELECT d.id, d.doc_title, d.status, d.pdf_path, d.office_path, d.review_due, d.accept_date,
                    d.sort_order, d.updated_at, d.updated_by
                FROM documents d
                LEFT JOIN document_sections ds ON ds.doc_id = d.id
                WHERE d.company_id=? AND ds.doc_id IS NULL
                ORDER BY d.sort_order ASC, d.id ASC;
            """, (company_id,))
            return cur.fetchall()

        # Конкретный раздел
        cur.execute("""
            SELECT d.id, d.doc_title, d.status, d.pdf_path, d.office_path, d.review_due, d.accept_date,
                d.sort_order, d.updated_at, d.updated_by
            FROM documents d
            JOIN document_sections ds ON ds.doc_id = d.id
            WHERE d.company_id=? AND ds.section_id=?
            ORDER BY d.sort_order ASC, d.id ASC;
        """, (company_id, section_id))
        return cur.fetchall()

    # ---------- sections ----------
    def list_sections(self, company_id: int):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT id, name, sort_order
            FROM sections
            WHERE company_id=?
            ORDER BY sort_order ASC, id ASC;
        """, (company_id,))
        return cur.fetchall()

    def list_sections_with_doc_counts(self, company_id: int):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                s.id,
                s.name,
                s.sort_order,
                COUNT(DISTINCT d.id) AS docs_count
            FROM sections s
            LEFT JOIN document_sections ds ON ds.section_id = s.id
            LEFT JOIN documents d ON d.id = ds.doc_id AND d.company_id = s.company_id
            WHERE s.company_id = ?
            GROUP BY s.id, s.name, s.sort_order
            ORDER BY s.sort_order ASC, s.id ASC;
        """, (company_id,))
        return cur.fetchall()

    def count_documents_by_company(self, company_id: int) -> int:
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) AS cnt FROM documents WHERE company_id=?;", (company_id,))
        return int(cur.fetchone()["cnt"])

    def count_documents_without_section(self, company_id: int) -> int:
        cur = self.conn.cursor()
        cur.execute("""
            SELECT COUNT(*) AS cnt
            FROM documents d
            LEFT JOIN document_sections ds ON ds.doc_id = d.id
            WHERE d.company_id = ? AND ds.doc_id IS NULL;
        """, (company_id,))
        return int(cur.fetchone()["cnt"])

    def add_section(self, company_id: int, name: str) -> int:
        name = (name or "").strip()
        if not name:
            raise ValueError("Название раздела не может быть пустым.")

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("SELECT COALESCE(MAX(sort_order),0)+1 AS nxt FROM sections WHERE company_id=?;", (company_id,))
            nxt = int(cur.fetchone()["nxt"])
            cur.execute("""
                INSERT INTO sections(company_id, name, sort_order, created_at, created_by)
                VALUES (?, ?, ?, ?, ?);
            """, (company_id, name, nxt, now_iso(), self.user))
            self.conn.commit()
            return cur.lastrowid

        return self._retry_write(work)

    def rename_section(self, section_id: int, new_name: str):
        new_name = (new_name or "").strip()
        if not new_name:
            raise ValueError("Название раздела не может быть пустым.")

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("""
                UPDATE sections
                SET name=?
                WHERE id=?;
            """, (new_name, section_id))
            self.conn.commit()

        self._retry_write(work)

    def delete_section(self, section_id: int):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            # связи удалятся каскадом через FK на section_id, но можно явно:
            cur.execute("DELETE FROM document_sections WHERE section_id=?;", (section_id,))
            cur.execute("DELETE FROM sections WHERE id=?;", (section_id,))
            self.conn.commit()

        self._retry_write(work)

    # ---------- СОПД ----------
    def _normalize_sopd_payload(self, values: Dict[str, Any]) -> Dict[str, str]:
        transfer = str(values.get("third_party_transfer") or "").strip()
        if transfer not in {"Да", "Нет", "Не указано"}:
            transfer = "Не указано"
        transfer_to = str(values.get("transfer_to") or "").strip()
        if transfer != "Да":
            transfer_to = ""
        payload = {
            "consent_type": str(values.get("consent_type") or "").strip(),
            "purpose": str(values.get("purpose") or "").strip(),
            "legal_basis": str(values.get("legal_basis") or "").strip(),
            "pd_categories": str(values.get("pd_categories") or "").strip(),
            "data_subjects": str(values.get("data_subjects") or "").strip(),
            "pd_list": str(values.get("pd_list") or "").strip(),
            "processing_operations": str(values.get("processing_operations") or "").strip(),
            "processing_method": str(values.get("processing_method") or "").strip(),
            "third_party_transfer": transfer,
            "transfer_to": transfer_to,
            "sopd_description": str(values.get("sopd_description") or "").strip(),
            "validity_period": str(values.get("validity_period") or "").strip(),
        }
        if not payload["consent_type"]:
            raise ValueError("Укажи вид согласия.")
        if not payload["purpose"]:
            raise ValueError("Укажи цель обработки.")
        if payload["third_party_transfer"] == "Да" and not payload["transfer_to"]:
            raise ValueError("При передаче третьим лицам укажи, кому передаются данные.")
        return payload

    def list_sopd_records(self, company_id: int):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                id, company_id, consent_type, purpose, legal_basis, pd_categories, data_subjects, pd_list,
                processing_operations, processing_method, third_party_transfer, transfer_to, sopd_description,
                attachment_path, validity_period,
                sort_order, created_at, created_by, updated_at, updated_by
            FROM sopd_records
            WHERE company_id=?
            ORDER BY sort_order ASC, id ASC;
        """, (company_id,))
        return cur.fetchall()

    def count_sopd_records(self, company_id: int) -> int:
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) AS cnt FROM sopd_records WHERE company_id=?;", (company_id,))
        return int(cur.fetchone()["cnt"] or 0)

    def get_sopd_record(self, record_id: int):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                id, company_id, consent_type, purpose, legal_basis, pd_categories, data_subjects, pd_list,
                processing_operations, processing_method, third_party_transfer, transfer_to, sopd_description,
                attachment_path, validity_period,
                sort_order, created_at, created_by, updated_at, updated_by
            FROM sopd_records
            WHERE id=?;
        """, (record_id,))
        return cur.fetchone()

    def add_sopd_record(self, company_id: int, values: Dict[str, Any]) -> int:
        payload = self._normalize_sopd_payload(values)

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("SELECT COALESCE(MAX(sort_order),0)+1 AS nxt FROM sopd_records WHERE company_id=?;", (company_id,))
            nxt = int(cur.fetchone()["nxt"] or 1)
            ts = now_iso()
            cur.execute("""
                INSERT INTO sopd_records(
                    company_id, consent_type, purpose, legal_basis, pd_categories, data_subjects, pd_list,
                    processing_operations, processing_method, third_party_transfer, transfer_to, sopd_description, validity_period,
                    sort_order, created_at, created_by, updated_at, updated_by
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
            """, (
                company_id,
                payload["consent_type"],
                payload["purpose"],
                payload["legal_basis"],
                payload["pd_categories"],
                payload["data_subjects"],
                payload["pd_list"],
                payload["processing_operations"],
                payload["processing_method"],
                payload["third_party_transfer"],
                payload["transfer_to"],
                payload["sopd_description"],
                payload["validity_period"],
                nxt,
                ts,
                self.user,
                ts,
                self.user,
            ))
            self.conn.commit()
            return cur.lastrowid

        return self._retry_write(work)

    def update_sopd_record(self, record_id: int, values: Dict[str, Any]):
        payload = self._normalize_sopd_payload(values)

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("""
                UPDATE sopd_records
                SET
                    consent_type=?,
                    purpose=?,
                    legal_basis=?,
                    pd_categories=?,
                    data_subjects=?,
                    pd_list=?,
                    processing_operations=?,
                    processing_method=?,
                    third_party_transfer=?,
                    transfer_to=?,
                    sopd_description=?,
                    validity_period=?,
                    updated_at=?,
                    updated_by=?
                WHERE id=?;
            """, (
                payload["consent_type"],
                payload["purpose"],
                payload["legal_basis"],
                payload["pd_categories"],
                payload["data_subjects"],
                payload["pd_list"],
                payload["processing_operations"],
                payload["processing_method"],
                payload["third_party_transfer"],
                payload["transfer_to"],
                payload["sopd_description"],
                payload["validity_period"],
                now_iso(),
                self.user,
                record_id,
            ))
            self.conn.commit()

        self._retry_write(work)

    def delete_sopd_record(self, record_id: int):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("DELETE FROM sopd_records WHERE id=?;", (record_id,))
            self.conn.commit()

        self._retry_write(work)

    def update_sopd_attachment_path(self, record_id: int, path: Optional[str]):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute(
                "UPDATE sopd_records SET attachment_path=?, updated_at=?, updated_by=? WHERE id=?;",
                (path, now_iso(), self.user, record_id),
            )
            self.conn.commit()

        self._retry_write(work)

    # ---------- bindings ----------
    def get_doc_section_ids(self, doc_id: int) -> List[int]:
        cur = self.conn.cursor()
        cur.execute("SELECT section_id FROM document_sections WHERE doc_id=? ORDER BY section_id;", (doc_id,))
        return [int(r["section_id"]) for r in cur.fetchall()]

    def set_doc_sections(self, doc_id: int, section_ids: List[int]):
        # полностью заменяем набор разделов у документа
        section_ids = sorted({int(x) for x in (section_ids or [])})

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            old_ids = [
                int(row["section_id"])
                for row in cur.execute(
                    "SELECT section_id FROM document_sections WHERE doc_id=? ORDER BY section_id;",
                    (doc_id,),
                ).fetchall()
            ]
            cur.execute("DELETE FROM document_sections WHERE doc_id=?;", (doc_id,))
            ts = now_iso()
            for sid in section_ids:
                cur.execute("""
                    INSERT OR IGNORE INTO document_sections(doc_id, section_id, created_at, created_by)
                    VALUES (?, ?, ?, ?);
                """, (doc_id, sid, ts, self.user))
            # отметим обновление документа
            cur.execute("UPDATE documents SET updated_at=?, updated_by=? WHERE id=?;", (ts, self.user, doc_id))
            if old_ids != section_ids:
                if section_ids:
                    section_names = self._section_names_for_history(cur, section_ids)
                    text = f"Обновлены разделы: {section_names or 'список разделов изменён'}."
                else:
                    text = "У документа больше нет разделов."
                self._append_document_history(cur, doc_id, "sections", text, created_at=ts)
            self.conn.commit()

        self._retry_write(work)

    def add_doc_to_section(self, doc_id: int, section_id: int):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("""
                INSERT OR IGNORE INTO document_sections(doc_id, section_id, created_at, created_by)
                VALUES (?, ?, ?, ?);
            """, (doc_id, section_id, now_iso(), self.user))
            cur.execute("UPDATE documents SET updated_at=?, updated_by=? WHERE id=?;", (now_iso(), self.user, doc_id))
            self.conn.commit()
        self._retry_write(work)

    def remove_doc_from_section(self, doc_id: int, section_id: int):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("DELETE FROM document_sections WHERE doc_id=? AND section_id=?;", (doc_id, section_id))
            cur.execute("UPDATE documents SET updated_at=?, updated_by=? WHERE id=?;", (now_iso(), self.user, doc_id))
            self.conn.commit()
        self._retry_write(work)

    def list_documents(self, company_id: int):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT id, doc_title, status, pdf_path, office_path, review_due, accept_date,
                sort_order, updated_at, updated_by
            FROM documents
            WHERE company_id=?
            ORDER BY sort_order ASC, id ASC;
        """, (company_id,))
        return cur.fetchall()

    def add_document(self, company_id: int, doc_title: str) -> int:
        doc_title = doc_title.strip()
        if not doc_title:
            raise ValueError("Название документа не может быть пустым.")

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 AS nxt FROM documents WHERE company_id=?;", (company_id,))
            nxt = cur.fetchone()["nxt"]
            ts = now_iso()
            cur.execute("""
                INSERT INTO documents(
                    company_id, doc_title, status, pdf_path, office_path, comment, needs_office, review_due,
                    sort_order, created_at, created_by, updated_at, updated_by
                )
                VALUES (?, ?, ?, NULL, NULL, '', 0, NULL, ?, ?, ?, ?, ?);
            """, (company_id, doc_title, DOC_STATUS_DEFAULT, nxt, ts, self.user, ts, self.user))
            doc_id = int(cur.lastrowid)
            company_name = self._company_name_for_history(cur, company_id)
            self._append_document_history(
                cur,
                doc_id,
                "create",
                f"Создан документ «{doc_title}» в компании «{company_name}».",
                created_at=ts,
            )
            self.conn.commit()
            return doc_id

        return self._retry_write(work)

    def delete_document(self, doc_id: int):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("DELETE FROM documents WHERE id=?;", (doc_id,))
            self.conn.commit()

        self._retry_write(work)

    def update_file_path(self, doc_id: int, field: str, path: Optional[str]):
        if field not in ("pdf_path", "office_path"):
            raise ValueError("Неверный тип пути.")

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            row = cur.execute(f"SELECT {field} FROM documents WHERE id=?;", (doc_id,)).fetchone()
            cur.execute(
                f"UPDATE documents SET {field}=?, updated_at=?, updated_by=? WHERE id=?;",
                (path, now_iso(), self.user, doc_id),
            )
            old_value = str(row[field] or "").strip() if row else ""
            new_value = str(path or "").strip()
            if old_value != new_value:
                file_label = "PDF" if field == "pdf_path" else "дополнительный файл"
                if new_value and old_value:
                    text = f"Заменён {file_label}: {os.path.basename(old_value)} -> {os.path.basename(new_value)}."
                    event_type = "file_replace"
                elif new_value:
                    text = f"Загружен {file_label}: {os.path.basename(new_value)}."
                    event_type = "file_upload"
                else:
                    text = f"Удалён {file_label}: {os.path.basename(old_value)}."
                    event_type = "file_delete"
                self._append_document_history(cur, doc_id, event_type, text)
            self.conn.commit()

        self._retry_write(work)

    def get_doc_info(self, doc_id: int):
        cur = self.conn.cursor()
        cur.execute("SELECT pdf_path, office_path FROM documents WHERE id=?;", (doc_id,))
        return cur.fetchone()

    def get_document_record(self, doc_id: int):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                d.id,
                d.company_id,
                c.name AS company_name,
                d.doc_title,
                d.status,
                d.pdf_path,
                d.office_path,
                d.comment,
                d.needs_office,
                d.review_due,
                d.accept_date,
                d.sort_order,
                d.created_at,
                d.created_by,
                d.updated_at,
                d.updated_by,
                COALESCE((
                    SELECT GROUP_CONCAT(s.name, ' • ')
                    FROM document_sections ds
                    LEFT JOIN sections s ON s.id = ds.section_id
                    WHERE ds.doc_id = d.id
                ), '') AS section_names,
                COALESCE((
                    SELECT GROUP_CONCAT(ds.section_id, ',')
                    FROM document_sections ds
                    WHERE ds.doc_id = d.id
                ), '') AS section_ids
            FROM documents d
            JOIN companies c ON c.id = d.company_id
            WHERE d.id=?;
        """, (doc_id,))
        return cur.fetchone()

    def update_document_record(
        self,
        doc_id: int,
        company_id: int,
        doc_title: str,
        status: str,
        accept_date: Optional[str],
        review_due: Optional[str],
        pdf_path: Optional[str],
        office_path: Optional[str],
        comment: Optional[str],
        needs_office: bool,
    ):
        doc_title = (doc_title or "").strip()
        if not doc_title:
            raise ValueError("Название документа не может быть пустым.")
        status = self._normalize_doc_status(status)
        comment = (comment or "").strip()
        needs_office_value = 1 if bool(needs_office) else 0

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            old_row = cur.execute("""
                SELECT company_id, doc_title, status, accept_date, review_due, comment, needs_office
                FROM documents
                WHERE id=?;
            """, (doc_id,)).fetchone()
            ts = now_iso()
            cur.execute("""
                UPDATE documents
                SET
                    company_id=?,
                    doc_title=?,
                    status=?,
                    accept_date=?,
                    review_due=?,
                    pdf_path=?,
                    office_path=?,
                    comment=?,
                    needs_office=?,
                    updated_at=?,
                    updated_by=?
                WHERE id=?;
            """, (
                company_id,
                doc_title,
                status,
                accept_date,
                review_due,
                pdf_path,
                office_path,
                comment,
                needs_office_value,
                ts,
                self.user,
                doc_id,
            ))
            changes: List[str] = []
            if old_row:
                old_company_id = int(old_row["company_id"])
                if old_company_id != company_id:
                    changes.append(
                        f"Компания: «{self._company_name_for_history(cur, old_company_id)}» -> «{self._company_name_for_history(cur, company_id)}»"
                    )
                old_title = str(old_row["doc_title"] or "").strip()
                if old_title != doc_title:
                    changes.append(f"Название: «{old_title or 'без названия'}» -> «{doc_title}»")
                old_status = self._normalize_doc_status(old_row["status"])
                if old_status != status:
                    changes.append(f"Статус: {old_status} -> {status}")
                old_accept_date = (old_row["accept_date"] or "").strip()
                new_accept_date = (accept_date or "").strip()
                if old_accept_date != new_accept_date:
                    changes.append(
                        f"Дата утверждения: {fmt_date_ddmmyyyy(old_accept_date) or old_accept_date or 'не была указана'} -> {fmt_date_ddmmyyyy(new_accept_date) or new_accept_date or 'не указана'}"
                    )
                old_review_due = (old_row["review_due"] or "").strip()
                new_review_due = (review_due or "").strip()
                if old_review_due != new_review_due:
                    changes.append(
                        f"Дата пересмотра: {fmt_date_ddmmyyyy(old_review_due) or old_review_due or 'не была указана'} -> {fmt_date_ddmmyyyy(new_review_due) or new_review_due or 'не указана'}"
                    )
                old_comment = (old_row["comment"] or "").strip()
                if old_comment != comment:
                    changes.append("Комментарий обновлён" if comment else "Комментарий очищен")
                old_needs_office = 1 if bool(old_row["needs_office"]) else 0
                if old_needs_office != needs_office_value:
                    changes.append(
                        "Дополнительный файл помечен как обязательный"
                        if needs_office_value
                        else "Дополнительный файл больше не обязателен"
                    )
            if changes:
                self._append_document_history(
                    cur,
                    doc_id,
                    "update",
                    "Обновлена карточка документа: " + "; ".join(changes) + ".",
                    created_at=ts,
                )
            self.conn.commit()

        self._retry_write(work)

    def list_document_catalog(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                d.id,
                d.company_id,
                c.name AS company_name,
                d.doc_title,
                d.status,
                d.pdf_path,
                d.office_path,
                d.comment,
                d.needs_office,
                d.review_due,
                d.accept_date,
                d.sort_order,
                d.updated_at,
                d.updated_by,
                COALESCE(GROUP_CONCAT(s.name, ' • '), '') AS section_names,
                COALESCE(GROUP_CONCAT(ds.section_id, ','), '') AS section_ids,
                COUNT(ds.section_id) AS sections_count
            FROM documents d
            JOIN companies c ON c.id = d.company_id
            LEFT JOIN document_sections ds ON ds.doc_id = d.id
            LEFT JOIN sections s ON s.id = ds.section_id
            GROUP BY
                d.id,
                d.company_id,
                c.name,
                d.doc_title,
                d.status,
                d.pdf_path,
                d.office_path,
                d.comment,
                d.needs_office,
                d.review_due,
                d.accept_date,
                d.sort_order,
                d.updated_at,
                d.updated_by
            ORDER BY c.name COLLATE NOCASE ASC, d.sort_order ASC, d.id ASC;
        """)
        return cur.fetchall()

    def get_company_sopd_file_path(self, company_id: int) -> Optional[str]:
        cur = self.conn.cursor()
        cur.execute("SELECT sopd_file_path FROM companies WHERE id=?;", (company_id,))
        r = cur.fetchone()
        if not r:
            return None
        return (r["sopd_file_path"] or "").strip() or None

    def update_company_sopd_file_path(self, company_id: int, path: Optional[str]):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("UPDATE companies SET sopd_file_path=? WHERE id=?;", (path, company_id))
            self.conn.commit()

        self._retry_write(work)

    def get_document_title(self, doc_id: int) -> Optional[str]:
        cur = self.conn.cursor()
        cur.execute("SELECT doc_title FROM documents WHERE id=?;", (doc_id,))
        r = cur.fetchone()
        return r["doc_title"] if r else None
    
    def set_review_due(self, doc_id: int, review_due: Optional[str]):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            row = cur.execute("SELECT review_due FROM documents WHERE id=?;", (doc_id,)).fetchone()
            ts = now_iso()
            cur.execute(
                "UPDATE documents SET review_due=?, updated_at=?, updated_by=? WHERE id=?;",
                (review_due, ts, self.user, doc_id),
            )
            old_value = (row["review_due"] or "").strip() if row else ""
            new_value = (review_due or "").strip()
            if old_value != new_value:
                if new_value:
                    text = f"Изменена дата пересмотра: {fmt_date_ddmmyyyy(old_value) or old_value or 'не была указана'} -> {fmt_date_ddmmyyyy(new_value) or new_value}."
                else:
                    text = "Дата пересмотра очищена."
                self._append_document_history(cur, doc_id, "review_due", text, created_at=ts)
            self.conn.commit()
        self._retry_write(work)

    def list_due_reviews(self, today_ymd: str):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                c.name AS company_name,
                d.id AS doc_id,
                d.doc_title,
                d.review_due
            FROM documents d
            JOIN companies c ON c.id = d.company_id
            WHERE d.review_due IS NOT NULL
            AND TRIM(d.review_due) <> ''
            AND date(TRIM(d.review_due)) IS NOT NULL
            AND date(TRIM(d.review_due)) <= date(?)
            AND (d.status IS NULL OR TRIM(d.status) <> 'Архив')
            ORDER BY date(TRIM(d.review_due)) ASC, c.name ASC, d.sort_order ASC, d.id ASC;
        """, (today_ymd,))
        return cur.fetchall()

    def count_due_reviews(self, today_ymd: str) -> int:
        cur = self.conn.cursor()
        cur.execute("""
            SELECT COUNT(*) AS cnt
            FROM documents
            WHERE review_due IS NOT NULL
            AND TRIM(review_due) <> ''
            AND date(TRIM(review_due)) IS NOT NULL
            AND date(TRIM(review_due)) <= date(?)
            AND (status IS NULL OR TRIM(status) <> 'Архив');
        """, (today_ymd,))
        return int(cur.fetchone()["cnt"])

    def set_company_doc_order(self, company_id: int, ordered_doc_ids: List[int]):
        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            for idx, doc_id in enumerate(ordered_doc_ids, start=1):
                cur.execute("""
                    UPDATE documents SET sort_order=?, updated_at=?, updated_by=?
                    WHERE id=? AND company_id=?;
                """, (idx, now_iso(), self.user, doc_id, company_id))
            self.conn.commit()

        self._retry_write(work)

    def get_all_registry_rows(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT
                c.name AS company_name,
                d.id AS doc_id,
                d.doc_title,
                d.status,
                d.pdf_path,
                d.office_path,
                d.review_due,
                d.accept_date,
                d.sort_order,
                d.updated_at,
                d.updated_by
            FROM documents d
            JOIN companies c ON c.id = d.company_id
            ORDER BY c.name, d.sort_order ASC, d.id ASC;
        """)
        return cur.fetchall()

    def copy_company_structure(self, src_company_id: int, new_name: str) -> int:
        new_company_id = self.add_company(new_name)

        def work():
            cur = self.conn.cursor()
            cur.execute("BEGIN IMMEDIATE;")
            cur.execute("""
                SELECT doc_title, status, sort_order
                FROM documents
                WHERE company_id=?
                ORDER BY sort_order ASC, id ASC;
            """, (src_company_id,))
            docs = cur.fetchall()
            ts = now_iso()
            for d in docs:
                cur.execute("""
                    INSERT INTO documents(company_id, doc_title, status, pdf_path, office_path, sort_order, created_at, created_by, updated_at, updated_by)
                    VALUES (?, ?, ?, NULL, NULL, ?, ?, ?, ?, ?);
                """, (new_company_id, d["doc_title"], self._normalize_doc_status(d["status"]), d["sort_order"], ts, self.user, ts, self.user))

            cur.execute("""
                SELECT
                    consent_type, purpose, legal_basis, pd_categories, data_subjects, pd_list,
                    processing_operations, processing_method,
                    third_party_transfer, transfer_to, sopd_description, validity_period, sort_order
                FROM sopd_records
                WHERE company_id=?
                ORDER BY sort_order ASC, id ASC;
            """, (src_company_id,))
            sopd_rows = cur.fetchall()
            for s in sopd_rows:
                cur.execute("""
                    INSERT INTO sopd_records(
                        company_id, consent_type, purpose, legal_basis, pd_categories, data_subjects, pd_list,
                        processing_operations, processing_method, third_party_transfer, transfer_to, sopd_description, validity_period,
                        sort_order, created_at, created_by, updated_at, updated_by
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                """, (
                    new_company_id,
                    s["consent_type"],
                    s["purpose"],
                    s["legal_basis"],
                    s["pd_categories"],
                    s["data_subjects"],
                    s["pd_list"],
                    s["processing_operations"],
                    s["processing_method"],
                    s["third_party_transfer"],
                    s["transfer_to"],
                    s["sopd_description"],
                    s["validity_period"],
                    s["sort_order"],
                    ts,
                    self.user,
                    ts,
                    self.user,
                ))
            self.conn.commit()

        self._retry_write(work)
        return new_company_id

    def list_doc_ids_by_company(self, company_id: int) -> List[int]:
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM documents WHERE company_id=? ORDER BY sort_order ASC, id ASC;", (company_id,))
        return [r["id"] for r in cur.fetchall()]

    def list_sopd_ids_by_company(self, company_id: int) -> List[int]:
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM sopd_records WHERE company_id=? ORDER BY sort_order ASC, id ASC;", (company_id,))
        return [r["id"] for r in cur.fetchall()]

class FileChipButton(QToolButton):
    def __init__(self, parent, base_icon: QIcon, ok_badged_icon: QIcon, missing_badged_icon: QIcon,
                 text: str, get_path_fn, upload_fn, open_fn, download_fn, delete_fn):
        super().__init__(parent)
        self._base_icon = base_icon
        self._miss_icon = missing_badged_icon
        self._ok_icon = ok_badged_icon
        self._text = text
        self._get_path = get_path_fn
        self._upload = upload_fn
        self._open = open_fn
        self._download = download_fn
        self._delete = delete_fn

        self.setProperty("chip", True)
        self.setFixedSize(112, 30)
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.setAutoRaise(False)

        self.clicked.connect(self._on_click)
        self.refresh_visual()

        self.setToolButtonStyle(Qt.ToolButtonTextOnly)
        self.setIcon(QIcon())
        self.setIconSize(QSize(0, 0))

        self.setFixedSize(96, 28)

        slot = "pdf" if self._text.strip().lower() == "pdf" else "office"
        self.setProperty("slot", slot)

    def _on_click(self):
        path = self._get_path()
        if path and os.path.exists(path):
            self._open()
        else:
            self._upload()

    def refresh_visual(self):
        path = self._get_path()
        has_path = bool(path)
        exists = bool(path) and os.path.exists(path)

        if not has_path:
            state = "empty"
            self.setToolTip("Файл не загружен")
        elif exists:
            state = "ok"
            self.setToolTip(path)
        else:
            state = "missing"
            self.setToolTip(f"Файл указан, но не найден:\n{path}")

        self.setText(self._text)
        self.setProperty("state", state)
        self.style().unpolish(self)
        self.style().polish(self)

    def mouseDoubleClickEvent(self, event):
        path = self._get_path()
        if path and os.path.exists(path):
            self._open()
        else:
            self._upload()
        event.accept()

    def contextMenuEvent(self, event):
        menu = QMenu(self)
        path = self._get_path()
        exists = bool(path) and os.path.exists(path)

        if exists:
            a_open = menu.addAction("Открыть")
            a_dl = menu.addAction("Скачать как…")
            menu.addSeparator()
            a_del = menu.addAction("Удалить файл")
            act = menu.exec(event.globalPos())
            if act == a_open:
                self._open()
            elif act == a_dl:
                self._download()
            elif act == a_del:
                self._delete()
        else:
            a_up = menu.addAction("Загрузить файл…")
            act = menu.exec(event.globalPos())
            if act == a_up:
                self._upload()

class DocumentTable(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._main = None

        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDragDropMode(QAbstractItemView.InternalMove)

        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._ctx_menu)

        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setWordWrap(False)
        self.setTextElideMode(Qt.ElideRight)
        self.setShowGrid(False)
        self.verticalHeader().setDefaultSectionSize(52)

        hh = self.horizontalHeader()
        hh.setHighlightSections(False)
        hh.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)

    def set_main(self, main):
        self._main = main

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            if not self._main:
                return
            urls = event.mimeData().urls()
            if not urls:
                return

            pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
            row = self.rowAt(pos.y())
            if row < 0:
                return

            doc_id_item = self.item(row, 0)
            if not doc_id_item or not doc_id_item.text().isdigit():
                return
            doc_id = int(doc_id_item.text())

            local_path = urls[0].toLocalFile()
            if not local_path or not os.path.isfile(local_path):
                return

            ext = os.path.splitext(local_path)[1].lower()
            if ext == ".pdf":
                self._main.upload_file_from_path(doc_id, "pdf", local_path)
            elif ext in (".doc", ".docx", ".xls", ".xlsx"):
                self._main.upload_file_from_path(doc_id, "office", local_path)
            else:
                self._main.warn("Формат не поддерживается. Можно: PDF, DOC/DOCX, XLS/XLSX.")
            event.acceptProposedAction()
            return

        super().dropEvent(event)
        if self._main:
            self._main.persist_current_doc_order()

    def _ctx_menu(self, pos: QPoint):
        if not self._main:
            return

        row = self.rowAt(pos.y())
        if row < 0:
            return

        it = self.item(row, 0)
        if not it or not it.text().isdigit():
            return
        doc_id = int(it.text())
        current_status = DOC_STATUS_DEFAULT
        st_it = self.item(row, 2)
        if st_it:
            current_status = (st_it.data(Qt.UserRole) or st_it.text() or DOC_STATUS_DEFAULT).strip()
        if current_status not in DOC_STATUSES:
            current_status = DOC_STATUS_DEFAULT

        menu = QMenu(self)
        m_status = menu.addMenu("Статус")
        status_actions: Dict[QAction, str] = {}
        for st in DOC_STATUSES:
            a = m_status.addAction(st)
            a.setCheckable(True)
            a.setChecked(st == current_status)
            status_actions[a] = st

        menu.addSeparator()
        a_accept = menu.addAction("Дата принятия…")
        a_clear_accept = menu.addAction("Очистить дату принятия")
        menu.addSeparator()
        a_review = menu.addAction("Дата пересмотра…")
        a_clear_review = menu.addAction("Очистить дату пересмотра")
        menu.addSeparator()
        a_sections = menu.addAction("Разделы…")
        a_rename = menu.addAction("Переименовать")
        a_del = menu.addAction("Удалить документ")
        act = menu.exec(self.viewport().mapToGlobal(pos))

        if act in status_actions:
            self._main.set_document_status(doc_id, status_actions[act])
        elif act == a_accept:
            self._main.set_accept_date_dialog(doc_id)
        elif act == a_clear_accept:
            self._main.clear_accept_date(doc_id)
        elif act == a_review:
            self._main.set_review_date_dialog(doc_id)
        elif act == a_clear_review:
            self._main.clear_review_date(doc_id)
        elif act == a_sections:
            self._main.edit_doc_sections_dialog(doc_id)
        elif act == a_rename:
            self._main.rename_document_row(doc_id)
        elif act == a_del:
            self._main.delete_document_row(doc_id)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.current_account = f"{getpass.getuser()}@{platform.node()}"
        self.current_section_id: Optional[int] = None
        self.shared_root = ""
        self.db_path = ""
        self.storage_dir = ""
        self.db = None
        self.doc_cache: Dict[int, Dict[str, Optional[str]]] = {}
        self._due_reviews_count = 0
        self._sidebar_visible = True
        self._content_mode = "docs"
        self._thread_pool = QThreadPool.globalInstance()
        self._bg_tasks: set[BackgroundTask] = set()
        self._busy_jobs = 0
        self._busy_cursor_set = False
        self._busy_label = ""

        self.setWindowTitle(APP_TITLE)
        self.setFixedSize(WINDOW_W, WINDOW_H)

        self.app_stack = QStackedWidget()
        self.app_stack.setObjectName("appStack")
        self.setCentralWidget(self.app_stack)

        self._doc_search_timer = QTimer(self)
        self._doc_search_timer.setSingleShot(True)
        self._doc_search_timer.timeout.connect(self.apply_doc_filter)

        self.page_empty_app = self._build_empty_app_page()
        self.page_main = self._build_main_page()

        self.app_stack.addWidget(self.page_empty_app)  # 0
        self.app_stack.addWidget(self.page_main)       # 1

        self._apply_style()
        menu = self.menuBar().addMenu("Настройки")
        act_change = QAction("Сменить общую папку…", self)
        act_change.triggered.connect(self.change_shared_root)
        menu.addAction(act_change)

        menu_registry = self.menuBar().addMenu("Реестр")
        menu_view = self.menuBar().addMenu("Вид")
        self._force_viewport_color(self.company_list, "#162437")
        self._force_viewport_color(self.table, "#131f31")
        act_new_company = QAction(self)
        act_new_company.setShortcut("Ctrl+N")
        act_new_company.triggered.connect(self.add_company)
        self.addAction(act_new_company)

        act_new_doc = QAction(self)
        act_new_doc.setShortcut("Ctrl+D")
        act_new_doc.triggered.connect(self.add_document)
        self.addAction(act_new_doc)

        act_export = QAction("Выгрузить в Excel…", self)
        act_export.setShortcut("Ctrl+E")
        act_export.triggered.connect(self.export_registry)
        menu_registry.addAction(act_export)
        self.addAction(act_export)

        act_notifications = QAction("Центр уведомлений", self)
        act_notifications.triggered.connect(self.open_notifications_center)
        menu_registry.addAction(act_notifications)

        self._act_toggle_sidebar = QAction("Скрыть боковую панель", self)
        self._act_toggle_sidebar.setShortcut("Ctrl+B")
        self._act_toggle_sidebar.triggered.connect(self.toggle_sidebar)
        menu_view.addAction(self._act_toggle_sidebar)
        self.addAction(self._act_toggle_sidebar)

        self._notif_timer = QTimer(self)
        self._notif_timer.setInterval(60_000)
        self._notif_timer.timeout.connect(self.refresh_notifications)
        self._notif_timer.start()

        self._update_statusbar()

        QTimer.singleShot(0, self.bootstrap)

    def _force_viewport_color(self, w: QWidget, hex_color: str = "#17182a"):
        c = QColor(hex_color)
        pal = w.palette()
        pal.setColor(QPalette.Base, c)
        pal.setColor(QPalette.Window, c) 
        w.setAutoFillBackground(True)
        if hasattr(w, "viewport"):
            w.viewport().setPalette(pal)
            w.viewport().setAutoFillBackground(True)

    def add_section(self):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            self.warn("Сначала выбери компанию, затем добавь раздел.")
            return

        dlg = InputDialog(self, "Новый раздел", "Название раздела:")
        if dlg.exec() != QDialog.Accepted:
            return

        try:
            new_section_id = self.db.add_section(company_id, dlg.value())
            self.load_sections(company_id, preferred_section_id=new_section_id)
        except sqlite3.IntegrityError:
            self.warn("Раздел с таким названием уже существует.")
        except Exception as e:
            self.warn(str(e))

    def section_context_menu(self, pos: QPoint):
        it = self.section_list.itemAt(pos)
        if not it:
            return

        section_id = it.data(Qt.UserRole)
        name = it.data(Qt.UserRole + 1) or it.text()
        name = str(name)

        # виртуальные пункты не трогаем
        if section_id in (None, -1):
            menu = QMenu(self)
            a_add = menu.addAction("Добавить раздел")
            act = menu.exec(self.section_list.viewport().mapToGlobal(pos))
            if act == a_add:
                self.add_section()
            return

        menu = QMenu(self)
        a_rename = menu.addAction("Переименовать")
        a_del = menu.addAction("Удалить раздел")
        menu.addSeparator()
        a_add = menu.addAction("Добавить раздел")

        act = menu.exec(self.section_list.viewport().mapToGlobal(pos))
        company_id, _ = self.current_company_id_name()

        if act == a_add:
            self.add_section()
        elif act == a_rename:
            dlg = InputDialog(self, "Переименовать раздел", "Новое название:", default=name)
            if dlg.exec() == QDialog.Accepted:
                try:
                    self.db.rename_section(int(section_id), dlg.value())
                    self.load_sections(company_id, preferred_section_id=int(section_id))
                except sqlite3.IntegrityError:
                    self.warn("Раздел с таким названием уже существует.")
                except Exception as e:
                    self.warn(str(e))
        elif act == a_del:
            if self.confirm(f"Удалить раздел «{name}»?\nДокументы останутся, просто исчезнет привязка."):
                try:
                    self.db.delete_section(int(section_id))
                    self.load_sections(company_id, preferred_section_id=None)
                except Exception as e:
                    self.warn(str(e))

    def edit_doc_sections_dialog(self, doc_id: int):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return

        sections = self.db.list_sections(company_id)
        chosen = set(self.db.get_doc_section_ids(doc_id))

        dlg = PrettyDialog(self, "Разделы документа", self.style().standardIcon(QStyle.SP_DirIcon))
        box = dlg.content_layout()

        info = QLabel("Отметь разделы, в которых должен быть этот документ:")
        info.setWordWrap(True)
        box.addWidget(info)

        # скролл с чекбоксами
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        host = QWidget()
        v = QVBoxLayout(host)
        v.setContentsMargins(10, 10, 10, 10)
        v.setSpacing(6)

        checks: List[Tuple[int, QCheckBox]] = []
        for r in sections:
            sid = int(r["id"])
            cb = QCheckBox(r["name"])
            cb.setChecked(sid in chosen)
            v.addWidget(cb)
            checks.append((sid, cb))

        v.addStretch(1)
        scroll.setWidget(host)
        scroll.setMinimumHeight(260)
        box.addWidget(scroll)

        # кнопки
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.button(QDialogButtonBox.Ok).setText("Сохранить")
        bb.button(QDialogButtonBox.Cancel).setText("Отмена")
        style_dialog_buttons(bb)
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        box.addWidget(bb)

        fix_dialog_size(dlg)

        if dlg.exec() != QDialog.Accepted:
            return

        new_ids = [sid for sid, cb in checks if cb.isChecked()]
        try:
            self.db.set_doc_sections(doc_id, new_ids)
            self.load_documents()
        except Exception as e:
            self.warn(str(e))

    def _read_registry(self, root: str) -> Optional[dict]:
        path = os.path.join(root, REGISTRY_FILE)
        if not os.path.exists(path):
            return None
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _write_registry(self, root: str):
        path = os.path.join(root, REGISTRY_FILE)
        data = {
            "db": "pd_docs.db",
            "storage": ".pd_docs"
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def refresh_notifications(self, due_reviews_count: Optional[int] = None):
        if not self.db and due_reviews_count is None:
            return
        if due_reviews_count is None:
            self._due_reviews_count = self.db.count_due_reviews(today_ymd())
        else:
            self._due_reviews_count = max(0, int(due_reviews_count))
        self._update_statusbar()

    def open_notifications_center(self):
        if not self.db:
            return
        rows = self.db.list_due_reviews(today_ymd())
        if not rows:
            self.info("Пока нет документов, которые пора пересмотреть.")
            return

        lines = []
        for r in rows:
            due = r["review_due"] or ""
            lines.append(f"• {r['company_name']} → {r['doc_title']} (до {fmt_date_ddmmyyyy(due)})")

        ThemedMessageDialog(
            self,
            "Центр уведомлений",
            "Пора пересмотреть документы:\n\n" + "\n".join(lines),
            kind="warn",
            buttons=("ok",)
        ).exec()

    def _normalize_status(self, status: Optional[str]) -> str:
        st = (status or "").strip()
        return st if st in DOC_STATUSES else DOC_STATUS_DEFAULT

    def _status_colors(self, status: str) -> Tuple[str, str]:
        return DOC_STATUS_COLORS.get(status, DOC_STATUS_COLORS[DOC_STATUS_DEFAULT])

    def set_document_status(self, doc_id: int, status: str):
        try:
            self.db.set_document_status(doc_id, status)
            self.load_documents()
        except Exception as e:
            self.warn(str(e))

    def _render_documents(self, company_id: int, rows):
        if len(rows) == 0:
            has_sopd = bool(self.db and self.db.count_sopd_records(company_id) > 0)
            if not has_sopd:
                self.right_stack.setCurrentIndex(0)
                return
            self.right_stack.setCurrentIndex(1)
            self.doc_cache = {}
            self.table.clearContents()
            self.table.setRowCount(0)
            self._hide_docs_empty_search()
            return
        self.right_stack.setCurrentIndex(1)

        self.doc_cache = {
            int(r["id"]): {
                "pdf": self._abs_storage_path(r["pdf_path"]),
                "office": self._abs_storage_path(r["office_path"]),
            }
            for r in rows
        }

        if not hasattr(self, "_chip_icons_ready"):
            base_pdf = self.style().standardIcon(QStyle.SP_FileIcon)
            base_off = self.style().standardIcon(QStyle.SP_DirIcon)
            badge_ok = self.style().standardIcon(QStyle.SP_DialogApplyButton)
            badge_warn = self.style().standardIcon(QStyle.SP_MessageBoxWarning)

            self._icon_base_pdf = base_pdf
            self._icon_base_off = base_off
            self._icon_ok_pdf = compose_badged_icon(base_pdf, badge_ok, size=16)
            self._icon_ok_off = compose_badged_icon(base_off, badge_ok, size=16)
            self._icon_miss_pdf = compose_badged_icon(base_pdf, badge_warn, size=16)
            self._icon_miss_off = compose_badged_icon(base_off, badge_warn, size=16)
            self._chip_icons_ready = True

        self.table.setUpdatesEnabled(False)
        self.table.blockSignals(True)
        try:
            self.table.clearContents()
            self.table.setRowCount(len(rows))

            for row_i, r in enumerate(rows):
                self.table.setRowHeight(row_i, 52)

                doc_id = int(r["id"])
                title = r["doc_title"]
                status = self._normalize_status(r["status"])
                review_due = (r["review_due"] or "").strip()
                accept_date = (r["accept_date"] or "").strip()

                self.table.setItem(row_i, 0, QTableWidgetItem(str(doc_id)))

                t_item = QTableWidgetItem(title)
                t_item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self.table.setItem(row_i, 1, t_item)

                s_item = QTableWidgetItem(status)
                s_item.setData(Qt.UserRole, status)
                s_item.setTextAlignment(Qt.AlignCenter)
                fg, bg = self._status_colors(status)
                s_item.setForeground(QBrush(QColor(fg)))
                s_item.setBackground(QBrush(QColor(bg)))
                s_item.setToolTip(f"Статус: {status}")
                self.table.setItem(row_i, 2, s_item)

                # PDF/Office колонки заполняем “пустышками” под виджеты
                for col in (3, 4):
                    filler = QTableWidgetItem("")
                    filler.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    self.table.setItem(row_i, col, filler)

                rev_item = QTableWidgetItem(fmt_date_ddmmyyyy(review_due))
                rev_item.setToolTip(review_due)
                self.table.setItem(row_i, 5, rev_item)

                acc_item = QTableWidgetItem(fmt_date_ddmmyyyy(accept_date))
                acc_item.setToolTip(accept_date)
                self.table.setItem(row_i, 6, acc_item)

                def get_pdf_path(did=doc_id):
                    return self.doc_cache.get(did, {}).get("pdf")

                def get_off_path(did=doc_id):
                    return self.doc_cache.get(did, {}).get("office")

                pdf_btn = FileChipButton(
                    self.table,
                    self._icon_base_pdf, self._icon_ok_pdf, self._icon_miss_pdf,
                    "PDF", get_pdf_path,
                    upload_fn=lambda *args, did=doc_id: self.upload_file_dialog(did, "pdf"),
                    open_fn=lambda did=doc_id: self.open_file(did, "pdf"),
                    download_fn=lambda did=doc_id: self.download_file(did, "pdf"),
                    delete_fn=lambda did=doc_id: self.delete_file(did, "pdf"),
                )
                off_btn = FileChipButton(
                    self.table,
                    self._icon_base_off, self._icon_ok_off, self._icon_miss_off,
                    "Office", get_off_path,
                    upload_fn=lambda *args, did=doc_id: self.upload_file_dialog(did, "office"),
                    open_fn=lambda did=doc_id: self.open_file(did, "office"),
                    download_fn=lambda did=doc_id: self.download_file(did, "office"),
                    delete_fn=lambda did=doc_id: self.delete_file(did, "office"),
                )

                self.table.setCellWidget(row_i, 3, self._center_cell_widget(pdf_btn))
                self.table.setCellWidget(row_i, 4, self._center_cell_widget(off_btn))

            # обновим внешний вид чипов
            for i in range(self.table.rowCount()):
                for col in (3, 4):
                    host = self.table.cellWidget(i, col)
                    if host:
                        btn = host.findChild(FileChipButton)
                        if btn:
                            btn.refresh_visual()

        finally:
            self.table.blockSignals(False)
            self.table.setUpdatesEnabled(True)
            self.table.viewport().update()

        self.apply_doc_filter()

    def load_documents(self):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            self._render_sopd_cards([])
            return
        rows = self.db.list_documents_filtered(company_id, self.current_section_id)
        self._render_documents(company_id, rows)
        self.load_sopd_records()

    def _clear_layout_widgets(self, layout: QVBoxLayout):
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()
                continue
            child = item.layout()
            if child is not None:
                self._clear_layout_widgets(child)  # type: ignore[arg-type]

    def _sopd_transfer_state(self, value: str) -> str:
        v = (value or "").strip().lower()
        if v == "да":
            return "yes"
        if v == "нет":
            return "no"
        return "na"

    def _sopd_fill_state(self, filled: int, total: int) -> str:
        if total <= 0:
            return "bad"
        if filled >= total:
            return "ok"
        if filled >= max(1, int(total * 0.6)):
            return "mid"
        return "bad"

    def _sopd_record_file_abs(self, record_id: int) -> Optional[str]:
        if not self.db:
            return None
        rec = self.db.get_sopd_record(record_id)
        if not rec:
            return None
        return self._abs_storage_path(rec["attachment_path"])

    def _sopd_card_completeness(self, row: sqlite3.Row) -> Tuple[int, int, List[str]]:
        transfer_value = (row["third_party_transfer"] or "").strip()
        transfer_norm = transfer_value.lower()
        transfer_known = transfer_norm in {"да", "нет"}
        file_abs = self._abs_storage_path((row["attachment_path"] or "").strip())
        checks: List[Tuple[str, bool]] = [
            ("Вид согласия", bool((row["consent_type"] or "").strip())),
            ("Цель", bool((row["purpose"] or "").strip())),
            ("Категории ПД", bool((row["pd_categories"] or "").strip())),
            ("Перечень ПД", bool((row["pd_list"] or "").strip())),
            ("Способ обработки", bool((row["processing_method"] or "").strip())),
            ("Передача 3-м лицам", transfer_known),
            ("Срок действия", bool((row["validity_period"] or "").strip())),
            ("Описание СОПД", bool((row["sopd_description"] or "").strip())),
            ("Файл СОПД", bool(file_abs) and os.path.exists(file_abs)),
        ]
        if transfer_norm == "да":
            checks.append(("Кому передаётся", bool((row["transfer_to"] or "").strip())))
        filled = sum(1 for _, ok in checks if ok)
        missing = [name for name, ok in checks if not ok]
        return filled, len(checks), missing

    def _build_sopd_card(self, row: sqlite3.Row) -> QFrame:
        card = QFrame()
        card.setObjectName("sopdCard")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(8)

        head = QHBoxLayout()
        head.setSpacing(8)

        title = QLabel((row["consent_type"] or "").strip() or "Согласие")
        title.setObjectName("sopdCardTitle")
        title.setWordWrap(True)
        head.addWidget(title, 1)

        transfer_value = (row["third_party_transfer"] or "").strip() or "Не указано"
        filled, total, missing = self._sopd_card_completeness(row)
        badges = QVBoxLayout()
        badges.setSpacing(5)

        transfer_badge = QLabel(f"Передача: {transfer_value}")
        transfer_badge.setProperty("sopdBadge", True)
        transfer_badge.setProperty("state", self._sopd_transfer_state(transfer_value))
        badges.addWidget(transfer_badge, 0, Qt.AlignRight)

        fill_badge = QLabel(f"Заполнено {filled}/{total}")
        fill_badge.setObjectName("sopdProgress")
        fill_badge.setProperty("state", self._sopd_fill_state(filled, total))
        badges.addWidget(fill_badge, 0, Qt.AlignRight)

        head.addLayout(badges, 0)

        lay.addLayout(head)

        def add_line(caption: str, value: str):
            lbl = QLabel(f"{caption}: {value if value else '—'}")
            lbl.setObjectName("sopdLine")
            lbl.setWordWrap(True)
            lbl.setTextFormat(Qt.PlainText)
            lay.addWidget(lbl)

        add_line("Цель", (row["purpose"] or "").strip())
        add_line("Категории ПД", (row["pd_categories"] or "").strip())
        add_line("Перечень ПД", (row["pd_list"] or "").strip())
        add_line("Способ обработки", (row["processing_method"] or "").strip())
        add_line("Кому передаётся", (row["transfer_to"] or "").strip())
        add_line("Описание СОПД", (row["sopd_description"] or "").strip())
        add_line("Срок действия", (row["validity_period"] or "").strip())

        if missing:
            miss_preview = ", ".join(missing[:5])
            if len(missing) > 5:
                miss_preview += f" и ещё {len(missing) - 5}"
            miss = QLabel(f"Не заполнено: {miss_preview}")
            miss.setObjectName("sopdMissing")
            miss.setWordWrap(True)
            lay.addWidget(miss)

        rid = int(row["id"])
        if not hasattr(self, "_sopd_chip_icons"):
            base_icon = self.style().standardIcon(QStyle.SP_FileIcon)
            ok_badge = compose_badged_icon(base_icon, self.style().standardIcon(QStyle.SP_DialogApplyButton), size=16)
            miss_badge = compose_badged_icon(base_icon, self.style().standardIcon(QStyle.SP_MessageBoxWarning), size=16)
            self._sopd_chip_icons = (base_icon, ok_badge, miss_badge)
        base_icon, ok_badge, miss_badge = self._sopd_chip_icons

        file_row = QHBoxLayout()
        file_row.setSpacing(8)
        file_lbl = QLabel("Файл СОПД:")
        file_lbl.setObjectName("sopdFileLabel")
        file_row.addWidget(file_lbl)
        file_chip = FileChipButton(
            card,
            base_icon,
            ok_badge,
            miss_badge,
            "Файл",
            lambda record_id=rid: self._sopd_record_file_abs(record_id),
            lambda *args, record_id=rid: self.upload_sopd_record_file_dialog(record_id),
            lambda record_id=rid: self.open_sopd_record_file(record_id),
            lambda record_id=rid: self.download_sopd_record_file(record_id),
            lambda record_id=rid: self.delete_sopd_record_file(record_id),
        )
        file_chip.setFixedSize(96, 28)
        file_row.addWidget(file_chip, 0, Qt.AlignLeft)
        file_row.addStretch(1)
        lay.addLayout(file_row)

        actions = QHBoxLayout()
        actions.setSpacing(8)
        actions.addStretch(1)

        btn_edit = QToolButton()
        btn_edit.setText("Изменить")
        btn_edit.setProperty("kind", "ghost")
        btn_edit.clicked.connect(lambda _=False, record_id=rid: self.edit_sopd_record(record_id))
        actions.addWidget(btn_edit)

        btn_delete = QToolButton()
        btn_delete.setText("Удалить")
        btn_delete.setProperty("kind", "ghost")
        btn_delete.clicked.connect(lambda _=False, record_id=rid: self.delete_sopd_record(record_id))
        actions.addWidget(btn_delete)

        lay.addLayout(actions)
        return card

    def _render_sopd_cards(self, rows):
        if not hasattr(self, "sopd_cards_layout"):
            return

        rows = list(rows or [])
        self._clear_layout_widgets(self.sopd_cards_layout)

        if hasattr(self, "sopd_count_label"):
            self.sopd_count_label.setText(str(len(rows)))

        if not rows:
            empty = QLabel("Карточек СОПД пока нет. Нажми «Добавить карточку».")
            empty.setObjectName("sopdHint")
            empty.setWordWrap(True)
            self.sopd_cards_layout.addWidget(empty)
            self.sopd_cards_layout.addStretch(1)
            return

        for row in rows:
            self.sopd_cards_layout.addWidget(self._build_sopd_card(row))
        self.sopd_cards_layout.addStretch(1)

    def load_sopd_records(self):
        if not self.db:
            return
        company_id, _ = self.current_company_id_name()
        if not company_id:
            self._render_sopd_cards([])
            if hasattr(self, "sopd_file_chip"):
                self.sopd_file_chip.refresh_visual()
            return
        rows = self.db.list_sopd_records(company_id)
        self._render_sopd_cards(rows)
        if hasattr(self, "sopd_file_chip"):
            self.sopd_file_chip.refresh_visual()

    def _current_sopd_file_abs(self) -> Optional[str]:
        if not self.db:
            return None
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return None
        rel = self.db.get_company_sopd_file_path(company_id)
        return self._abs_storage_path(rel)

    def upload_company_sopd_file_dialog(self):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            self.warn("Сначала выбери компанию.")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выбери файл СОПД",
            "",
            "Файлы СОПД (*.pdf *.doc *.docx *.xls *.xlsx);;Все файлы (*)",
        )
        if not file_path:
            return
        self.upload_company_sopd_file_from_path(file_path)

    def upload_company_sopd_file_from_path(self, file_path: str):
        company_id, _ = self.current_company_id_name()
        if not company_id or not self.db:
            return

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".pdf", ".doc", ".docx", ".xls", ".xlsx"):
            self.warn("Для файла СОПД поддерживаются только pdf/doc/docx/xls/xlsx.")
            return

        base_name = self._safe_filename(os.path.splitext(os.path.basename(file_path))[0])
        dst = os.path.join(self._sopd_dir(company_id), f"{base_name}{ext}")
        rel_dst = self._rel_storage_path(dst)

        old_rel = self.db.get_company_sopd_file_path(company_id)
        old_abs = self._abs_storage_path(old_rel) if old_rel else None

        def work():
            if old_abs and os.path.exists(old_abs):
                try:
                    os.remove(old_abs)
                except Exception:
                    pass
            atomic_copy(file_path, dst)
            return rel_dst

        def on_success(new_rel: Any):
            self.db.update_company_sopd_file_path(company_id, str(new_rel))
            self.load_sopd_records()

        self._run_background_task(
            label="Загрузка файла СОПД",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось загрузить файл СОПД:",
        )

    def open_company_sopd_file(self):
        path = self._current_sopd_file_abs()
        if not path or not os.path.exists(path):
            self.warn("Файл СОПД не найден.")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(path)))

    def download_company_sopd_file(self):
        path = self._current_sopd_file_abs()
        if not path or not os.path.exists(path):
            self.warn("Файл СОПД не найден.")
            return

        base = os.path.basename(path)
        save_path, _ = QFileDialog.getSaveFileName(self, "Скачать файл СОПД как…", base, "Все файлы (*)")
        if not save_path:
            return

        def work():
            atomic_copy(path, save_path)
            return None

        def on_success(_: Any):
            self.info("Файл СОПД сохранён ✅")

        self._run_background_task(
            label="Скачивание файла СОПД",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось сохранить файл СОПД:",
        )

    def delete_company_sopd_file(self):
        company_id, _ = self.current_company_id_name()
        if not company_id or not self.db:
            return

        rel = self.db.get_company_sopd_file_path(company_id)
        path = self._abs_storage_path(rel)
        if not rel:
            self.db.update_company_sopd_file_path(company_id, None)
            self.load_sopd_records()
            return

        if not self.confirm("Удалить файл СОПД из программы и с диска?"):
            return

        try:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass
            self.db.update_company_sopd_file_path(company_id, None)
            self.load_sopd_records()
        except Exception as e:
            self.warn(str(e))

    def upload_sopd_record_file_dialog(self, record_id: int):
        rec = self.db.get_sopd_record(record_id) if self.db else None
        if not rec:
            self.warn("Карточка СОПД не найдена.")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Загрузить файл СОПД",
            "",
            "Файлы Word (*.doc *.docx)",
        )
        if not file_path:
            return
        self.upload_sopd_record_file_from_path(record_id, file_path)

    def upload_sopd_record_file_from_path(self, record_id: int, file_path: str):
        if not self.db:
            return
        rec = self.db.get_sopd_record(record_id)
        if not rec:
            self.warn("Карточка СОПД не найдена.")
            return

        company_id = int(rec["company_id"])
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".doc", ".docx"):
            self.warn("Для карточки СОПД нужен файл в формате doc или docx.")
            return
        base_name = self._safe_filename(os.path.splitext(os.path.basename(file_path))[0])
        dst = os.path.join(self._sopd_record_dir(company_id, record_id), f"{base_name}{ext}")
        rel_dst = self._rel_storage_path(dst)

        old_rel = (rec["attachment_path"] or "").strip()
        old_abs = self._abs_storage_path(old_rel) if old_rel else None

        def work():
            if old_abs and os.path.exists(old_abs) and os.path.abspath(old_abs) != os.path.abspath(dst):
                try:
                    os.remove(old_abs)
                except Exception:
                    pass
            atomic_copy(file_path, dst)
            return rel_dst

        def on_success(new_rel: Any):
            self.db.update_sopd_attachment_path(record_id, str(new_rel))
            self.load_sopd_records()
            if hasattr(self, "_load_sopd_into_editor") and getattr(self, "current_sopd_id", None) == record_id:
                self._load_sopd_into_editor(record_id)
            self.info("Файл СОПД загружен.")

        self._run_background_task(
            label="Загрузка файла карточки СОПД",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось загрузить файл для карточки СОПД:",
        )

    def open_sopd_record_file(self, record_id: int):
        path = self._sopd_record_file_abs(record_id)
        if not path or not os.path.exists(path):
            self.warn("Файл карточки СОПД не найден.")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(path)))

    def download_sopd_record_file(self, record_id: int):
        path = self._sopd_record_file_abs(record_id)
        if not path or not os.path.exists(path):
            self.warn("Файл карточки СОПД не найден.")
            return

        base = os.path.basename(path)
        save_path, _ = QFileDialog.getSaveFileName(self, "Скачать файл карточки СОПД как…", base, "Все файлы (*)")
        if not save_path:
            return

        def work():
            atomic_copy(path, save_path)
            return None

        def on_success(_: Any):
            self.info("Файл карточки СОПД сохранён ✅")

        self._run_background_task(
            label="Скачивание файла карточки СОПД",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось сохранить файл карточки СОПД:",
        )

    def delete_sopd_record_file(
        self,
        record_id: int,
        ask_confirm: bool = True,
        silent: bool = False,
        refresh_after: bool = True,
    ) -> bool:
        rec = self.db.get_sopd_record(record_id) if self.db else None
        if not rec:
            if not silent:
                self.warn("Карточка СОПД не найдена.")
            return False

        rel = (rec["attachment_path"] or "").strip()
        if not rel:
            if not silent:
                self.warn("У этой карточки файл не прикреплён.")
            return False
        path = self._abs_storage_path(rel)

        if ask_confirm and not self.confirm("Удалить файл карточки СОПД из программы и с диска?"):
            return False

        try:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass

                # Чистим пустую папку карточки, чтобы не копить мусор.
                parent = os.path.dirname(path)
                if os.path.isdir(parent):
                    try:
                        if not os.listdir(parent):
                            os.rmdir(parent)
                    except Exception:
                        pass

            self.db.update_sopd_attachment_path(record_id, None)
            if refresh_after:
                self.load_sopd_records()
            return True
        except Exception as e:
            if not silent:
                self.warn(str(e))
            return False

    def add_sopd_record(self):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            self.warn("Сначала выбери компанию.")
            return

        dlg = SopdDialog(self, "Добавить карточку СОПД")
        if dlg.exec() != QDialog.Accepted:
            return

        try:
            self.db.add_sopd_record(company_id, dlg.values())
            self.load_documents()
            self.set_content_mode("sopd")
        except Exception as e:
            self.warn(str(e))

    def edit_sopd_record(self, record_id: int):
        rec = self.db.get_sopd_record(record_id) if self.db else None
        if not rec:
            self.warn("Карточка СОПД не найдена.")
            return

        dlg = SopdDialog(self, "Изменить карточку СОПД", data={
            "consent_type": rec["consent_type"] or "",
            "purpose": rec["purpose"] or "",
            "pd_categories": rec["pd_categories"] or "",
            "pd_list": rec["pd_list"] or "",
            "processing_method": rec["processing_method"] or "",
            "third_party_transfer": rec["third_party_transfer"] or "Не указано",
            "transfer_to": rec["transfer_to"] or "",
            "sopd_description": rec["sopd_description"] or "",
            "validity_period": rec["validity_period"] or "",
        })
        if dlg.exec() != QDialog.Accepted:
            return

        try:
            self.db.update_sopd_record(record_id, dlg.values())
            self.load_documents()
        except Exception as e:
            self.warn(str(e))

    def delete_sopd_record(self, record_id: int):
        rec = self.db.get_sopd_record(record_id) if self.db else None
        if not rec:
            self.warn("Карточка СОПД не найдена.")
            return

        name = (rec["consent_type"] or "").strip() or "карточку"
        if not self.confirm(f"Удалить карточку СОПД «{name}»?"):
            return

        try:
            self.delete_sopd_record_file(record_id, ask_confirm=False, silent=True, refresh_after=False)
            self.db.delete_sopd_record(record_id)
            self.load_documents()
        except Exception as e:
            self.warn(str(e))

    def on_section_changed(self):
        it = self.section_list.currentItem()
        if not it:
            return
        self.current_section_id = it.data(Qt.UserRole)
        self.load_documents()

    def _try_hide_windows_dir(self, path: str):
        if os.name != "nt":
            return
        try:
            FILE_ATTRIBUTE_HIDDEN = 0x2
            FILE_ATTRIBUTE_SYSTEM = 0x4
            attrs = FILE_ATTRIBUTE_HIDDEN | FILE_ATTRIBUTE_SYSTEM
            ctypes.windll.kernel32.SetFileAttributesW(str(path), attrs)
        except Exception:
            pass

    def _abs_storage_path(self, p: Optional[str]) -> Optional[str]:
        if not p:
            return None
        if os.path.isabs(p) or p.startswith("\\\\"):
            return p
        return os.path.join(self.storage_dir, p)

    def _rel_storage_path(self, abs_path: str) -> str:
        return os.path.relpath(abs_path, self.storage_dir)
    
    def _title_from_file(self, file_path: str) -> str:
        base = os.path.basename(file_path)
        name = os.path.splitext(base)[0]
        return name.strip()
    
    def _safe_filename(self, name: str, max_len: int = 120) -> str:
        name = (name or "").strip()

        bad = '<>:"/\\|?*'
        for ch in bad:
            name = name.replace(ch, "_")

        name = name.rstrip(" .")

        if len(name) > max_len:
            name = name[:max_len].rstrip(" .")

        if not name:
            name = "file"
        return name

    def _is_auto_title(self, title: str) -> bool:
        t = (title or "").strip().lower()
        if not t:
            return True
        return t in ("новый документ", "документ", "без названия") or t.startswith("новый документ")

    def apply_doc_filter(self):
        q = self.doc_filter.text().strip().lower()
        visible = 0

        self.table.setUpdatesEnabled(False)
        try:
            for row_i in range(self.table.rowCount()):
                it = self.table.item(row_i, 1)
                title = it.text().lower() if it else ""
                st_it = self.table.item(row_i, 2)
                status = ""
                if st_it:
                    status = str(st_it.data(Qt.UserRole) or st_it.text() or "").lower()
                hide = bool(q) and (q not in title) and (q not in status)
                self.table.setRowHidden(row_i, hide)
                if not hide:
                    visible += 1
        finally:
            self.table.setUpdatesEnabled(True)
            self.table.viewport().update()

        if self.table.rowCount() > 0 and visible == 0 and q:
            self._show_docs_empty_search()
        else:
            self._hide_docs_empty_search()

    def _ensure_docs_empty_overlay(self):
        if hasattr(self, "_docs_empty_overlay"):
            return
        host = self.docs_page if hasattr(self, "docs_page") else self.right_full
        self._docs_empty_overlay = QFrame(host)
        self._docs_empty_overlay.setObjectName("emptyOverlay")
        self._docs_empty_overlay.hide()

        lay = QVBoxLayout(self._docs_empty_overlay)
        lay.setContentsMargins(22, 22, 22, 22)
        lay.setSpacing(8)

        t = QLabel("Ничего не найдено")
        t.setObjectName("emptyOverlayTitle")
        s = QLabel("Попробуй изменить запрос или очистить поиск.")
        s.setObjectName("emptyOverlayText")
        s.setWordWrap(True)

        lay.addStretch(1)
        lay.addWidget(t, 0, Qt.AlignCenter)
        lay.addWidget(s, 0, Qt.AlignCenter)
        lay.addStretch(1)

    def _show_docs_empty_search(self):
        self._ensure_docs_empty_overlay()
        self._docs_empty_overlay.setGeometry(self.table.geometry())
        self._docs_empty_overlay.show()
        self._fade_in(self._docs_empty_overlay, 140)

    def _hide_docs_empty_search(self):
        if hasattr(self, "_docs_empty_overlay"):
            self._docs_empty_overlay.hide()
    
    def rename_document_row(self, doc_id: int):
        current = ""
        for i in range(self.table.rowCount()):
            it = self.table.item(i, 0)
            if it and it.text().isdigit() and int(it.text()) == doc_id:
                t = self.table.item(i, 1)
                current = t.text() if t else ""
                break

        dlg = InputDialog(self, "Переименовать документ", "Новое название:", default=current)
        if dlg.exec() != QDialog.Accepted:
            return

        try:
            self.db.rename_document(doc_id, dlg.value())
            self.load_documents()
        except Exception as e:
            self.warn(str(e))

    def set_accept_date_dialog(self, doc_id: int):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return

        docs = self.db.list_documents(company_id)
        current = ""
        for r in docs:
            if int(r["id"]) == doc_id:
                current = r["accept_date"] or ""
                break

        dlg = InputDialog(
            self,
            "Дата принятия",
            "Введите дату (дд.мм.гггг или гггг-мм-дд):",
            default=fmt_date_ddmmyyyy(current) or ""
        )
        if dlg.exec() != QDialog.Accepted:
            return

        ymd = parse_date_to_ymd(dlg.value())
        if not ymd:
            self.warn("Не поняла дату 😕\nПример: 25.01.2026 или 2026-01-25")
            return

        try:
            self.db.set_accept_date(doc_id, ymd)
            self.load_documents()
        except Exception as e:
            self.warn(str(e))


    def clear_accept_date(self, doc_id: int):
        try:
            self.db.set_accept_date(doc_id, None)
            self.load_documents()
        except Exception as e:
            self.warn(str(e))

    def set_review_date_dialog(self, doc_id: int):
        row = None
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return

        docs = self.db.list_documents(company_id)
        current = ""
        for r in docs:
            if int(r["id"]) == doc_id:
                current = r["review_due"] or ""
                break

        dlg = InputDialog(self, "Дата пересмотра", "Введите дату (дд.мм.гггг или гггг-мм-дд):", default=fmt_date_ddmmyyyy(current) or "")
        if dlg.exec() != QDialog.Accepted:
            return

        ymd = parse_date_to_ymd(dlg.value())
        if not ymd:
            self.warn("Не поняла дату 😕\nПример: 25.01.2026 или 2026-01-25")
            return

        try:
            self.db.set_review_due(doc_id, ymd)
            self.load_documents()
            self.refresh_notifications()
        except Exception as e:
            self.warn(str(e))

    def clear_review_date(self, doc_id: int):
        try:
            self.db.set_review_due(doc_id, None)
            self.load_documents()
            self.refresh_notifications()
        except Exception as e:
            self.warn(str(e))

    def _pick_directory_dialog(self, title: str, start_dir: str = "") -> str:
        base_dir = (start_dir or self.shared_root or "").strip()
        if not base_dir:
            base_dir = (
                QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
                or os.path.expanduser("~")
            )

        dialog = QFileDialog(self, title, base_dir)
        dialog.setOption(QFileDialog.DontUseNativeDialog, True)
        dialog.setOption(QFileDialog.ShowDirsOnly, True)
        dialog.setFileMode(QFileDialog.Directory)
        dialog.setAcceptMode(QFileDialog.AcceptOpen)
        dialog.setLabelText(QFileDialog.Accept, "Выбрать папку")
        dialog.setLabelText(QFileDialog.Reject, "Отмена")
        dialog.setObjectName("prettyDialog")
        dialog.setAttribute(Qt.WA_StyledBackground, True)
        dialog.setStyleSheet(DIRECTORY_DIALOG_STYLESHEET)
        dialog.setMinimumSize(880, 560)
        dialog.resize(920, 600)
        if dialog.exec() != QDialog.Accepted:
            return ""
        selected = dialog.selectedFiles()
        return selected[0] if selected else ""

    def _load_or_pick_shared_root(self) -> str:
        last = None
        sp = settings_path()

        if os.path.exists(sp):
            try:
                with open(sp, "r", encoding="utf-8") as f:
                    last = json.load(f).get("shared_root")
            except Exception:
                pass

        if last and os.path.isdir(last):
            reg = self._read_registry(last)
            if reg:
                return last

        intro = WorkspaceSetupDialog(self)
        if intro.exec() != QDialog.Accepted:
            raise SystemExit("Не выбрана рабочая папка.")

        root = self._pick_directory_dialog("Выберите рабочую папку", last or "")
        if not root:
            raise SystemExit("Не выбрана рабочая папка.")

        reg = self._read_registry(root)
        if not reg:
            self._write_registry(root)
        try:
            with open(sp, "w", encoding="utf-8") as f:
                json.dump({"shared_root": root}, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        return root

    def change_shared_root(self):
        root = self._pick_directory_dialog("Выберите новую рабочую папку", self.shared_root)
        if not root:
            return

        try:
            test_dir = os.path.join(root, ".pd_registry_test")
            ensure_dir(test_dir)
            test_file = os.path.join(test_dir, f"test_{int(time.time())}.tmp")
            with open(test_file, "w", encoding="utf-8") as f:
                f.write("ok")
            os.remove(test_file)
        except Exception as e:
            self.warn(f"Нет прав на запись в эту папку:\n{e}")
            return

        new_db_path = os.path.join(root, "pd_docs.db")
        new_storage_dir = os.path.join(root, ".pd_docs")
        try:
            ensure_dir(new_storage_dir)
            self._try_hide_windows_dir(new_storage_dir)
            new_db = Db(new_db_path, self.current_account)
        except Exception as e:
            self.warn(f"Не удалось подключиться к новой папке:\n{e}")
            return

        try:
            with open(settings_path(), "w", encoding="utf-8") as f:
                json.dump({"shared_root": root}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.warn(f"Не получилось сохранить settings.json:\n{e}")
            try:
                new_db.close()
            except Exception:
                pass
            return

        try:
            if self.db:
                self.db.close()
            self.db = new_db
            self.shared_root = root
            self.db_path = new_db_path
            self.storage_dir = new_storage_dir

            self._due_reviews_count = 0
            self._update_statusbar()
            self.doc_filter.setText("")
            if hasattr(self, "clear_right_panel_state"):
                self.clear_right_panel_state(self._page_key, preferred_company_id=None)
            company_rows = [dict(row) for row in self.db.list_companies_with_stats()]
            if hasattr(self, "_startup_catalog_pending"):
                self._startup_catalog_rows = None
                self._document_catalog_cache = None
                self._startup_catalog_loading = False
                self._startup_catalog_pending = bool(company_rows)
            self.load_companies(
                initial=True,
                rows=company_rows,
                refresh_page=not bool(getattr(self, "_startup_catalog_pending", False)),
            )
            self.refresh_notifications()
            if bool(getattr(self, "_startup_catalog_pending", False)):
                self._show_catalog_loading_state()
                self._warm_initial_document_catalog()
            QTimer.singleShot(600, self.check_updates_and_offer)
            self.info("Папка изменена ✅")
        except Exception as e:
            self.warn(f"Не получилось переключиться:\n{e}")

    def _releases_dir(self) -> str:
        p = os.path.join(self.shared_root, "releases")
        ensure_dir(p)
        return p

    def _manifest_path(self) -> str:
        return os.path.join(self._releases_dir(), "manifest.json")

    def _read_update_manifest(self) -> Optional[dict]:
        mp = self._manifest_path()
        if not os.path.exists(mp):
            return None
        try:
            with open(mp, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def check_updates_and_offer(self):
        m = self._read_update_manifest()
        if not m:
            return

        new_ver = str(m.get("version", "")).strip()
        exe_name = str(m.get("exe", "")).strip()
        expected_sha = str(m.get("sha256", "")).strip().lower()

        if not new_ver or not exe_name or not expected_sha:
            return

        if parse_ver(new_ver) <= parse_ver(APP_VERSION):
            return

        src_exe = os.path.join(self._releases_dir(), exe_name)
        if not os.path.exists(src_exe):
            return

        try:
            real_sha = sha256_file(src_exe).lower()
        except Exception:
            self.warn("Не удалось прочитать файл обновления.")
            return

        if real_sha != expected_sha:
            self.warn("Обновление отклонено: не совпал SHA-256 (файл повреждён или подменён).")
            return

        if not self.confirm(f"Найдена новая версия {new_ver}. Обновить сейчас?"):
            return

        try:
            self._run_updater_apply(src_exe, new_ver)
        except Exception as e:
            self.warn(f"Не получилось запустить обновление:\n{e}")

    def _run_updater_apply(self, src_exe: str, new_ver: str):
        # staging (не temp)
        staging_dir = os.path.join(local_app_dir("PDDocsRegistry"), "staging")
        ensure_dir(staging_dir)
        staged = os.path.join(staging_dir, f"PD_Docs_Registry_{new_ver}.exe")
        atomic_copy(src_exe, staged)

        # ещё раз проверим, что staging не битый
        m = self._read_update_manifest()
        if not m:
            raise RuntimeError("manifest.json пропал")
        expected_sha = str(m.get("sha256", "")).strip().lower()
        if sha256_file(staged).lower() != expected_sha:
            raise RuntimeError("SHA-256 staging не совпал")

        target = sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
        upd = updater_exe_path()

        args = [upd, "--apply", staged, "--target", target, "--pid", str(os.getpid()), "--restart"]
        creationflags = 0
        if os.name == "nt":
            creationflags = subprocess.CREATE_NO_WINDOW

        subprocess.Popen(args, close_fds=True, creationflags=creationflags)

        QApplication.quit()       

    def bootstrap(self):
        try:
            self.shared_root = self._load_or_pick_shared_root()
            self.db_path = os.path.join(self.shared_root, "pd_docs.db")
            self.storage_dir = os.path.join(self.shared_root, ".pd_docs")
            ensure_dir(self.storage_dir)
            self._try_hide_windows_dir(self.storage_dir)


            self.db = Db(self.db_path, self.current_account)

            self._force_viewport_color(self.company_list, "#162437")
            self._force_viewport_color(self.table, "#131f31")

            self._due_reviews_count = 0
            self._update_statusbar()
            self.load_companies(initial=True)
            self.refresh_notifications()
        except Exception as e:
            self.warn(f"Не удалось запуститься:\n{e}")

    def _center_cell_widget(self, w: QWidget) -> QWidget:
        host = QWidget()
        host.setObjectName("cellHost")
        host.setStyleSheet("background: transparent;")

        host.setAutoFillBackground(False)
        host.setAttribute(Qt.WA_TranslucentBackground, True)
        host.setAttribute(Qt.WA_NoSystemBackground, True)
        host.setAttribute(Qt.WA_StyledBackground, True)

        lay = QHBoxLayout(host)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        lay.addStretch(1)
        lay.addWidget(w, 0, Qt.AlignCenter)
        lay.addStretch(1)

        return host

    def _fade_in(self, widget: QWidget, duration_ms: int = 170):
        eff = QGraphicsOpacityEffect(widget)
        widget.setGraphicsEffect(eff)
        eff.setOpacity(0.0)

        anim = QPropertyAnimation(eff, b"opacity", widget)
        anim.setDuration(duration_ms)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        anim.start(QPropertyAnimation.DeleteWhenStopped)

        def cleanup():
            widget.setGraphicsEffect(None)

        anim.finished.connect(cleanup)

    def switch_app_page(self, index: int):
        if self.app_stack.currentIndex() == index:
            return
        self.app_stack.setCurrentIndex(index)
        self._fade_in(self.app_stack.currentWidget(), 200)

    def _update_statusbar(self):
        base = f"Аккаунт: {self.current_account}   |   База: {self.shared_root}"
        if self._due_reviews_count > 0:
            base += f"   |   На пересмотр: {self._due_reviews_count}"
        if self._busy_jobs > 0 and self._busy_label:
            base += f"   |   {self._busy_label}..."
        self.statusBar().showMessage(base)

    def _set_busy_state(self, is_busy: bool, label: str = ""):
        if is_busy:
            self._busy_jobs += 1
            if label:
                self._busy_label = label
            if not self._busy_cursor_set:
                QApplication.setOverrideCursor(Qt.WaitCursor)
                self._busy_cursor_set = True
            self._update_statusbar()
            return

        if self._busy_jobs > 0:
            self._busy_jobs -= 1
        if self._busy_jobs == 0:
            self._busy_label = ""
            if self._busy_cursor_set:
                QApplication.restoreOverrideCursor()
                self._busy_cursor_set = False
        self._update_statusbar()

    def _run_background_task(
        self,
        label: str,
        work_fn: Callable[[], Any],
        on_success: Callable[[Any], None],
        on_error_prefix: str,
        show_busy_state: bool = True,
        reject_if_busy: bool = True,
        on_failure: Optional[Callable[[str], None]] = None,
    ):
        if reject_if_busy and self._busy_jobs > 0:
            self.warn("Дождитесь завершения текущей операции.")
            return

        task = BackgroundTask(work_fn)
        self._bg_tasks.add(task)
        if show_busy_state:
            self._set_busy_state(True, label)

        def cleanup():
            self._bg_tasks.discard(task)
            if show_busy_state:
                self._set_busy_state(False)

        def ok(result: Any):
            cleanup()
            try:
                on_success(result)
            except Exception as e:
                self.warn(str(e))

        def fail(msg: str):
            cleanup()
            if on_failure is not None:
                on_failure(msg)
            else:
                self.warn(f"{on_error_prefix}\n{msg}")

        task.signals.finished.connect(ok)
        task.signals.failed.connect(fail)
        self._thread_pool.start(task)

    # ---------- message helpers ----------
    def info(self, text: str):
        ThemedMessageDialog(self, "Готово", text, kind="info", buttons=("ok",)).exec()

    def warn(self, text: str):
        ThemedMessageDialog(self, "Внимание", text, kind="warn", buttons=("ok",)).exec()

    def confirm(self, text: str) -> bool:
        dlg = ThemedMessageDialog(self, "Подтверждение", text, kind="ask", buttons=("yes", "no"))
        dlg.exec()
        return dlg.yes()

    # ---------- storage ----------
    def current_company_id_name(self):
        item = self.company_list.currentItem() if hasattr(self, "company_list") else None
        if not item:
            return None, None
        company_name = item.data(Qt.UserRole + 1)
        if not company_name:
            company_name = item.text()
        return item.data(Qt.UserRole), str(company_name)

    def _company_dir(self, company_id: int) -> str:
        path = os.path.join(self.storage_dir, f"company_{company_id}")
        ensure_dir(path)
        return path

    def _doc_dir(self, company_id: int, doc_id: int) -> str:
        path = os.path.join(self._company_dir(company_id), f"doc_{doc_id}")
        ensure_dir(path)
        return path

    def _sopd_dir(self, company_id: int) -> str:
        path = os.path.join(self._company_dir(company_id), "_sopd")
        ensure_dir(path)
        return path

    def _sopd_record_dir(self, company_id: int, record_id: int) -> str:
        path = os.path.join(self._sopd_dir(company_id), f"record_{record_id}")
        ensure_dir(path)
        return path

    def _doc_lock(self, company_id: int, doc_id: int) -> FileLock:
        lock_path = os.path.join(self._doc_dir(company_id, doc_id), ".lock")
        return FileLock(lock_path, owner=self.current_account, stale_seconds=180)

    def _build_empty_app_page(self) -> QWidget:
        root = QWidget()
        root.setObjectName("emptyRoot")

        lay = QVBoxLayout(root)
        lay.setContentsMargins(26, 26, 26, 26)

        card = QFrame()
        card.setObjectName("emptyCard")
        card_w = 760
        card.setFixedWidth(card_w)

        c = QVBoxLayout(card)
        c.setContentsMargins(22, 22, 22, 20)
        c.setSpacing(12)

        title = QLabel("Тут пока пусто ✨")
        title.setObjectName("emptyTitle")

        subtitle = QLabel(
            "Создай первую компанию — и дальше удобно веди общий реестр документов."
        )
        subtitle.setObjectName("emptyText")
        subtitle.setWordWrap(True)
        subtitle.setTextFormat(Qt.PlainText)
        subtitle.setFixedWidth(card_w - 44)
        subtitle.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)

        btn = QToolButton()
        btn.setText("Добавить компанию")
        btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogNewFolder))
        btn.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        btn.setProperty("kind", "solid")
        btn.clicked.connect(self.add_company)

        c.addWidget(title)
        c.addWidget(subtitle)
        c.addSpacing(4)
        c.addWidget(btn, 0, Qt.AlignLeft)

        lay.addWidget(card, 0, Qt.AlignCenter)
        return root

    def _build_right_before_first_doc(self) -> QWidget:
        page = QWidget()
        page_lay = QVBoxLayout(page)
        page_lay.setContentsMargins(0, 0, 0, 0)

        card = QFrame()
        card.setObjectName("emptyCardSmall")
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        card.setMinimumHeight(520)

        c = QVBoxLayout(card)
        c.setContentsMargins(30, 28, 30, 26)
        c.setSpacing(16)

        top = QHBoxLayout()
        top.setSpacing(14)

        ico = QLabel()
        ico.setObjectName("heroIcon")
        ico.setFixedSize(48, 48)
        ico.setAlignment(Qt.AlignCenter)
        ico.setPixmap(self.style().standardIcon(QStyle.SP_FileDialogNewFolder).pixmap(24, 24))
        top.addWidget(ico, 0, Qt.AlignTop)

        text_col = QVBoxLayout()
        text_col.setSpacing(6)

        t = QLabel("Документы этой компании")
        t.setObjectName("heroTitle")
        t.setWordWrap(True)

        s = QLabel(
            "Пока тут пусто.\n"
            "Добавь первый документ — после этого появятся поиск и таблица.\n"
            "Можно сразу перейти в раздел СОПД и завести карточки согласий."
        )
        s.setObjectName("heroText")
        s.setWordWrap(True)
        s.setTextFormat(Qt.PlainText)

        text_col.addWidget(t)
        text_col.addWidget(s)
        top.addLayout(text_col, 1)

        c.addLayout(top)

        chips = QHBoxLayout()
        chips.setSpacing(10)

        def chip(title: str, subtitle: str, icon: QStyle.StandardPixmap) -> QFrame:
            f = QFrame()
            f.setObjectName("heroChip")
            f.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            f.setMinimumHeight(70)
            l = QHBoxLayout(f)
            l.setContentsMargins(12, 10, 12, 10)
            l.setSpacing(10)

            ic = QLabel()
            ic.setObjectName("heroChipIcon")
            ic.setFixedSize(34, 34)
            ic.setAlignment(Qt.AlignCenter)
            ic.setPixmap(self.style().standardIcon(icon).pixmap(18, 18))

            txt = QVBoxLayout()
            txt.setSpacing(2)

            a = QLabel(title)
            a.setObjectName("heroChipTitle")

            b = QLabel(subtitle)
            b.setObjectName("heroChipText")
            b.setWordWrap(True)

            txt.addWidget(a)
            txt.addWidget(b)

            l.addWidget(ic)
            l.addLayout(txt, 1)
            return f

        chips.addWidget(chip("PDF", "Скан документа", QStyle.SP_FileIcon))
        chips.addWidget(chip("Office", "Word / Excel", QStyle.SP_DirIcon))
        chips.addWidget(chip("Drag & Drop", "Перетяни файл", QStyle.SP_ArrowDown))

        c.addLayout(chips)
        c.addStretch(1)

        cta_row = QHBoxLayout()
        cta_row.setSpacing(10)

        btn = QToolButton()
        btn.setObjectName("heroCta")
        btn.setText("Добавить документ")
        btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogNewFolder))
        btn.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        btn.setProperty("kind", "solid")
        btn.setMinimumHeight(52)
        btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn.clicked.connect(self.add_document)
        cta_row.addWidget(btn, 1)

        sopd_btn = QToolButton()
        sopd_btn.setText("Добавить СОПД")
        sopd_btn.setProperty("kind", "ghost")
        sopd_btn.setMinimumHeight(52)
        sopd_btn.clicked.connect(lambda: (self.set_content_mode("sopd"), self.add_sopd_record()))
        cta_row.addWidget(sopd_btn, 0)

        c.addLayout(cta_row)

        page_lay.addWidget(card, 1)
        return page

    def _apply_company_filter(self, preferred_company_id: Optional[int] = None):
        if not self.db:
            return

        current_item = self.company_list.currentItem()
        current_id = preferred_company_id
        if current_id is None and current_item:
            current_id = current_item.data(Qt.UserRole)

        self.company_list.blockSignals(True)
        self.company_list.clear()

        rows = self.db.list_companies_with_stats()
        selected_row = -1

        for r in rows:
            company_id = int(r["id"])
            name = str(r["name"])
            docs_count = int(r["docs_count"] or 0)
            sections_count = int(r["sections_count"] or 0)

            text = self._company_item_text(name, docs_count, sections_count)
            it = QListWidgetItem(text)
            it.setData(Qt.UserRole, company_id)
            it.setData(Qt.UserRole + 1, name)
            it.setToolTip(f"{name}\nДокументов: {docs_count}\nРазделов: {sections_count}")
            self.company_list.addItem(it)

            if company_id == current_id:
                selected_row = self.company_list.count() - 1

        self.company_list.blockSignals(False)

        total = len(rows)
        self.lbl_left.setText("Компании")
        self.company_count.setText(str(total))

        has_items = self.company_list.count() > 0
        if hasattr(self, "quick_add_section_btn"):
            self.quick_add_section_btn.setEnabled(has_items)
        if hasattr(self, "add_sopd_btn"):
            self.add_sopd_btn.setEnabled(has_items)
        if has_items:
            if selected_row < 0:
                selected_row = 0
            self.company_list.setCurrentRow(selected_row)
        else:
            self.lbl_right.setText("Документы и СОПД")
            self.right_stack.setCurrentIndex(0)
            self._sections_company_id = None
            self.section_list.clear()
            self.current_section_id = None
            self.sections_title.setText("Разделы")
            self._render_sopd_cards([])
            if hasattr(self, "sopd_file_chip"):
                self.sopd_file_chip.refresh_visual()

    def _build_right_full(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self.right_content_stack = QStackedWidget()
        self.right_content_stack.setObjectName("rightContentStack")
        self.docs_page = self._build_docs_page()
        self.sopd_page = self._build_sopd_page()
        self.right_content_stack.addWidget(self.docs_page)  # 0
        self.right_content_stack.addWidget(self.sopd_page)  # 1
        lay.addWidget(self.right_content_stack, 1)
        return w

    def _build_docs_page(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(10)

        self.doc_filter = QLineEdit()
        self.doc_filter.setPlaceholderText("Поиск по названию…")
        self.doc_filter.textChanged.connect(lambda: self._doc_search_timer.start(140))
        top.addWidget(self.doc_filter, 1)

        self.add_doc_btn = QToolButton()
        self.add_doc_btn.setText("Добавить документ")
        self.add_doc_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogNewFolder))
        self.add_doc_btn.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.add_doc_btn.setProperty("kind", "solid")
        self.add_doc_btn.clicked.connect(self.add_document)
        top.addWidget(self.add_doc_btn)
        lay.addLayout(top)

        self.table = DocumentTable()
        self.table.set_main(self)
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "ID",
            "Название",
            "\u00A0\u00A0Статус",
            "\u00A0\u00A0PDF",
            "\u00A0\u00A0Office",
            "\u00A0\u00A0Пересмотр",
            "\u00A0\u00A0Принято",
        ])
        self.table.verticalHeader().setVisible(False)
        self.table.setColumnHidden(0, True)
        self.table.setAlternatingRowColors(True)

        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.Fixed)
        hh.setSectionResizeMode(4, QHeaderView.Fixed)
        hh.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.setColumnWidth(2, 150)
        self.table.setColumnWidth(3, 132)
        self.table.setColumnWidth(4, 148)

        lay.addWidget(self.table, 1)
        return w

    def _build_sopd_page(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        self.sopd_panel = QFrame()
        self.sopd_panel.setObjectName("sopdPanel")
        panel_l = QVBoxLayout(self.sopd_panel)
        panel_l.setContentsMargins(12, 10, 12, 10)
        panel_l.setSpacing(8)

        panel_head = QHBoxLayout()
        panel_head.setSpacing(8)

        panel_title = QLabel("СОПД")
        panel_title.setObjectName("h2")
        panel_head.addWidget(panel_title)

        self.sopd_count_label = QLabel("0")
        self.sopd_count_label.setObjectName("muted")
        panel_head.addWidget(self.sopd_count_label, 0, Qt.AlignVCenter)
        panel_head.addStretch(1)

        self.add_sopd_btn = QToolButton()
        self.add_sopd_btn.setText("Добавить карточку")
        self.add_sopd_btn.setProperty("kind", "solid")
        self.add_sopd_btn.clicked.connect(self.add_sopd_record)
        panel_head.addWidget(self.add_sopd_btn)
        panel_l.addLayout(panel_head)

        self.sopd_scroll = QScrollArea()
        self.sopd_scroll.setWidgetResizable(True)
        self.sopd_scroll.setFrameShape(QFrame.NoFrame)
        self.sopd_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        self.sopd_cards_host = QWidget()
        self.sopd_cards_layout = QVBoxLayout(self.sopd_cards_host)
        self.sopd_cards_layout.setContentsMargins(0, 0, 0, 0)
        self.sopd_cards_layout.setSpacing(8)
        self.sopd_scroll.setWidget(self.sopd_cards_host)
        panel_l.addWidget(self.sopd_scroll, 1)

        lay.addWidget(self.sopd_panel, 1)
        return w

    def _sync_content_mode_buttons(self):
        if not hasattr(self, "mode_docs_btn") or not hasattr(self, "mode_sopd_btn"):
            return
        on_docs = (self._content_mode == "docs")
        self.mode_docs_btn.setProperty("active", on_docs)
        self.mode_sopd_btn.setProperty("active", not on_docs)
        self.mode_docs_btn.style().unpolish(self.mode_docs_btn)
        self.mode_docs_btn.style().polish(self.mode_docs_btn)
        self.mode_sopd_btn.style().unpolish(self.mode_sopd_btn)
        self.mode_sopd_btn.style().polish(self.mode_sopd_btn)
        if hasattr(self, "section_list"):
            self.section_list.setEnabled(on_docs)

    def set_content_mode(self, mode: str):
        mode = (mode or "").strip().lower()
        if mode not in {"docs", "sopd"}:
            mode = "docs"
        self._content_mode = mode
        if hasattr(self, "right_content_stack"):
            self.right_content_stack.setCurrentIndex(0 if mode == "docs" else 1)
        if mode == "docs":
            if hasattr(self, "table"):
                self.apply_doc_filter()
        else:
            self._hide_docs_empty_search()
        self._sync_content_mode_buttons()

    def _build_main_page(self) -> QWidget:
        root = QWidget()
        root.setObjectName("centralWidget")

        main = QVBoxLayout(root)
        main.setContentsMargins(16, 16, 16, 16)
        main.setSpacing(12)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        left_header = QWidget()
        left_header.setFixedWidth(LEFT_W)
        self.left_header = left_header
        lh = QHBoxLayout(left_header)
        lh.setContentsMargins(0, 0, 0, 0)
        lh.setSpacing(0)
        lh.addStretch(1)

        right_header = QWidget()
        rh = QHBoxLayout(right_header)
        rh.setContentsMargins(0, 0, 0, 0)
        rh.setSpacing(8)

        self.lbl_right = QLabel("Документы и СОПД")
        self.lbl_right.setObjectName("h1")
        rh.addWidget(self.lbl_right, 1)

        self.mode_docs_btn = QToolButton()
        self.mode_docs_btn.setText("Документы")
        self.mode_docs_btn.setProperty("modeToggle", True)
        self.mode_docs_btn.clicked.connect(lambda: self.set_content_mode("docs"))
        rh.addWidget(self.mode_docs_btn, 0, Qt.AlignRight)

        self.mode_sopd_btn = QToolButton()
        self.mode_sopd_btn.setText("СОПД")
        self.mode_sopd_btn.setProperty("modeToggle", True)
        self.mode_sopd_btn.clicked.connect(lambda: self.set_content_mode("sopd"))
        rh.addWidget(self.mode_sopd_btn, 0, Qt.AlignRight)

        header_row.addWidget(left_header)
        header_row.addWidget(right_header, 1)

        main.addLayout(header_row)

        body = QHBoxLayout()
        body.setSpacing(14)

        left = QWidget()
        left.setObjectName("leftPanel")
        self.left_panel = left
        left.setFixedWidth(LEFT_W)
        left_l = QVBoxLayout(left)
        left_l.setContentsMargins(0, 0, 0, 0)
        left_l.setSpacing(0)

        nav_shell = QFrame()
        nav_shell.setObjectName("sideShell")
        shell_l = QVBoxLayout(nav_shell)
        shell_l.setContentsMargins(12, 12, 12, 12)
        shell_l.setSpacing(12)

        company_block = QFrame()
        company_block.setObjectName("sideCard")
        company_l = QVBoxLayout(company_block)
        company_l.setContentsMargins(10, 10, 10, 10)
        company_l.setSpacing(8)

        company_head = QHBoxLayout()
        company_head.setSpacing(8)

        self.lbl_left = QLabel("Компании")
        self.lbl_left.setObjectName("h2")
        company_head.addWidget(self.lbl_left)
        company_head.addStretch(1)

        self.company_count = QLabel("")
        self.company_count.setObjectName("muted")
        company_head.addWidget(self.company_count, 0, Qt.AlignRight)

        self.quick_add_btn = QToolButton()
        self.quick_add_btn.setObjectName("quickAddMain")
        self.quick_add_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogNewFolder))
        self.quick_add_btn.setToolButtonStyle(Qt.ToolButtonIconOnly)
        self.quick_add_btn.setPopupMode(QToolButton.InstantPopup)
        self.quick_add_btn.setProperty("kind", "icon")
        self.quick_add_btn.setToolTip("Создать")
        company_head.addWidget(self.quick_add_btn, 0, Qt.AlignRight)
        company_l.addLayout(company_head)

        self.company_list = QListWidget()
        self.company_list.setObjectName("navList")
        self.company_list.setFocusPolicy(Qt.NoFocus)
        self.company_list.setIconSize(QSize(0, 0))
        self.company_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.company_list.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.company_list.setSpacing(6)
        self.company_list.setUniformItemSizes(True)
        self.company_list.itemSelectionChanged.connect(self.on_company_changed)
        self.company_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.company_list.customContextMenuRequested.connect(self.company_context_menu)
        company_l.addWidget(self.company_list, 1)
        shell_l.addWidget(company_block, 1)

        section_block = QFrame()
        section_block.setObjectName("sideCard")
        section_l = QVBoxLayout(section_block)
        section_l.setContentsMargins(10, 10, 10, 10)
        section_l.setSpacing(8)

        section_head = QHBoxLayout()
        section_head.setSpacing(8)

        self.sections_title = QLabel("Разделы")
        self.sections_title.setObjectName("h2")
        section_head.addWidget(self.sections_title)
        section_head.addStretch(1)

        self.quick_add_section_btn = QToolButton()
        self.quick_add_section_btn.setObjectName("quickAddSectionBtn")
        self.quick_add_section_btn.setProperty("kind", "icon")
        self.quick_add_section_btn.setToolButtonStyle(Qt.ToolButtonIconOnly)
        self.quick_add_section_btn.setToolTip("Добавить раздел")
        self.quick_add_section_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogNewFolder))
        self.quick_add_section_btn.clicked.connect(self.add_section)
        section_head.addWidget(self.quick_add_section_btn, 0, Qt.AlignRight)
        section_l.addLayout(section_head)

        self.section_list = QListWidget()
        self.section_list.setObjectName("navList")
        self.section_list.setFocusPolicy(Qt.NoFocus)
        self.section_list.setIconSize(QSize(0, 0))
        self.section_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.section_list.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.section_list.setSpacing(6)
        self.section_list.setUniformItemSizes(True)
        self.section_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.section_list.customContextMenuRequested.connect(self.section_context_menu)
        self.section_list.itemSelectionChanged.connect(self.on_section_changed)
        section_l.addWidget(self.section_list, 1)
        shell_l.addWidget(section_block, 1)

        add_menu = QMenu(self.quick_add_btn)
        self.quick_add_company_action = add_menu.addAction("Добавить компанию")
        self.quick_add_section_action = add_menu.addAction("Добавить раздел")
        self.quick_add_sopd_action = add_menu.addAction("Добавить СОПД")
        self.quick_add_company_action.triggered.connect(self.add_company)
        self.quick_add_section_action.triggered.connect(self.add_section)
        self.quick_add_sopd_action.triggered.connect(lambda: (self.set_content_mode("sopd"), self.add_sopd_record()))
        add_menu.aboutToShow.connect(self._refresh_quick_add_menu)
        self.quick_add_btn.setMenu(add_menu)

        left_l.addWidget(nav_shell, 1)

        self.right_stack = QStackedWidget()
        self.right_stack.setObjectName("rightStack")

        self.right_before = self._build_right_before_first_doc()
        self.right_full = self._build_right_full()

        self.right_stack.addWidget(self.right_before)  # 0
        self.right_stack.addWidget(self.right_full)    # 1

        body.addWidget(left)
        body.addWidget(self.right_stack, 1)

        main.addLayout(body, 1)
        self.set_content_mode("docs")
        self._sync_sidebar_toggle_ui()
        return root

    def _sync_sidebar_toggle_ui(self):
        if hasattr(self, "_act_toggle_sidebar"):
            if self._sidebar_visible:
                self._act_toggle_sidebar.setText("Скрыть боковую панель")
            else:
                self._act_toggle_sidebar.setText("Показать боковую панель")

    def set_sidebar_visible(self, visible: bool):
        self._sidebar_visible = bool(visible)
        if hasattr(self, "left_panel"):
            self.left_panel.setVisible(self._sidebar_visible)
        if hasattr(self, "left_header"):
            self.left_header.setVisible(self._sidebar_visible)
        self._sync_sidebar_toggle_ui()

    def toggle_sidebar(self):
        self.set_sidebar_visible(not self._sidebar_visible)

    def _refresh_quick_add_menu(self):
        if not hasattr(self, "quick_add_section_action"):
            return
        company_id, _ = self.current_company_id_name()
        can_add_context_items = bool(company_id)
        self.quick_add_section_action.setEnabled(can_add_context_items)
        if hasattr(self, "quick_add_sopd_action"):
            self.quick_add_sopd_action.setEnabled(can_add_context_items)
        if hasattr(self, "quick_add_section_btn"):
            self.quick_add_section_btn.setEnabled(can_add_context_items)
        if hasattr(self, "add_sopd_btn"):
            self.add_sopd_btn.setEnabled(can_add_context_items)

    def _company_item_text(self, name: str, docs_count: int, sections_count: int) -> str:
        return f"{name}  ·  {docs_count} док. / {sections_count} разд."

    def _section_item_text(self, name: str, docs_count: int) -> str:
        return f"{name}  ·  {docs_count} док."

    def load_sections(self, company_id: int, preferred_section_id=KEEP_SELECTION):
        self._sections_company_id = company_id
        self._section_rows = self.db.list_sections_with_doc_counts(company_id)
        self._section_total_docs = self.db.count_documents_by_company(company_id)
        self._section_without_docs = self.db.count_documents_without_section(company_id)
        self._apply_section_filter(preferred_section_id)

    def _apply_section_filter(self, preferred_section_id=KEEP_SELECTION):
        company_id = getattr(self, "_sections_company_id", None)
        if not company_id:
            return

        rows = getattr(self, "_section_rows", []) or []
        total_docs = int(getattr(self, "_section_total_docs", 0) or 0)
        without_docs = int(getattr(self, "_section_without_docs", 0) or 0)

        self.section_list.blockSignals(True)
        self.section_list.clear()

        it_all = QListWidgetItem(self._section_item_text("Все документы", total_docs))
        it_all.setData(Qt.UserRole, None)
        it_all.setData(Qt.UserRole + 1, "Все документы")
        it_all.setToolTip("Показать все документы компании")
        self.section_list.addItem(it_all)

        it_none = QListWidgetItem(self._section_item_text("Без раздела", without_docs))
        it_none.setData(Qt.UserRole, -1)
        it_none.setData(Qt.UserRole + 1, "Без раздела")
        it_none.setToolTip("Документы без привязки к разделам")
        self.section_list.addItem(it_none)

        total_sections = len(rows)
        selected_id = self.current_section_id if preferred_section_id is KEEP_SELECTION else preferred_section_id

        for r in rows:
            sid = int(r["id"])
            name = str(r["name"])
            docs_count = int(r["docs_count"] or 0)

            it = QListWidgetItem(self._section_item_text(name, docs_count))
            it.setData(Qt.UserRole, sid)
            it.setData(Qt.UserRole + 1, name)
            it.setToolTip(f"{name}\nДокументов: {docs_count}")
            self.section_list.addItem(it)

        self.sections_title.setText(f"Разделы · {total_sections}")

        target_row = 0
        target_section_id = selected_id
        if target_section_id == -1:
            target_row = 1
        elif target_section_id is not None:
            for i in range(self.section_list.count()):
                if self.section_list.item(i).data(Qt.UserRole) == target_section_id:
                    target_row = i
                    break

        if self.section_list.count() > 0:
            self.section_list.setCurrentRow(target_row)
            current_item = self.section_list.currentItem()
            self.current_section_id = current_item.data(Qt.UserRole) if current_item else None
        else:
            self.current_section_id = None

        self.section_list.blockSignals(False)
        self.load_documents()

    def _apply_style(self):
        self.setStyleSheet("""
            QWidget {
                font-size: 14px;
                color: #e8f0fb;
            }
            QMainWindow {
                background: #0f1621;
            }
            QWidget#centralWidget, QWidget#emptyRoot {
                background: transparent;
            }

            QStackedWidget#appStack, QStackedWidget#rightStack, QStackedWidget#rightContentStack {
                background: transparent;
                border: none;
            }

            QLabel#h1 {
                font-size: 19px;
                font-weight: 820;
                color: #f4f8ff;
            }
            QLabel#h2 {
                font-size: 16px;
                font-weight: 760;
                color: #eef6ff;
            }
            QLabel#muted {
                color: #8ba3be;
                font-size: 12px;
                padding-left: 2px;
            }

            QFrame#sideShell {
                background: #111b2a;
                border: 1px solid #263a55;
                border-radius: 18px;
            }
            QFrame#sideCard {
                background: #162437;
                border: 1px solid #2f4768;
                border-radius: 14px;
            }

            QLineEdit {
                background: #18283d;
                border: 1px solid #355173;
                border-radius: 10px;
                padding: 9px 11px;
                selection-background-color: rgba(51, 188, 181, 0.42);
            }
            QLineEdit:focus {
                border: 1px solid #44b9b4;
                background: #1b2d45;
            }
            QPlainTextEdit, QComboBox {
                background: #18283d;
                border: 1px solid #355173;
                border-radius: 10px;
                padding: 8px 10px;
                selection-background-color: rgba(51, 188, 181, 0.42);
            }
            QPlainTextEdit:focus, QComboBox:focus {
                border: 1px solid #44b9b4;
                background: #1b2d45;
            }
            QComboBox::drop-down {
                border: none;
                width: 0px;
            }
            QComboBox::down-arrow {
                image: none;
                width: 0px;
                height: 0px;
            }

            QListWidget#navList {
                background: transparent;
                border: none;
                outline: 0;
                padding: 0px;
            }
            QListWidget#navList::item {
                padding: 8px 10px;
                border-radius: 9px;
                color: #edf3ff;
            }
            QListWidget#navList::item:hover {
                background: rgba(255,255,255,0.07);
            }
            QListWidget#navList::item:selected,
            QListWidget#navList::item:selected:active,
            QListWidget#navList::item:selected:!active {
                background: rgba(51, 188, 181, 0.28);
                color: #f5fbff;
            }

            QListView::viewport, QAbstractItemView::viewport, QAbstractScrollArea::viewport {
                background: transparent;
                border: none;
            }

            QTableWidget {
                background: #131f31;
                border: 1px solid #2e4666;
                border-radius: 14px;
                padding: 0px;
                alternate-background-color: rgba(255,255,255,0.03);
            }
            QTableWidget::viewport {
                background: #131f31;
                border-radius: 14px;
            }
            QTableWidget::item {
                padding: 8px;
                border: none;
            }
            QTableWidget::item:selected {
                background: rgba(51, 188, 181, 0.20);
            }
            QTableView::item:hover {
                background: rgba(255,255,255,0.04);
            }

            QHeaderView::section {
                background: #101a2a;
                color: #dae8f8;
                border: none;
                padding: 10px 12px;
                font-weight: 820;
            }

            QWidget#cellHost {
                background: transparent;
                border: none;
            }

            QToolButton {
                border-radius: 11px;
                padding: 7px 10px;
                font-weight: 740;
                color: #edf5ff;
            }
            QToolButton:disabled {
                color: rgba(237,243,255,0.45);
                border-color: rgba(255,255,255,0.08);
            }

            QToolButton[kind="solid"] {
                background: #219f9c;
                border: 1px solid #34bcb8;
            }
            QToolButton[kind="solid"]:hover {
                background: #27b0ac;
            }
            QToolButton[kind="solid"]:pressed {
                background: #1b8784;
            }

            QToolButton[kind="ghost"] {
                background: rgba(255,255,255,0.05);
                border: 1px solid rgba(255,255,255,0.14);
            }
            QToolButton[kind="ghost"]:hover {
                background: rgba(255,255,255,0.09);
            }

            QToolButton[kind="icon"] {
                background: transparent;
                border: 1px solid transparent;
                padding: 6px;
                min-width: 30px;
                min-height: 30px;
            }
            QToolButton[kind="icon"]:hover {
                background: rgba(255,255,255,0.08);
                border: 1px solid rgba(255,255,255,0.18);
            }
            QToolButton[kind="icon"]:pressed {
                background: rgba(255,255,255,0.05);
            }
            QToolButton[modeToggle="true"] {
                background: rgba(255,255,255,0.04);
                border: 1px solid rgba(255,255,255,0.14);
                border-radius: 10px;
                padding: 7px 12px;
                min-width: 96px;
            }
            QToolButton[modeToggle="true"]:hover {
                background: rgba(255,255,255,0.09);
            }
            QToolButton[modeToggle="true"][active="true"] {
                background: rgba(33, 159, 156, 0.28);
                border: 1px solid rgba(52, 188, 184, 0.72);
                color: #f4fbff;
            }
            QToolButton#quickAddMain {
                background: rgba(255,255,255,0.04);
                border: 1px solid rgba(255,255,255,0.14);
                border-radius: 10px;
                min-width: 30px;
                min-height: 30px;
                max-width: 30px;
                max-height: 30px;
                padding: 4px;
            }
            QToolButton#quickAddMain:hover {
                background: rgba(255,255,255,0.10);
                border: 1px solid rgba(255,255,255,0.22);
            }
            QToolButton#quickAddMain:pressed {
                background: rgba(255,255,255,0.07);
            }

            QFrame#emptyCard, QFrame#emptyCardSmall {
                background: #111b2a;
                border: 1px solid #2a3f5c;
                border-radius: 18px;
            }
            QFrame#sopdPanel {
                background: #111b2a;
                border: 1px solid #2d435f;
                border-radius: 14px;
            }
            QFrame#sopdCard {
                background: #162437;
                border: 1px solid #314b6b;
                border-radius: 12px;
            }
            QLabel#sopdCardTitle {
                font-size: 15px;
                font-weight: 860;
                color: #f4f8ff;
            }
            QLabel#sopdLine {
                color: #d9e6f6;
                font-size: 13px;
            }
            QLabel#sopdProgress {
                padding: 3px 8px;
                border-radius: 10px;
                font-size: 12px;
                font-weight: 760;
                background: rgba(148, 163, 184, 0.16);
                border: 1px solid rgba(148, 163, 184, 0.40);
                color: #f2f7ff;
            }
            QLabel#sopdProgress[state="ok"] {
                background: rgba(34, 197, 94, 0.16);
                border: 1px solid rgba(34, 197, 94, 0.45);
            }
            QLabel#sopdProgress[state="mid"] {
                background: rgba(245, 158, 11, 0.16);
                border: 1px solid rgba(245, 158, 11, 0.48);
            }
            QLabel#sopdProgress[state="bad"] {
                background: rgba(239, 68, 68, 0.15);
                border: 1px solid rgba(239, 68, 68, 0.45);
            }
            QLabel#sopdMissing {
                color: rgba(255, 210, 190, 0.88);
                font-size: 12px;
                padding: 1px 1px 2px 1px;
            }
            QLabel#sopdFileLabel {
                color: rgba(232,240,252,0.75);
                font-size: 12px;
            }
            QLabel#sopdHint {
                color: rgba(232,240,252,0.68);
                font-size: 13px;
                padding: 6px 2px;
            }
            QLabel[sopdBadge="true"] {
                padding: 3px 8px;
                border-radius: 10px;
                font-size: 12px;
                font-weight: 760;
                background: rgba(255,255,255,0.08);
                color: #f2f7ff;
                border: 1px solid rgba(255,255,255,0.15);
            }
            QLabel[sopdBadge="true"][state="yes"] {
                background: rgba(251, 146, 60, 0.18);
                border: 1px solid rgba(251, 146, 60, 0.50);
            }
            QLabel[sopdBadge="true"][state="no"] {
                background: rgba(34, 197, 94, 0.16);
                border: 1px solid rgba(34, 197, 94, 0.45);
            }
            QLabel[sopdBadge="true"][state="na"] {
                background: rgba(148, 163, 184, 0.16);
                border: 1px solid rgba(148, 163, 184, 0.40);
            }
            QLabel#emptyTitle {
                font-size: 20px;
                font-weight: 900;
            }
            QLabel#emptyText {
                color: rgba(232,240,252,0.72);
            }

            QLabel#heroTitle {
                font-size: 20px;
                font-weight: 900;
            }
            QLabel#heroText {
                color: rgba(232,240,252,0.72);
            }
            QLabel#heroIcon {
                background: rgba(37, 173, 165, 0.17);
                border: 1px solid rgba(37, 173, 165, 0.40);
                border-radius: 14px;
            }
            QFrame#heroChip {
                background: rgba(255,255,255,0.03);
                border: 1px solid rgba(255,255,255,0.10);
                border-radius: 14px;
            }
            QLabel#heroChipIcon {
                background: rgba(37, 173, 165, 0.14);
                border: 1px solid rgba(37, 173, 165, 0.32);
                border-radius: 12px;
            }
            QLabel#heroChipTitle {
                font-weight: 900;
            }
            QLabel#heroChipText {
                color: rgba(232,240,252,0.70);
                font-size: 13px;
            }
            QToolButton#heroCta {
                font-size: 15px;
                font-weight: 880;
                padding: 12px 14px;
            }

            QDialog#prettyDialog {
                background: #0f1826;
            }
            QLabel#dlgTitle {
                font-size: 16px;
                font-weight: 900;
            }
            QLabel#dlgText {
                color: rgba(232,240,252,0.82);
            }
            QLabel#dlgIcon {
                background: rgba(37, 173, 165, 0.17);
                border: 1px solid rgba(37, 173, 165, 0.38);
                border-radius: 12px;
            }

            QMenu {
                background: #142237;
                border: 1px solid #2a3f5c;
                padding: 4px;
            }
            QMenu::item {
                padding: 6px 16px;
                border-radius: 6px;
            }
            QMenu::item:selected {
                background: rgba(37, 173, 165, 0.22);
            }

            QToolButton[chip="true"] {
                padding: 6px 14px;
                border-radius: 14px;
                font-weight: 880;
                letter-spacing: 0.2px;
            }
            QToolButton[chip="true"][state="empty"] {
                background: rgba(255,255,255,0.02);
                border: 1px dashed rgba(255,255,255,0.24);
                color: rgba(232,240,252,0.74);
            }
            QToolButton[chip="true"][state="empty"]:hover {
                background: rgba(255,255,255,0.05);
                border: 1px dashed rgba(255,255,255,0.30);
            }
            QToolButton[chip="true"][state="ok"][slot="pdf"] {
                background: rgba(34, 197, 94, 0.16);
                border: 1px solid rgba(34, 197, 94, 0.55);
                color: #f1f6ff;
            }
            QToolButton[chip="true"][state="ok"][slot="pdf"]:hover {
                background: rgba(34, 197, 94, 0.22);
                border: 1px solid rgba(34, 197, 94, 0.70);
            }
            QToolButton[chip="true"][state="ok"][slot="office"] {
                background: rgba(59, 130, 246, 0.16);
                border: 1px solid rgba(59, 130, 246, 0.55);
                color: #f1f6ff;
            }
            QToolButton[chip="true"][state="ok"][slot="office"]:hover {
                background: rgba(59, 130, 246, 0.22);
                border: 1px solid rgba(59, 130, 246, 0.70);
            }
            QToolButton[chip="true"][state="missing"] {
                background: rgba(251, 146, 60, 0.14);
                border: 1px solid rgba(251, 146, 60, 0.55);
                color: #f1f6ff;
            }
            QToolButton[chip="true"][state="missing"]:hover {
                background: rgba(251, 146, 60, 0.20);
                border: 1px solid rgba(251, 146, 60, 0.70);
            }

            QFrame#emptyOverlay {
                background: rgba(11, 18, 30, 0.62);
                border-radius: 14px;
            }
            QLabel#emptyOverlayTitle {
                font-size: 18px;
                font-weight: 900;
            }
            QLabel#emptyOverlayText {
                color: rgba(232,240,252,0.70);
            }

            QScrollBar:vertical {
                background: transparent;
                width: 10px;
                margin: 8px 4px 8px 0px;
            }
            QScrollBar::handle:vertical {
                background: rgba(255,255,255,0.18);
                border-radius: 5px;
                min-height: 28px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(255,255,255,0.28);
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
                background: transparent;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: transparent;
            }

            QScrollBar:horizontal {
                background: transparent;
                height: 10px;
                margin: 0px 8px 4px 8px;
            }
            QScrollBar::handle:horizontal {
                background: rgba(255,255,255,0.18);
                border-radius: 5px;
                min-width: 28px;
            }
            QScrollBar::handle:horizontal:hover {
                background: rgba(255,255,255,0.28);
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
                background: transparent;
            }
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                background: transparent;
            }
        """)

    def load_companies(self, initial: bool = False, preferred_company_id: Optional[int] = None):
        rows = self.db.list_companies_with_stats()
        if len(rows) == 0:
            self.switch_app_page(0)
            return
        self.switch_app_page(1)

        self._apply_company_filter(preferred_company_id=preferred_company_id)
        if initial:
            self._fade_in(self.page_main, 220)

    def add_company(self):
        dlg = InputDialog(self, "Новая компания", "Название компании:")
        if dlg.exec() != QDialog.Accepted:
            return
        name = dlg.value()
        try:
            new_company_id = self.db.add_company(name)
            self.load_companies(preferred_company_id=new_company_id)
        except sqlite3.IntegrityError:
            self.warn("Такая компания уже существует.")
        except Exception as e:
            self.warn(str(e))

    def company_context_menu(self, pos: QPoint):
        item = self.company_list.itemAt(pos)
        if not item:
            return
        company_id = int(item.data(Qt.UserRole))
        company_name = item.data(Qt.UserRole + 1) or item.text()
        company_name = str(company_name)

        menu = QMenu(self)
        a_rename = menu.addAction("Переименовать")
        a_copy = menu.addAction("Копировать компанию")
        menu.addSeparator()
        a_del = menu.addAction("Удалить")

        act = menu.exec(self.company_list.viewport().mapToGlobal(pos))
        if act == a_rename:
            dlg = InputDialog(self, "Переименовать компанию", "Новое название:", default=company_name)
            if dlg.exec() == QDialog.Accepted:
                try:
                    self.db.rename_company(company_id, dlg.value())
                    self.load_companies(preferred_company_id=company_id)
                except sqlite3.IntegrityError:
                    self.warn("Компания с таким названием уже существует.")
        elif act == a_copy:
            default_name = self._unique_copy_name(company_name)
            dlg = InputDialog(self, "Копировать компанию", "Название копии:", default=default_name)
            if dlg.exec() == QDialog.Accepted:
                self.copy_company(company_id, company_name, dlg.value())
        elif act == a_del:
            if self.confirm(f"Удалить компанию «{company_name}» и все документы?"):
                try:
                    self.db.delete_company(company_id)
                    try:
                        shutil.rmtree(os.path.join(self.storage_dir, f"company_{company_id}"), ignore_errors=True)
                    except Exception:
                        pass
                    self.load_companies()
                except Exception as e:
                    self.warn(str(e))

    def _unique_copy_name(self, base_name: str) -> str:
        existing = {r["name"] for r in self.db.list_companies()}
        candidate = f"{base_name} (копия)"
        if candidate not in existing:
            return candidate
        i = 2
        while True:
            cand = f"{base_name} (копия {i})"
            if cand not in existing:
                return cand
            i += 1

    def copy_company(self, src_company_id: int, src_company_name: str, new_name: str):
        try:
            new_company_id = self.db.copy_company_structure(src_company_id, new_name)
            src_doc_ids = self.db.list_doc_ids_by_company(src_company_id)
            new_doc_ids = self.db.list_doc_ids_by_company(new_company_id)
            src_sopd_ids = self.db.list_sopd_ids_by_company(src_company_id)
            new_sopd_ids = self.db.list_sopd_ids_by_company(new_company_id)
            src_sopd_rel = self.db.get_company_sopd_file_path(src_company_id)
            src_sopd_abs = self._abs_storage_path(src_sopd_rel) if src_sopd_rel else None

            copy_plan = []
            for src_doc_id, new_doc_id in zip(src_doc_ids, new_doc_ids):
                src_info = self.db.get_doc_info(src_doc_id)
                if not src_info:
                    continue

                src_pdf_abs = self._abs_storage_path(src_info["pdf_path"])
                src_off_abs = self._abs_storage_path(src_info["office_path"])

                if src_pdf_abs and os.path.exists(src_pdf_abs):
                    dst = os.path.join(self._doc_dir(new_company_id, new_doc_id), os.path.basename(src_pdf_abs))
                    copy_plan.append((src_pdf_abs, dst, new_doc_id, "pdf_path", self._rel_storage_path(dst)))

                if src_off_abs and os.path.exists(src_off_abs):
                    dst = os.path.join(self._doc_dir(new_company_id, new_doc_id), os.path.basename(src_off_abs))
                    copy_plan.append((src_off_abs, dst, new_doc_id, "office_path", self._rel_storage_path(dst)))

            sopd_record_copy_plan = []
            for src_sopd_id, new_sopd_id in zip(src_sopd_ids, new_sopd_ids):
                src_sopd = self.db.get_sopd_record(src_sopd_id)
                if not src_sopd:
                    continue
                src_attachment_abs = self._abs_storage_path(src_sopd["attachment_path"])
                if not src_attachment_abs or not os.path.exists(src_attachment_abs):
                    continue
                dst = os.path.join(self._sopd_record_dir(new_company_id, new_sopd_id), os.path.basename(src_attachment_abs))
                sopd_record_copy_plan.append((src_attachment_abs, dst, new_sopd_id, self._rel_storage_path(dst)))

            sopd_copy = None
            if src_sopd_abs and os.path.exists(src_sopd_abs):
                sopd_dst = os.path.join(self._sopd_dir(new_company_id), os.path.basename(src_sopd_abs))
                sopd_copy = (src_sopd_abs, sopd_dst, self._rel_storage_path(sopd_dst))
        except Exception as e:
            self.warn(f"Не получилось скопировать:\n{e}")
            return

        def work():
            updates = []
            for src_abs, dst_abs, dst_doc_id, field, rel_path in copy_plan:
                atomic_copy(src_abs, dst_abs)
                updates.append((dst_doc_id, field, rel_path))
            sopd_record_updates = []
            for src_abs, dst_abs, dst_record_id, rel_path in sopd_record_copy_plan:
                atomic_copy(src_abs, dst_abs)
                sopd_record_updates.append((dst_record_id, rel_path))
            sopd_rel = None
            if sopd_copy is not None:
                src_abs, dst_abs, rel_path = sopd_copy
                atomic_copy(src_abs, dst_abs)
                sopd_rel = rel_path
            return {
                "company_id": new_company_id,
                "updates": updates,
                "sopd_rel": sopd_rel,
                "sopd_record_updates": sopd_record_updates,
            }

        def on_success(result: Any):
            for dst_doc_id, field, rel_path in result["updates"]:
                self.db.update_file_path(dst_doc_id, field, rel_path)
            for dst_record_id, rel_path in result.get("sopd_record_updates", []):
                self.db.update_sopd_attachment_path(dst_record_id, rel_path)
            if result.get("sopd_rel"):
                self.db.update_company_sopd_file_path(result["company_id"], result["sopd_rel"])
            self.load_companies(preferred_company_id=result["company_id"])
            self.info("Компания скопирована ✅")

        self._run_background_task(
            label="Копирование компании",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось скопировать:",
        )

    def on_company_changed(self):
        company_id, company_name = self.current_company_id_name()
        if not company_id:
            return
        mode = self._content_mode
        self.lbl_right.setText(f"Документы и СОПД — {company_name}")
        self.load_sections(company_id, preferred_section_id=None)
        self.set_content_mode(mode)
        self._fade_in(self.right_stack, 140)

    def add_document(self):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return
        dlg = AddDocDialog(self)
        if dlg.exec() != QDialog.Accepted:
            return
        title = dlg.get_value()
        try:
            self.db.add_document(company_id, title)
            self.load_documents()
        except Exception as e:
            self.warn(str(e))

    def refresh_documents(self):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return

        rows = self.db.list_documents(company_id)

        q = self.doc_filter.text().strip().lower()
        if q:
            rows = [r for r in rows if q in (r["doc_title"] or "").lower()]

        if q and len(rows) == 0:
            self._show_docs_empty_search()
        else:
            self._hide_docs_empty_search()

        self._render_documents(company_id, rows)

    def persist_current_doc_order(self):
        if getattr(self, "current_section_id", None) is not None:
            return

        company_id, _ = self.current_company_id_name()
        if not company_id:
            return

        ordered = []
        for i in range(self.table.rowCount()):
            it = self.table.item(i, 0)
            if it and it.text().isdigit():
                ordered.append(int(it.text()))

        if ordered:
            try:
                self.db.set_company_doc_order(company_id, ordered)
            except Exception as e:
                self.warn(str(e))

    def delete_document_row(self, doc_id: int):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return
        if not self.confirm("Удалить эту строку документа?"):
            return

        try:
            with self._doc_lock(company_id, doc_id):
                info = self.db.get_doc_info(doc_id)
                if info:
                    for stored in [info["pdf_path"], info["office_path"]]:
                        abs_path = self._abs_storage_path(stored)
                        if abs_path and os.path.exists(abs_path):
                            try:
                                os.remove(abs_path)
                            except Exception:
                                pass
                self.db.delete_document(doc_id)
        except TimeoutError:
            self.warn("Документ сейчас редактирует другой пользователь. Попробуй чуть позже.")
            return
        except Exception as e:
            self.warn(str(e))
            return

        self.load_documents()

    def upload_file_dialog(self, doc_id: int, kind: str):
        if kind == "pdf":
            file_path, _ = QFileDialog.getOpenFileName(self, "Выбери PDF (скан)", "", "PDF (*.pdf)")
        else:
            file_path, _ = QFileDialog.getOpenFileName(self, "Выбери Word/Excel", "", "Файлы Word и Excel (*.doc *.docx *.xls *.xlsx)")
        if not file_path:
            return
        self.upload_file_from_path(doc_id, kind, file_path)

    def upload_file_from_path(self, doc_id: int, kind: str, file_path: str):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return

        ext = os.path.splitext(file_path)[1].lower()
        doc_dir = self._doc_dir(company_id, doc_id)
        base_name = self._safe_filename(os.path.splitext(os.path.basename(file_path))[0])
        if kind == "pdf":
            if ext != ".pdf":
                self.warn("Для PDF слота нужен файл .pdf")
                return
            field = "pdf_path"
            dst = os.path.join(doc_dir, f"{base_name}.pdf")
        else:
            if ext not in (".doc", ".docx", ".xls", ".xlsx"):
                self.warn("Для Office слота нужен doc/docx/xls/xlsx")
                return
            field = "office_path"
            dst = os.path.join(doc_dir, f"{base_name}{ext}")

        info = self.db.get_doc_info(doc_id)
        old_rel = info[field] if info else None
        old_abs = self._abs_storage_path(old_rel) if old_rel else None
        rel_dst = self._rel_storage_path(dst)

        def work():
            with self._doc_lock(company_id, doc_id):
                if old_abs and os.path.exists(old_abs):
                    try:
                        os.remove(old_abs)
                    except Exception:
                        pass
                atomic_copy(file_path, dst)
            return {"doc_id": doc_id, "field": field, "rel_path": rel_dst}

        def on_success(result: Any):
            self.db.update_file_path(result["doc_id"], result["field"], result["rel_path"])
            self.load_documents()

        self._run_background_task(
            label="Загрузка файла",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось загрузить файл:",
        )

    def open_file(self, doc_id: int, kind: str):
        info = self.db.get_doc_info(doc_id)
        if not info:
            self.warn("Запись документа не найдена.")
            return

        stored = info["pdf_path"] if kind == "pdf" else info["office_path"]
        path = self._abs_storage_path(stored)

        if not path or not os.path.exists(path):
            self.warn("Файл не найден. Возможно, его удалили вручную.")
            return

        QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(path)))

    def download_file(self, doc_id: int, kind: str):
        info = self.db.get_doc_info(doc_id)
        if not info:
            self.warn("Запись документа не найдена.")
            return

        stored = info["pdf_path"] if kind == "pdf" else info["office_path"]
        path = self._abs_storage_path(stored)

        if not path or not os.path.exists(path):
            self.warn("Файл не найден.")
            return

        base = os.path.basename(path)
        save_path, _ = QFileDialog.getSaveFileName(self, "Скачать как…", base, "Все файлы (*)")
        if not save_path:
            return

        def work():
            atomic_copy(path, save_path)
            return None

        def on_success(_: Any):
            self.info("Файл сохранён.")

        self._run_background_task(
            label="Скачивание файла",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось сохранить:",
        )

    def delete_file(self, doc_id: int, kind: str):
        company_id, _ = self.current_company_id_name()
        if not company_id:
            return

        field = "pdf_path" if kind == "pdf" else "office_path"
        info = self.db.get_doc_info(doc_id)
        if not info:
            self.warn("Запись документа не найдена.")
            return

        stored = info[field]
        abs_path = self._abs_storage_path(stored)

        if not stored:
            self.db.update_file_path(doc_id, field, None)
            self.load_documents()
            return

        if not self.confirm("Удалить файл из реестра и с диска?"):
            return

        try:
            with self._doc_lock(company_id, doc_id):
                if abs_path and os.path.exists(abs_path):
                    try:
                        os.remove(abs_path)
                    except Exception:
                        pass
                self.db.update_file_path(doc_id, field, None)

            self.load_documents()
        except TimeoutError:
            self.warn("Документ сейчас редактирует другой пользователь. Попробуй чуть позже.")
        except Exception as e:
            self.warn(str(e))

    def export_registry(self):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            self.warn("openpyxl не установлен. Установи: pip install openpyxl")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self, "Выгрузить реестр", "pd_docs_registry.xlsx", "Книга Excel (*.xlsx)"
        )
        if not save_path:
            return

        rows_payload = []
        for r in self.db.get_all_registry_rows():
            pdf_stored = r["pdf_path"] or ""
            office_stored = r["office_path"] or ""
            rows_payload.append(
                {
                    "company_name": r["company_name"],
                    "sort_order": r["sort_order"],
                    "doc_id": r["doc_id"],
                    "doc_title": r["doc_title"],
                    "status": r["status"] or DOC_STATUS_DEFAULT,
                    "accept_date": fmt_date_ddmmyyyy(r["accept_date"] or ""),
                    "review_due": fmt_date_ddmmyyyy(r["review_due"] or ""),
                    "pdf_stored": pdf_stored,
                    "office_stored": office_stored,
                    "pdf_abs": self._abs_storage_path(pdf_stored),
                    "office_abs": self._abs_storage_path(office_stored),
                    "updated_at": r["updated_at"],
                    "updated_by": r["updated_by"] or "",
                }
            )

        def work():
            wb = Workbook()
            ws = wb.active
            ws.title = "Registry"

            headers = [
                "Компания", "Порядок", "ID документа", "Название документа",
                "Статус",
                "Дата принятия",
                "Дата пересмотра",
                "PDF загружен", "Office загружен",
                "PDF путь", "Office путь", "Обновлено", "Кто изменил"
            ]
            ws.append(headers)

            for r in rows_payload:
                pdf_ok = "Да" if (r["pdf_abs"] and os.path.exists(r["pdf_abs"])) else "Нет"
                office_ok = "Да" if (r["office_abs"] and os.path.exists(r["office_abs"])) else "Нет"
                ws.append([
                    r["company_name"],
                    r["sort_order"],
                    r["doc_id"],
                    r["doc_title"],
                    r["status"],
                    r["accept_date"],
                    r["review_due"],
                    pdf_ok,
                    office_ok,
                    r["pdf_stored"],
                    r["office_stored"],
                    r["updated_at"],
                    r["updated_by"],
                ])

            widths = [28, 10, 12, 46, 18, 14, 14, 12, 13, 50, 50, 20, 22]
            for idx, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = w

            wb.save(save_path)
            return None

        def on_success(_: Any):
            self.info("Реестр выгружен ✅")

        self._run_background_task(
            label="Экспорт в Excel",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось сохранить Excel:",
        )


def _refresh_widget_style(widget: QWidget):
    widget.style().unpolish(widget)
    widget.style().polish(widget)
    widget.update()


class ClickableCard(QFrame):
    clicked = Signal(int)

    def __init__(self, record_id: int, parent=None):
        super().__init__(parent)
        self.record_id = int(record_id)
        self.setCursor(Qt.PointingHandCursor)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

    def mousePressEvent(self, event):
        self.clicked.emit(self.record_id)
        super().mousePressEvent(event)

    def enable_child_click_proxy(self):
        for child in self.findChildren(QWidget):
            if child is self or isinstance(child, QToolButton):
                continue
            child.installEventFilter(self)

    def eventFilter(self, watched, event):
        if (
            watched is not self
            and event.type() == QEvent.MouseButtonPress
            and hasattr(event, "button")
            and event.button() == Qt.LeftButton
        ):
            self.clicked.emit(self.record_id)
        return super().eventFilter(watched, event)

    def set_selected(self, selected: bool):
        self.setProperty("selectedCard", selected)
        _refresh_widget_style(self)


OldMainWindow = MainWindow


class MainWindow(OldMainWindow):
    PAGE_TITLES = {
        "dashboard": "Дашборд",
        "attention": "Требует внимания",
        "documents": "Документы",
        "sopd": "СОПД",
        "settings": "Компании и разделы",
    }

    PAGE_SUBTITLES = {
        "dashboard": "",
        "attention": "",
        "documents": "",
        "sopd": "",
        "settings": "",
    }

    def __init__(self):
        QMainWindow.__init__(self)

        self.current_account = f"{getpass.getuser()}@{platform.node()}"
        self.shared_root = ""
        self.db_path = ""
        self.storage_dir = ""
        self.db = None
        self._due_reviews_count = 0
        self._thread_pool = QThreadPool.globalInstance()
        self._bg_tasks: set[BackgroundTask] = set()
        self._busy_jobs = 0
        self._busy_cursor_set = False
        self._busy_label = ""
        self._page_key = "dashboard"
        self.current_document_id: Optional[int] = None
        self.current_sopd_id: Optional[int] = None
        self._doc_form_loading = False
        self._doc_form_dirty = False
        self._sopd_form_loading = False
        self._sopd_form_dirty = False
        self._document_cards: Dict[int, ClickableCard] = {}
        self._sopd_cards: Dict[int, ClickableCard] = {}
        self._review_cards: Dict[int, ClickableCard] = {}
        self._pending_document_files: Dict[str, Optional[str]] = {"pdf": None, "office": None}
        self._settings_company_ids: List[int] = []
        self._settings_section_ids: List[int] = []
        self._startup_catalog_rows: Optional[List[Dict[str, Any]]] = None
        self._document_catalog_cache: Optional[List[Dict[str, Any]]] = None
        self._startup_catalog_pending = False
        self._startup_catalog_loading = False
        self._last_company_filter_value: Optional[int] = None
        self._last_section_filter_value: Optional[int] = None
        self._last_status_filter_value: Optional[str] = None
        self._last_problem_filter_value: Optional[str] = None

        self.setWindowTitle(APP_TITLE)
        self.resize(1480, 920)
        self.setMinimumSize(1280, 820)

        self.app_stack = QStackedWidget()
        self.app_stack.setObjectName("appStack")
        self.setCentralWidget(self.app_stack)

        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self.refresh_current_page)

        self.page_empty_app = self._build_empty_app_page()
        self.page_main = self._build_main_page()
        self.app_stack.addWidget(self.page_empty_app)
        self.app_stack.addWidget(self.page_main)

        self._apply_style()
        self._setup_menu()
        self._setup_shortcuts()
        self._update_nav_state()
        self._update_toolbar_state()
        self.right_stack.setCurrentWidget(self.right_placeholder_page)

        self._notif_timer = QTimer(self)
        self._notif_timer.setInterval(60_000)
        self._notif_timer.timeout.connect(self.refresh_notifications)
        self._notif_timer.start()

        self._update_statusbar()
        QTimer.singleShot(0, self.bootstrap)

    def _setup_menu(self):
        menu_file = self.menuBar().addMenu("Файл")
        act_export = QAction("Выгрузить в Excel…", self)
        act_export.setShortcut("Ctrl+E")
        act_export.triggered.connect(self.export_registry)
        menu_file.addAction(act_export)

        menu_tools = self.menuBar().addMenu("Ещё")
        act_notifications = QAction("Центр уведомлений", self)
        act_notifications.triggered.connect(self.open_notifications_center)
        menu_tools.addAction(act_notifications)

        menu_settings = self.menuBar().addMenu("Папка")
        act_change = QAction("Сменить рабочую папку…", self)
        act_change.triggered.connect(self.change_shared_root)
        menu_settings.addAction(act_change)

    def _setup_shortcuts(self):
        act_new_doc = QAction(self)
        act_new_doc.setShortcut("Ctrl+N")
        act_new_doc.triggered.connect(self._handle_primary_action)
        self.addAction(act_new_doc)

        act_search = QAction(self)
        act_search.setShortcut("Ctrl+F")
        act_search.triggered.connect(lambda: self.search_input.setFocus())
        self.addAction(act_search)

    def _create_button(
        self,
        text: str,
        kind: str = "ghost",
        icon: Optional[QIcon] = None,
        checkable: bool = False,
    ) -> QToolButton:
        kind_map = {
            "solid": "primary",
            "ghost": "secondary",
            "soft": "secondary",
        }
        btn = AnimatedToolButton()
        btn.setText(text)
        btn.setProperty("kind", kind_map.get(kind, kind))
        btn.setCheckable(checkable)
        btn.setToolButtonStyle(Qt.ToolButtonTextBesideIcon if icon else Qt.ToolButtonTextOnly)
        if icon:
            btn.setIcon(icon)
        return btn

    def _create_menu_button(self, text: str, menu: QMenu, kind: str = "ghost") -> QToolButton:
        btn = self._create_button(text, kind=kind)
        btn.setProperty("menuButton", True)
        btn.setMenu(menu)
        btn.setPopupMode(QToolButton.InstantPopup)
        _refresh_widget_style(btn)
        return btn

    def _create_panel(self, title: str, subtitle: str = "") -> Tuple[QFrame, QVBoxLayout]:
        panel = QFrame()
        panel.setObjectName("contentCard")
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(22, 20, 22, 20)
        lay.setSpacing(14)

        head = QVBoxLayout()
        head.setSpacing(4)

        title_lbl = QLabel(title)
        title_lbl.setObjectName("sectionTitle")
        head.addWidget(title_lbl)

        if subtitle:
            subtitle_lbl = QLabel(subtitle)
            subtitle_lbl.setObjectName("sectionSubtitle")
            subtitle_lbl.setWordWrap(True)
            head.addWidget(subtitle_lbl)

        lay.addLayout(head)
        return panel, lay

    def _configure_scroll_area(self, area: QScrollArea):
        area.setWidgetResizable(True)
        area.setFrameShape(QFrame.NoFrame)
        area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        area.viewport().setAutoFillBackground(False)
        area.viewport().setStyleSheet("background: transparent; border: none;")

    def _build_empty_app_page(self) -> QWidget:
        root = QWidget()
        root.setObjectName("emptyRoot")

        outer = QVBoxLayout(root)
        outer.setContentsMargins(44, 44, 44, 44)
        outer.setSpacing(0)

        hero = QFrame()
        hero.setObjectName("emptyHero")
        hero_l = QVBoxLayout(hero)
        hero_l.setContentsMargins(34, 34, 34, 34)
        hero_l.setSpacing(18)

        badge = QLabel("Порядок в документах")
        badge.setObjectName("heroBadge")
        hero_l.addWidget(badge, 0, Qt.AlignLeft)

        title = QLabel("Соберите единое рабочее пространство для документов")
        title.setObjectName("heroTitle")
        title.setWordWrap(True)
        hero_l.addWidget(title)

        text = QLabel(
            "Добавьте первую компанию, и приложение сразу откроет дашборд, экран внимания, реестр документов и карточки СОПД."
        )
        text.setObjectName("heroText")
        text.setWordWrap(True)
        hero_l.addWidget(text)

        actions = QHBoxLayout()
        actions.setSpacing(10)
        add_btn = self._create_button(
            "Добавить компанию",
            kind="solid",
            icon=self.style().standardIcon(QStyle.SP_FileDialogNewFolder),
        )
        add_btn.clicked.connect(self.add_company)
        actions.addWidget(add_btn, 0)

        open_btn = self._create_button("Выбрать рабочую папку", kind="ghost")
        open_btn.clicked.connect(self.change_shared_root)
        actions.addWidget(open_btn, 0)
        actions.addStretch(1)
        hero_l.addLayout(actions)

        outer.addStretch(1)
        outer.addWidget(hero, 0, Qt.AlignCenter)
        outer.addStretch(1)
        return root

    def _build_main_page(self) -> QWidget:
        root = QWidget()
        root.setObjectName("centralWidget")

        main = QHBoxLayout(root)
        main.setContentsMargins(20, 20, 20, 20)
        main.setSpacing(16)

        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(232)
        side_l = QVBoxLayout(sidebar)
        side_l.setContentsMargins(18, 18, 18, 18)
        side_l.setSpacing(16)

        brand = QFrame()
        brand.setObjectName("brandCard")
        brand_l = QVBoxLayout(brand)
        brand_l.setContentsMargins(16, 16, 16, 16)
        brand_l.setSpacing(6)
        brand_title = QLabel("Реестр документов")
        brand_title.setObjectName("brandTitle")
        brand_l.addWidget(brand_title)
        brand_text = QLabel("Дашборд, экран внимания, документы, СОПД и структура компаний в одном рабочем пространстве.")
        brand_text.setObjectName("brandText")
        brand_text.setWordWrap(True)
        brand_l.addWidget(brand_text)
        side_l.addWidget(brand)

        self.nav_buttons: Dict[str, QToolButton] = {}
        nav_host = QVBoxLayout()
        nav_host.setSpacing(8)
        for key in ("dashboard", "attention", "documents", "sopd", "settings"):
            btn = self._create_button(self.PAGE_TITLES[key], kind="nav", checkable=False)
            btn.setProperty("navButton", True)
            btn.clicked.connect(lambda checked=False, page=key: self.set_active_page(page))
            self.nav_buttons[key] = btn
            nav_host.addWidget(btn)
        nav_host.addStretch(1)
        side_l.addLayout(nav_host, 1)

        center = QWidget()
        center_l = QVBoxLayout(center)
        center_l.setContentsMargins(0, 0, 0, 0)
        center_l.setSpacing(16)

        header = QFrame()
        header.setObjectName("topHeader")
        header_l = QVBoxLayout(header)
        header_l.setContentsMargins(22, 20, 22, 18)
        header_l.setSpacing(14)

        head_row = QHBoxLayout()
        head_row.setSpacing(12)
        head_copy = QVBoxLayout()
        head_copy.setSpacing(4)
        self.page_title_label = QLabel(self.PAGE_TITLES["dashboard"])
        self.page_title_label.setObjectName("pageTitle")
        head_copy.addWidget(self.page_title_label)
        self.page_subtitle_label = QLabel(self.PAGE_SUBTITLES["dashboard"])
        self.page_subtitle_label.setObjectName("pageSubtitle")
        self.page_subtitle_label.setWordWrap(True)
        self.page_subtitle_label.setVisible(bool(self.PAGE_SUBTITLES["dashboard"].strip()))
        head_copy.addWidget(self.page_subtitle_label)
        head_row.addLayout(head_copy, 1)

        self.primary_action_button = self._create_button("Добавить документ", kind="solid")
        self.primary_action_button.clicked.connect(self._handle_primary_action)
        head_row.addWidget(self.primary_action_button, 0, Qt.AlignTop)
        header_l.addLayout(head_row)

        filters = QFrame()
        filters.setObjectName("filterBar")
        filters_l = QVBoxLayout(filters)
        filters_l.setContentsMargins(14, 14, 14, 14)
        filters_l.setSpacing(10)

        self.filters_top_row = QWidget()
        filters_top_l = QHBoxLayout(self.filters_top_row)
        filters_top_l.setContentsMargins(0, 0, 0, 0)
        filters_top_l.setSpacing(10)

        self.filters_bottom_row = QWidget()
        filters_bottom_l = QHBoxLayout(self.filters_bottom_row)
        filters_bottom_l.setContentsMargins(0, 0, 0, 0)
        filters_bottom_l.setSpacing(14)

        self.search_input = QLineEdit()
        self.search_input.setObjectName("searchField")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.setPlaceholderText("Найти документ или компанию")
        self.search_input.textChanged.connect(lambda: self._search_timer.start(160))
        self.doc_filter = self.search_input
        filters_top_l.addWidget(self.search_input, 1)

        self.company_filter_combo = ToolbarFilterCombo()
        self.company_filter_combo.setObjectName("toolbarCombo")
        self.company_filter_combo.currentIndexChanged.connect(self._on_company_filter_changed)
        filters_bottom_l.addWidget(self.company_filter_combo, 2)

        self.status_filter_combo = ToolbarFilterCombo()
        self.status_filter_combo.setObjectName("toolbarCombo")
        self.status_filter_combo.currentIndexChanged.connect(self._on_status_filter_changed)
        filters_bottom_l.addWidget(self.status_filter_combo, 1)

        self.section_filter_combo = ToolbarFilterCombo()
        self.section_filter_combo.setObjectName("toolbarCombo")
        self.section_filter_combo.currentIndexChanged.connect(self._on_section_filter_changed)
        filters_bottom_l.addWidget(self.section_filter_combo, 1)

        self.problem_filter_combo = ToolbarFilterCombo()
        self.problem_filter_combo.setObjectName("toolbarCombo")
        self.problem_filter_combo.currentIndexChanged.connect(self._on_problem_filter_changed)
        filters_bottom_l.addWidget(self.problem_filter_combo, 1)

        self.clear_filters_button = self._create_button("Сбросить фильтры", kind="soft")
        self.clear_filters_button.clicked.connect(self.reset_filters)
        filters_top_l.addWidget(self.clear_filters_button, 0)
        filters_l.addWidget(self.filters_top_row)
        filters_l.addWidget(self.filters_bottom_row)
        header_l.addWidget(filters)
        center_l.addWidget(header)

        self.page_stack = QStackedWidget()
        self.page_stack.setObjectName("pageStack")
        center_l.addWidget(self.page_stack, 1)

        self.right_panel_frame = QFrame()
        self.right_panel_frame.setObjectName("rightPanel")
        self.right_panel_frame.setFixedWidth(430)
        right_l = QVBoxLayout(self.right_panel_frame)
        right_l.setContentsMargins(18, 18, 18, 18)
        right_l.setSpacing(14)

        right_head = QVBoxLayout()
        right_head.setSpacing(4)
        self.right_panel_title = QLabel("Детали")
        self.right_panel_title.setObjectName("rightTitle")
        right_head.addWidget(self.right_panel_title)
        self.right_panel_subtitle = QLabel("Выберите элемент слева.")
        self.right_panel_subtitle.setObjectName("rightSubtitle")
        self.right_panel_subtitle.setWordWrap(True)
        right_head.addWidget(self.right_panel_subtitle)
        right_l.addLayout(right_head)

        self.right_stack = QStackedWidget()
        self.right_stack.setObjectName("rightStack")
        right_l.addWidget(self.right_stack, 1)
        self.right_actions_stack = QStackedWidget()
        self.right_actions_stack.setObjectName("rightActionsStack")
        self.right_actions_stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        right_l.addWidget(self.right_actions_stack, 0)

        self.dashboard_page = self._build_dashboard_page()
        self.attention_page = self._build_attention_page()
        self.documents_page = self._build_documents_page()
        self.sopd_page = self._build_sopd_page()
        self.settings_page = self._build_settings_page()
        for page in (
            self.dashboard_page,
            self.attention_page,
            self.documents_page,
            self.sopd_page,
            self.settings_page,
        ):
            self.page_stack.addWidget(page)

        self.right_placeholder_page = self._build_right_placeholder_page()
        self.right_document_page = self._build_document_editor()
        self.right_sopd_page = self._build_sopd_editor()
        self.right_settings_page = self._build_settings_summary_page()
        for page in (
            self.right_placeholder_page,
            self.right_document_page,
            self.right_sopd_page,
            self.right_settings_page,
        ):
            self.right_stack.addWidget(page)

        self.right_actions_empty_page = QWidget()
        self.right_document_actions_page = self._build_document_action_bar()
        self.right_sopd_actions_page = self._build_sopd_action_bar()
        for page in (
            self.right_actions_empty_page,
            self.right_document_actions_page,
            self.right_sopd_actions_page,
        ):
            self.right_actions_stack.addWidget(page)
        self.right_actions_stack.setCurrentWidget(self.right_actions_empty_page)
        self.right_actions_stack.hide()

        main.addWidget(sidebar, 0)
        main.addWidget(center, 1)
        main.addWidget(self.right_panel_frame, 0)
        return root

    def _build_dashboard_page(self) -> QWidget:
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        self._configure_scroll_area(scroll)
        host = QWidget()
        lay = QVBoxLayout(host)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(16)

        overview_panel, overview_l = self._create_panel("Статистика")
        self.dashboard_scope_label = QLabel("Сейчас показываем: все компании и все разделы")
        self.dashboard_scope_label.setObjectName("summaryPill")
        self.dashboard_scope_label.setWordWrap(True)
        self.dashboard_scope_label.hide()
        overview_l.addWidget(self.dashboard_scope_label, 0, Qt.AlignLeft)
        self.dashboard_scope_note = QLabel("После загрузки базы здесь появится сводка по документам и карточкам СОПД.")
        self.dashboard_scope_note.setObjectName("summaryText")
        self.dashboard_scope_note.setWordWrap(True)
        self.dashboard_scope_note.hide()
        overview_l.addWidget(self.dashboard_scope_note)

        self.metrics_grid = QGridLayout()
        self.metrics_grid.setHorizontalSpacing(12)
        self.metrics_grid.setVerticalSpacing(12)
        for column in range(4):
            self.metrics_grid.setColumnStretch(column, 1)
        self.metric_cards: Dict[str, Tuple[QLabel, QLabel]] = {}
        metric_specs = [
            ("docs_total", "Всего документов"),
            ("docs_missing", "Без файла"),
            ("docs_overdue", "Просрочено"),
            ("docs_upcoming", "Скоро пересмотр"),
        ]
        for index, (key, title) in enumerate(metric_specs):
            card = QFrame()
            card.setObjectName("metricCard")
            card_l = QVBoxLayout(card)
            card_l.setContentsMargins(18, 16, 18, 16)
            card_l.setSpacing(8)
            title_lbl = QLabel(title)
            title_lbl.setObjectName("metricLabel")
            title_lbl.setWordWrap(True)
            title_lbl.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            title_lbl.setMinimumHeight(34)
            card_l.addWidget(title_lbl)
            value = QLabel("0")
            value.setObjectName("metricValue")
            card_l.addWidget(value)
            note = QLabel("Загрузка…")
            note.setObjectName("metricNote")
            note.setWordWrap(True)
            note.hide()
            card_l.addWidget(note)
            self.metric_cards[key] = (value, note)
            self.metrics_grid.addWidget(card, 0, index)
        overview_l.addLayout(self.metrics_grid)
        lay.addWidget(overview_panel)

        chart_panel, chart_l = self._create_panel("Главный график")
        self.dashboard_main_chart = DashboardBarChart()
        self.dashboard_main_chart.setMinimumHeight(336)
        chart_l.addWidget(self.dashboard_main_chart)
        lay.addWidget(chart_panel)

        insight_row = QHBoxLayout()
        insight_row.setSpacing(16)

        attention_panel, attention_l = self._create_panel("Требует внимания")
        self.dashboard_attention_host = QWidget()
        self.dashboard_attention_layout = QVBoxLayout(self.dashboard_attention_host)
        self.dashboard_attention_layout.setContentsMargins(0, 0, 0, 0)
        self.dashboard_attention_layout.setSpacing(10)
        attention_l.addWidget(self.dashboard_attention_host)
        insight_row.addWidget(attention_panel, 3)

        recent_panel, recent_l = self._create_panel("Последние изменения")
        self.dashboard_recent_host = QWidget()
        self.dashboard_recent_layout = QVBoxLayout(self.dashboard_recent_host)
        self.dashboard_recent_layout.setContentsMargins(0, 0, 0, 0)
        self.dashboard_recent_layout.setSpacing(10)
        recent_l.addWidget(self.dashboard_recent_host)
        insight_row.addWidget(recent_panel, 2)
        lay.addLayout(insight_row)
        lay.addStretch(1)

        scroll.setWidget(host)
        outer.addWidget(scroll)
        return page

    def _build_attention_page(self) -> QWidget:
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        self._configure_scroll_area(scroll)
        host = QWidget()
        lay = QVBoxLayout(host)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(16)

        summary_panel, summary_l = self._create_panel("Требует внимания")
        summary_row = QHBoxLayout()
        summary_row.setSpacing(10)
        self.attention_urgent_summary = QLabel("Срочно: 0")
        self.attention_urgent_summary.setObjectName("summaryPill")
        summary_row.addWidget(self.attention_urgent_summary, 0, Qt.AlignLeft)
        self.attention_issue_summary = QLabel("Нужно поправить: 0")
        self.attention_issue_summary.setObjectName("summaryPill")
        summary_row.addWidget(self.attention_issue_summary, 0, Qt.AlignLeft)
        self.attention_recent_summary = QLabel("Недавно изменены: 0")
        self.attention_recent_summary.setObjectName("summaryPill")
        summary_row.addWidget(self.attention_recent_summary, 0, Qt.AlignLeft)
        summary_row.addStretch(1)
        summary_l.addLayout(summary_row)
        lay.addWidget(summary_panel)

        urgent_panel, urgent_l = self._create_panel("Срочно", "Просроченный пересмотр, отсутствие PDF и другие блокирующие проблемы.")
        self.attention_urgent_host = QWidget()
        self.attention_urgent_layout = QVBoxLayout(self.attention_urgent_host)
        self.attention_urgent_layout.setContentsMargins(0, 0, 0, 0)
        self.attention_urgent_layout.setSpacing(10)
        urgent_l.addWidget(self.attention_urgent_host)
        lay.addWidget(urgent_panel)

        issue_panel, issue_l = self._create_panel("Нужно поправить", "Документы без раздела, даты, статуса или обязательного дополнительного файла.")
        self.attention_issue_host = QWidget()
        self.attention_issue_layout = QVBoxLayout(self.attention_issue_host)
        self.attention_issue_layout.setContentsMargins(0, 0, 0, 0)
        self.attention_issue_layout.setSpacing(10)
        issue_l.addWidget(self.attention_issue_host)
        lay.addWidget(issue_panel)

        recent_panel, recent_l = self._create_panel("Недавно изменены", "Последние обновления, которые стоит быстро проверить.")
        self.attention_recent_host = QWidget()
        self.attention_recent_layout = QVBoxLayout(self.attention_recent_host)
        self.attention_recent_layout.setContentsMargins(0, 0, 0, 0)
        self.attention_recent_layout.setSpacing(10)
        recent_l.addWidget(self.attention_recent_host)
        lay.addWidget(recent_panel)

        lay.addStretch(1)
        scroll.setWidget(host)
        outer.addWidget(scroll)
        return page

    def _build_documents_page(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(12)

        self.documents_count_label = QLabel("0 документов")
        self.documents_count_label.setObjectName("listMeta")
        lay.addWidget(self.documents_count_label)

        self.documents_scroll = QScrollArea()
        self._configure_scroll_area(self.documents_scroll)
        self.documents_host = QWidget()
        self.documents_layout = QVBoxLayout(self.documents_host)
        self.documents_layout.setContentsMargins(0, 0, 0, 0)
        self.documents_layout.setSpacing(10)
        self.documents_scroll.setWidget(self.documents_host)
        lay.addWidget(self.documents_scroll, 1)
        return page

    def _build_sopd_page(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(12)

        self.sopd_count_label = QLabel("0 карточек")
        self.sopd_count_label.setObjectName("listMeta")
        lay.addWidget(self.sopd_count_label)

        self.sopd_scroll = QScrollArea()
        self._configure_scroll_area(self.sopd_scroll)
        self.sopd_host = QWidget()
        self.sopd_layout = QVBoxLayout(self.sopd_host)
        self.sopd_layout.setContentsMargins(0, 0, 0, 0)
        self.sopd_layout.setSpacing(10)
        self.sopd_scroll.setWidget(self.sopd_host)
        lay.addWidget(self.sopd_scroll, 1)
        return page

    def _build_review_page(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(16)

        top = QHBoxLayout()
        top.setSpacing(12)
        self.review_overdue_summary = QLabel("Просрочено: 0")
        self.review_overdue_summary.setObjectName("summaryPill")
        top.addWidget(self.review_overdue_summary, 0, Qt.AlignLeft)
        self.review_upcoming_summary = QLabel("Скоро пересмотр: 0")
        self.review_upcoming_summary.setObjectName("summaryPill")
        top.addWidget(self.review_upcoming_summary, 0, Qt.AlignLeft)
        top.addStretch(1)
        lay.addLayout(top)

        overdue_panel, overdue_l = self._create_panel("Просрочено", "Документы, которые уже пора обновить.")
        self.review_overdue_host = QWidget()
        self.review_overdue_layout = QVBoxLayout(self.review_overdue_host)
        self.review_overdue_layout.setContentsMargins(0, 0, 0, 0)
        self.review_overdue_layout.setSpacing(10)
        overdue_l.addWidget(self.review_overdue_host)
        lay.addWidget(overdue_panel, 1)

        upcoming_panel, upcoming_l = self._create_panel("Скоро пересмотр", "Документы, которые скоро потребуют обновления даты.")
        self.review_upcoming_host = QWidget()
        self.review_upcoming_layout = QVBoxLayout(self.review_upcoming_host)
        self.review_upcoming_layout.setContentsMargins(0, 0, 0, 0)
        self.review_upcoming_layout.setSpacing(10)
        upcoming_l.addWidget(self.review_upcoming_host)
        lay.addWidget(upcoming_panel, 1)
        return page

    def _build_archive_page(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(12)

        self.archive_count_label = QLabel("0 документов в архиве")
        self.archive_count_label.setObjectName("listMeta")
        lay.addWidget(self.archive_count_label)

        self.archive_scroll = QScrollArea()
        self._configure_scroll_area(self.archive_scroll)
        self.archive_host = QWidget()
        self.archive_layout = QVBoxLayout(self.archive_host)
        self.archive_layout.setContentsMargins(0, 0, 0, 0)
        self.archive_layout.setSpacing(10)
        self.archive_scroll.setWidget(self.archive_host)
        lay.addWidget(self.archive_scroll, 1)
        return page

    def _build_settings_page(self) -> QWidget:
        page = QWidget()
        lay = QHBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(16)

        companies_panel, companies_l = self._create_panel("Компании")
        self.settings_company_list = QListWidget()
        self.settings_company_list.setObjectName("simpleList")
        self.settings_company_list.itemSelectionChanged.connect(self._on_settings_company_changed)
        companies_l.addWidget(self.settings_company_list, 1)
        self.settings_company_empty_label = QLabel("Нет компаний. Добавьте первую компанию, чтобы начать вести реестр.")
        self.settings_company_empty_label.setObjectName("summaryText")
        self.settings_company_empty_label.setWordWrap(True)
        companies_l.addWidget(self.settings_company_empty_label)

        company_actions = QHBoxLayout()
        company_actions.setSpacing(8)
        btn_add_company = self._create_button("Добавить компанию", kind="soft")
        btn_add_company.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_add_company.clicked.connect(self.settings_add_company)
        company_actions.addWidget(btn_add_company, 1)
        company_menu = QMenu(self)
        self.settings_company_rename_action = company_menu.addAction("Переименовать")
        self.settings_company_copy_action = company_menu.addAction("Копировать")
        company_menu.addSeparator()
        self.settings_company_delete_action = company_menu.addAction("Удалить компанию")
        self.settings_company_rename_action.triggered.connect(self.settings_rename_company)
        self.settings_company_copy_action.triggered.connect(self.settings_copy_company)
        self.settings_company_delete_action.triggered.connect(self.settings_delete_company)
        self.settings_company_menu_btn = self._create_menu_button("Еще", company_menu, kind="ghost")
        self.settings_company_menu_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.settings_company_menu_btn.setMinimumWidth(92)
        company_actions.addWidget(self.settings_company_menu_btn, 0)
        companies_l.addLayout(company_actions)

        sections_panel, sections_l = self._create_panel("Разделы")
        self.settings_section_stack = QStackedWidget()
        self.settings_section_stack.setObjectName("settingsSectionStack")
        self.settings_section_list_page = QWidget()
        section_list_l = QVBoxLayout(self.settings_section_list_page)
        section_list_l.setContentsMargins(0, 0, 0, 0)
        section_list_l.setSpacing(0)
        self.settings_section_list = QListWidget()
        self.settings_section_list.setObjectName("simpleList")
        self.settings_section_list.itemSelectionChanged.connect(self._refresh_settings_action_controls)
        section_list_l.addWidget(self.settings_section_list)
        self.settings_section_empty_page = QWidget()
        section_empty_l = QVBoxLayout(self.settings_section_empty_page)
        section_empty_l.setContentsMargins(0, 0, 0, 0)
        section_empty_l.setSpacing(0)
        section_empty_l.addStretch(1)
        self.settings_section_empty_label = QLabel("Пусто")
        self.settings_section_empty_label.setObjectName("emptyPanelTitle")
        self.settings_section_empty_label.setAlignment(Qt.AlignCenter)
        section_empty_l.addWidget(self.settings_section_empty_label)
        section_empty_l.addStretch(1)
        self.settings_section_stack.addWidget(self.settings_section_list_page)
        self.settings_section_stack.addWidget(self.settings_section_empty_page)
        sections_l.addWidget(self.settings_section_stack, 1)

        section_actions = QHBoxLayout()
        section_actions.setSpacing(8)
        self.settings_add_section_btn = self._create_button("Добавить раздел", kind="soft")
        self.settings_add_section_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.settings_add_section_btn.clicked.connect(self.settings_add_section)
        section_actions.addWidget(self.settings_add_section_btn, 1)
        section_menu = QMenu(self)
        self.settings_section_rename_action = section_menu.addAction("Переименовать")
        self.settings_section_delete_action = section_menu.addAction("Удалить раздел")
        self.settings_section_rename_action.triggered.connect(self.settings_rename_section)
        self.settings_section_delete_action.triggered.connect(self.settings_delete_section)
        self.settings_section_menu_btn = self._create_menu_button("Еще", section_menu, kind="ghost")
        self.settings_section_menu_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.settings_section_menu_btn.setMinimumWidth(92)
        section_actions.addWidget(self.settings_section_menu_btn, 0)
        sections_l.addLayout(section_actions)

        lay.addWidget(companies_panel, 1)
        lay.addWidget(sections_panel, 1)
        self._refresh_settings_action_controls()
        return page

    def _build_right_placeholder_page(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(12)
        lay.addStretch(1)
        self.right_placeholder_title = QLabel("Выберите документ")
        self.right_placeholder_title.setObjectName("emptyPanelTitle")
        self.right_placeholder_title.setAlignment(Qt.AlignCenter)
        lay.addWidget(self.right_placeholder_title)
        self.right_placeholder_text = QLabel("")
        self.right_placeholder_text.setObjectName("emptyPanelText")
        self.right_placeholder_text.setWordWrap(True)
        self.right_placeholder_text.setAlignment(Qt.AlignCenter)
        self.right_placeholder_text.hide()
        lay.addWidget(self.right_placeholder_text)
        lay.addStretch(1)
        return page

    def _build_document_editor(self) -> QWidget:
        wrapper = QScrollArea()
        self._configure_scroll_area(wrapper)

        root = QWidget()
        lay = QVBoxLayout(root)
        lay.setContentsMargins(2, 2, 8, 2)
        lay.setSpacing(12)

        self.doc_summary_status_badge = QLabel("Черновик")
        self.doc_summary_review_badge = QLabel("Дата пересмотра не указана")
        self.doc_summary_pdf_badge = QLabel("PDF")
        self.doc_summary_office_badge = QLabel("Нет файла")
        self.doc_summary_meta = QLabel("")
        self.doc_attention_label = QLabel("")
        for hidden_widget in (
            self.doc_summary_status_badge,
            self.doc_summary_review_badge,
            self.doc_summary_pdf_badge,
            self.doc_summary_office_badge,
            self.doc_summary_meta,
            self.doc_attention_label,
        ):
            hidden_widget.hide()

        basic_panel, basic_l = self._create_panel("Основные данные")
        basic_form = QGridLayout()
        basic_form.setContentsMargins(0, 0, 0, 0)
        basic_form.setHorizontalSpacing(18)
        basic_form.setVerticalSpacing(10)
        basic_form.setColumnMinimumWidth(0, 126)
        basic_form.setColumnStretch(1, 1)

        self.doc_title_edit = QLineEdit()
        self.doc_title_edit.setPlaceholderText("Например: Политика обработки персональных данных")
        self.doc_title_edit.textChanged.connect(self._set_document_form_dirty)
        basic_form.addWidget(self._form_key_label("Название"), 0, 0)
        basic_form.addWidget(self.doc_title_edit, 0, 1)

        self.doc_company_combo = ThemedComboBox()
        self.doc_company_combo.setObjectName("formCombo")
        self.doc_company_combo.currentIndexChanged.connect(self._on_document_company_changed)
        basic_form.addWidget(self._form_key_label("Компания"), 1, 0)
        basic_form.addWidget(self.doc_company_combo, 1, 1)

        self.doc_status_combo = ThemedComboBox()
        self.doc_status_combo.setObjectName("formCombo")
        for status in DOC_STATUSES:
            if status == "Архив":
                continue
            self.doc_status_combo.addItem(status)
        self.doc_status_combo.currentIndexChanged.connect(self._set_document_form_dirty)
        basic_form.addWidget(self._form_key_label("Статус"), 2, 0)
        basic_form.addWidget(self.doc_status_combo, 2, 1)

        self.doc_review_date_edit = QLineEdit()
        self.doc_review_date_edit.setPlaceholderText("дд.мм.гггг")
        self.doc_review_date_edit.textChanged.connect(self._set_document_form_dirty)
        basic_form.addWidget(self._form_key_label("Дата пересмотра"), 3, 0)
        basic_form.addWidget(self.doc_review_date_edit, 3, 1)

        self.doc_accept_date_edit = QLineEdit()
        self.doc_accept_date_edit.setPlaceholderText("дд.мм.гггг")
        self.doc_accept_date_edit.textChanged.connect(self._set_document_form_dirty)
        basic_form.addWidget(self._form_key_label("Дата принятия"), 4, 0)
        basic_form.addWidget(self.doc_accept_date_edit, 4, 1)
        basic_l.addLayout(basic_form)
        lay.addWidget(basic_panel)

        self.doc_needs_office_checkbox = QCheckBox()
        self.doc_needs_office_checkbox.hide()
        self.doc_needs_office_checkbox.stateChanged.connect(lambda *_: self._set_document_form_dirty())
        self.doc_sections_label = QLabel("")
        self.doc_sections_label.hide()
        self.doc_sections_hint = QLabel("")
        self.doc_sections_hint.hide()
        self.doc_sections_list = QListWidget()
        self.doc_sections_list.setObjectName("checkList")
        self.doc_sections_list.hide()
        self.doc_sections_list.itemChanged.connect(lambda item: self._set_document_form_dirty())
        self.doc_comment_edit = QPlainTextEdit()
        self.doc_comment_edit.hide()
        self.doc_comment_edit.textChanged.connect(self._set_document_form_dirty)

        files_panel, files_l = self._create_panel("Файлы")
        files_column = QVBoxLayout()
        files_column.setContentsMargins(0, 0, 0, 0)
        files_column.setSpacing(12)

        pdf_tile = QFrame()
        pdf_tile.setObjectName("fileTile")
        pdf_tile_l = QVBoxLayout(pdf_tile)
        pdf_tile_l.setContentsMargins(14, 12, 14, 12)
        pdf_tile_l.setSpacing(8)
        pdf_head = QHBoxLayout()
        pdf_head.setContentsMargins(0, 0, 0, 0)
        pdf_head.setSpacing(8)
        pdf_title = QLabel("PDF")
        pdf_title.setObjectName("fileTileTitle")
        pdf_head.addWidget(pdf_title)
        pdf_head.addStretch(1)
        pdf_tile_l.addLayout(pdf_head)
        self.doc_pdf_tag = QLabel("Основной PDF")
        self._set_file_tag_state(self.doc_pdf_tag, "Основной PDF", "pdf", False)
        self.doc_pdf_tag.hide()
        self.doc_pdf_caption = QLabel("Файл не загружен")
        self.doc_pdf_caption.setObjectName("fileCaption")
        self.doc_pdf_caption.setWordWrap(True)
        pdf_tile_l.addWidget(self.doc_pdf_caption)
        pdf_actions = QHBoxLayout()
        pdf_actions.setContentsMargins(0, 0, 0, 0)
        pdf_actions.setSpacing(8)
        self.doc_pdf_upload_btn = self._create_button("Загрузить", kind="soft")
        self.doc_pdf_upload_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_pdf_upload_btn.clicked.connect(lambda: self.upload_selected_document_file("pdf"))
        pdf_actions.addWidget(self.doc_pdf_upload_btn)
        self.doc_pdf_open_btn = self._create_button("Открыть", kind="ghost")
        self.doc_pdf_open_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_pdf_open_btn.clicked.connect(lambda: self.open_selected_document_file("pdf"))
        pdf_actions.addWidget(self.doc_pdf_open_btn)
        self.doc_pdf_delete_btn = self._create_button("Удалить", kind="danger")
        self.doc_pdf_delete_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_pdf_delete_btn.clicked.connect(lambda: self.delete_selected_document_file("pdf"))
        pdf_actions.addWidget(self.doc_pdf_delete_btn)
        pdf_tile_l.addLayout(pdf_actions)
        files_column.addWidget(pdf_tile)

        office_tile = QFrame()
        office_tile.setObjectName("fileTile")
        office_tile_l = QVBoxLayout(office_tile)
        office_tile_l.setContentsMargins(14, 12, 14, 12)
        office_tile_l.setSpacing(8)
        office_head = QHBoxLayout()
        office_head.setContentsMargins(0, 0, 0, 0)
        office_head.setSpacing(8)
        office_title = QLabel("Word / Excel")
        office_title.setObjectName("fileTileTitle")
        office_head.addWidget(office_title)
        office_head.addStretch(1)
        office_tile_l.addLayout(office_head)
        self.doc_office_tag = QLabel("Нет файла")
        self._set_file_tag_state(self.doc_office_tag, "Нет файла", "doc", False)
        self.doc_office_tag.hide()
        self.doc_office_caption = QLabel("Файл не загружен")
        self.doc_office_caption.setObjectName("fileCaption")
        self.doc_office_caption.setWordWrap(True)
        office_tile_l.addWidget(self.doc_office_caption)
        office_actions = QHBoxLayout()
        office_actions.setContentsMargins(0, 0, 0, 0)
        office_actions.setSpacing(8)
        self.doc_office_upload_btn = self._create_button("Загрузить", kind="soft")
        self.doc_office_upload_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_office_upload_btn.clicked.connect(lambda: self.upload_selected_document_file("office"))
        office_actions.addWidget(self.doc_office_upload_btn)
        self.doc_office_open_btn = self._create_button("Открыть", kind="ghost")
        self.doc_office_open_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_office_open_btn.clicked.connect(lambda: self.open_selected_document_file("office"))
        office_actions.addWidget(self.doc_office_open_btn)
        self.doc_office_delete_btn = self._create_button("Удалить", kind="danger")
        self.doc_office_delete_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_office_delete_btn.clicked.connect(lambda: self.delete_selected_document_file("office"))
        office_actions.addWidget(self.doc_office_delete_btn)
        office_tile_l.addLayout(office_actions)
        files_column.addWidget(office_tile)

        files_l.addLayout(files_column)
        lay.addWidget(files_panel)

        self.doc_history_host = QWidget()
        self.doc_history_layout = QVBoxLayout(self.doc_history_host)
        self.doc_history_layout.setContentsMargins(0, 0, 0, 0)
        self.doc_history_layout.setSpacing(10)
        self.doc_history_host.hide()
        lay.addStretch(1)

        wrapper.setWidget(root)
        return wrapper

    def _build_sopd_editor(self) -> QWidget:
        wrapper = QScrollArea()
        self._configure_scroll_area(wrapper)

        root = QWidget()
        lay = QVBoxLayout(root)
        lay.setContentsMargins(2, 2, 18, 2)
        lay.setSpacing(14)

        basic_panel, basic_l = self._create_panel("Основное")
        basic_form = QFormLayout()
        basic_form.setSpacing(10)
        basic_form.setHorizontalSpacing(18)
        basic_form.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        basic_form.setFormAlignment(Qt.AlignLeft | Qt.AlignTop)
        basic_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.sopd_company_combo = ThemedComboBox()
        self.sopd_company_combo.setObjectName("formCombo")
        self.sopd_company_combo.currentIndexChanged.connect(self._set_sopd_form_dirty)
        basic_form.addRow(self._form_key_label("Компания", width=132), self.sopd_company_combo)
        self.sopd_title_edit = QPlainTextEdit()
        self.sopd_title_edit.setPlaceholderText("Например: Согласие на обработку данных кандидатов")
        self.sopd_title_edit.textChanged.connect(self._set_sopd_form_dirty)
        basic_form.addRow(self._form_key_label("Название карточки", width=132), self.sopd_title_edit)
        basic_l.addLayout(basic_form)
        lay.addWidget(basic_panel)

        purpose_panel, purpose_l = self._create_panel("Цель обработки")
        self.sopd_purpose_edit = QPlainTextEdit()
        self.sopd_purpose_edit.setPlaceholderText("Опишите цель обработки")
        self.sopd_purpose_edit.textChanged.connect(self._set_sopd_form_dirty)
        purpose_l.addWidget(self.sopd_purpose_edit)
        lay.addWidget(purpose_panel)

        basis_panel, basis_l = self._create_panel("Правовое основание")
        self.sopd_legal_basis_edit = QPlainTextEdit()
        self.sopd_legal_basis_edit.setPlaceholderText("Например: согласие субъекта, договор, требование закона")
        self.sopd_legal_basis_edit.textChanged.connect(self._set_sopd_form_dirty)
        basis_l.addWidget(self.sopd_legal_basis_edit)
        lay.addWidget(basis_panel)

        categories_panel, categories_l = self._create_panel("Категории данных")
        self.sopd_categories_edit = QPlainTextEdit()
        self.sopd_categories_edit.setPlaceholderText("Категории персональных данных")
        self.sopd_categories_edit.textChanged.connect(self._set_sopd_form_dirty)
        categories_l.addWidget(self.sopd_categories_edit)
        self.sopd_pd_list_edit = QPlainTextEdit()
        self.sopd_pd_list_edit.setPlaceholderText("Подробный перечень данных")
        self.sopd_pd_list_edit.textChanged.connect(self._set_sopd_form_dirty)
        categories_l.addWidget(self.sopd_pd_list_edit)
        lay.addWidget(categories_panel)

        subjects_panel, subjects_l = self._create_panel("Субъекты данных")
        self.sopd_subjects_edit = QPlainTextEdit()
        self.sopd_subjects_edit.setPlaceholderText("Например: сотрудники, кандидаты, клиенты")
        self.sopd_subjects_edit.textChanged.connect(self._set_sopd_form_dirty)
        subjects_l.addWidget(self.sopd_subjects_edit)
        lay.addWidget(subjects_panel)

        operations_panel, operations_l = self._create_panel("Операции обработки")
        self.sopd_operations_edit = QPlainTextEdit()
        self.sopd_operations_edit.setPlaceholderText("Сбор, хранение, передача, удаление и т.д.")
        self.sopd_operations_edit.textChanged.connect(self._set_sopd_form_dirty)
        operations_l.addWidget(self.sopd_operations_edit)
        self.sopd_method_edit = QPlainTextEdit()
        self.sopd_method_edit.setPlaceholderText("Способ обработки")
        self.sopd_method_edit.textChanged.connect(self._set_sopd_form_dirty)
        operations_l.addWidget(self.sopd_method_edit)
        lay.addWidget(operations_panel)

        transfer_panel, transfer_l = self._create_panel("Передача третьим лицам")
        transfer_form = QFormLayout()
        transfer_form.setSpacing(10)
        transfer_form.setHorizontalSpacing(18)
        transfer_form.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        transfer_form.setFormAlignment(Qt.AlignLeft | Qt.AlignTop)
        transfer_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.sopd_transfer_combo = ThemedComboBox()
        self.sopd_transfer_combo.setObjectName("formCombo")
        for value in ("Не указано", "Да", "Нет"):
            self.sopd_transfer_combo.addItem(value)
        self.sopd_transfer_combo.currentIndexChanged.connect(self._set_sopd_form_dirty)
        transfer_form.addRow(self._form_key_label("Передача", width=132), self.sopd_transfer_combo)
        self.sopd_transfer_to_edit = QPlainTextEdit()
        self.sopd_transfer_to_edit.setPlaceholderText("Кому именно передаются данные")
        self.sopd_transfer_to_edit.textChanged.connect(self._set_sopd_form_dirty)
        transfer_form.addRow(self._form_key_label("Кому передаются", width=132), self.sopd_transfer_to_edit)
        transfer_l.addLayout(transfer_form)
        lay.addWidget(transfer_panel)

        extra_panel, extra_l = self._create_panel("Дополнительно")
        self.sopd_validity_edit = QLineEdit()
        self.sopd_validity_edit.setPlaceholderText("Например: до отзыва согласия")
        self.sopd_validity_edit.textChanged.connect(self._set_sopd_form_dirty)
        extra_l.addWidget(self.sopd_validity_edit)
        self.sopd_description_edit = QPlainTextEdit()
        self.sopd_description_edit.setPlaceholderText("Дополнительные пояснения")
        self.sopd_description_edit.textChanged.connect(self._set_sopd_form_dirty)
        extra_l.addWidget(self.sopd_description_edit)
        lay.addWidget(extra_panel)

        file_panel, file_l = self._create_panel("Файл СОПД")
        sopd_file_tile = QFrame()
        sopd_file_tile.setObjectName("fileTile")
        sopd_file_tile_l = QVBoxLayout(sopd_file_tile)
        sopd_file_tile_l.setContentsMargins(16, 14, 16, 14)
        sopd_file_tile_l.setSpacing(8)
        sopd_file_title = QLabel("Word / DOCX")
        sopd_file_title.setObjectName("fileTileTitle")
        sopd_file_tile_l.addWidget(sopd_file_title)
        self.sopd_file_caption = QLabel("Файл не загружен")
        self.sopd_file_caption.setObjectName("fileCaption")
        self.sopd_file_caption.setWordWrap(True)
        sopd_file_tile_l.addWidget(self.sopd_file_caption)
        self.sopd_file_upload_btn = self._create_button("Загрузить файл", kind="soft")
        self.sopd_file_upload_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.sopd_file_upload_btn.clicked.connect(self.upload_selected_sopd_file)
        sopd_file_tile_l.addWidget(self.sopd_file_upload_btn)
        file_actions = QHBoxLayout()
        file_actions.setSpacing(8)
        self.sopd_file_open_btn = self._create_button("Открыть", kind="ghost")
        self.sopd_file_open_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.sopd_file_open_btn.clicked.connect(self.open_selected_sopd_file)
        file_actions.addWidget(self.sopd_file_open_btn)
        self.sopd_file_delete_btn = self._create_button("Удалить", kind="danger")
        self.sopd_file_delete_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.sopd_file_delete_btn.clicked.connect(self.delete_selected_sopd_file)
        file_actions.addWidget(self.sopd_file_delete_btn)
        sopd_file_tile_l.addLayout(file_actions)
        file_l.addWidget(sopd_file_tile)
        lay.addWidget(file_panel)

        lay.addStretch(1)

        wrapper.setWidget(root)
        return wrapper

    def _build_document_action_bar(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)
        self.doc_save_btn = self._create_button("Сохранить документ", kind="solid")
        self.doc_save_btn.setProperty("panelPrimary", True)
        self.doc_save_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_save_btn.clicked.connect(self.save_document_form)
        _refresh_widget_style(self.doc_save_btn)
        lay.addWidget(self.doc_save_btn)
        self.doc_delete_btn = self._create_button("Удалить документ", kind="danger")
        self.doc_delete_btn.setProperty("panelDanger", True)
        self.doc_delete_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.doc_delete_btn.clicked.connect(self.delete_selected_document)
        _refresh_widget_style(self.doc_delete_btn)
        lay.addWidget(self.doc_delete_btn)
        return page

    def _build_sopd_action_bar(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)
        self.sopd_save_btn = self._create_button("Сохранить карточку", kind="solid")
        self.sopd_save_btn.setProperty("panelPrimary", True)
        self.sopd_save_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.sopd_save_btn.clicked.connect(self.save_sopd_form)
        _refresh_widget_style(self.sopd_save_btn)
        lay.addWidget(self.sopd_save_btn)
        self.sopd_delete_btn = self._create_button("Удалить карточку", kind="danger")
        self.sopd_delete_btn.setProperty("panelDanger", True)
        self.sopd_delete_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.sopd_delete_btn.clicked.connect(self.delete_selected_sopd)
        _refresh_widget_style(self.sopd_delete_btn)
        lay.addWidget(self.sopd_delete_btn)
        return page

    def _set_right_actions_mode(self, mode: str = "none"):
        if not hasattr(self, "right_actions_stack"):
            return
        mode_key = (mode or "none").strip().lower()
        page_map = {
            "document": self.right_document_actions_page,
            "sopd": self.right_sopd_actions_page,
        }
        target = page_map.get(mode_key, self.right_actions_empty_page)
        self.right_actions_stack.setCurrentWidget(target)
        self.right_actions_stack.setVisible(target is not self.right_actions_empty_page)

    def _build_settings_summary_page(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(2, 2, 2, 2)
        lay.setSpacing(14)

        workspace_panel, workspace_l = self._create_panel("Рабочая папка")
        self.settings_workspace_label = QLabel("Папка ещё не выбрана.")
        self.settings_workspace_label.setObjectName("summaryText")
        self.settings_workspace_label.setWordWrap(True)
        workspace_l.addWidget(self.settings_workspace_label)
        self.settings_db_label = QLabel("")
        self.settings_db_label.setObjectName("summaryText")
        self.settings_db_label.setWordWrap(True)
        workspace_l.addWidget(self.settings_db_label)

        workspace_actions = QHBoxLayout()
        workspace_actions.setSpacing(8)
        open_folder_btn = self._create_button("Открыть папку", kind="soft")
        open_folder_btn.clicked.connect(self._open_workspace_folder)
        workspace_actions.addWidget(open_folder_btn)
        change_folder_btn = self._create_button("Сменить папку", kind="ghost")
        change_folder_btn.clicked.connect(self.change_shared_root)
        workspace_actions.addWidget(change_folder_btn)
        workspace_actions.addStretch(1)
        workspace_l.addLayout(workspace_actions)
        lay.addWidget(workspace_panel)

        stats_panel, stats_l = self._create_panel("Сводка")
        self.settings_summary_label = QLabel("")
        self.settings_summary_label.setObjectName("summaryText")
        self.settings_summary_label.setWordWrap(True)
        stats_l.addWidget(self.settings_summary_label)
        lay.addWidget(stats_panel)
        lay.addStretch(1)
        return page

    def _apply_style(self):
        self.setStyleSheet("""
            QWidget {
                color: #F3EEFF;
                font-size: 14px;
                font-family: "Bahnschrift";
            }
            QMainWindow {
                background: #0B0713;
            }
            QWidget#centralWidget, QWidget#emptyRoot, QStackedWidget#appStack, QStackedWidget#pageStack, QStackedWidget#rightStack {
                background: transparent;
            }
            QMenuBar {
                background: #120D1C;
                color: #F3EEFF;
                border: none;
            }
            QMenuBar::item:selected {
                background: rgba(200, 174, 255, 0.14);
                border-radius: 8px;
            }
            QMenu {
                background: #171224;
                border: 1px solid rgba(200, 174, 255, 0.12);
                padding: 6px;
            }
            QMenu::item {
                padding: 8px 12px;
                border-radius: 8px;
            }
            QMenu::item:selected {
                background: rgba(200, 174, 255, 0.16);
            }
            QFrame#emptyHero, QFrame#brandCard, QFrame#topHeader, QFrame#filterBar,
            QFrame#contentCard, QFrame#sidebar, QFrame#rightPanel {
                background: #120D1C;
                border: none;
                border-radius: 24px;
            }
            QFrame#sidebar {
                background: #171224;
            }
            QFrame#metricCard {
                background: #171224;
                border: 1px solid rgba(200, 174, 255, 0.08);
                border-radius: 22px;
            }
            QFrame#analyticsRow, QFrame#fileTile {
                background: rgba(255, 255, 255, 0.03);
                border: 1px solid rgba(200, 174, 255, 0.08);
                border-radius: 20px;
            }
            QLabel#heroBadge {
                color: #C8AEFF;
                background: rgba(200, 174, 255, 0.10);
                border-radius: 12px;
                padding: 6px 10px;
                font-size: 12px;
                font-weight: 700;
            }
            QLabel#heroTitle, QLabel#pageTitle {
                font-size: 28px;
                font-weight: 900;
                color: #F8F4FF;
            }
            QLabel#heroText, QLabel#pageSubtitle, QLabel#brandText, QLabel#sectionSubtitle,
            QLabel#rightSubtitle, QLabel#metricNote, QLabel#summaryText, QLabel#emptyPanelText, QLabel#analyticsHint {
                color: #B8B0C9;
            }
            QLabel#brandTitle, QLabel#rightTitle, QLabel#sectionTitle {
                font-size: 18px;
                font-weight: 850;
                color: #F8F4FF;
            }
            QLabel#metricLabel, QLabel#fileTileTitle {
                color: #B8B0C9;
                font-size: 13px;
                font-weight: 800;
            }
            QLabel#metricValue {
                font-size: 34px;
                font-weight: 900;
                color: #C8AEFF;
            }
            QLabel#analyticsTitle {
                font-size: 15px;
                font-weight: 800;
                color: #F8F4FF;
            }
            QLabel#analyticsCount {
                font-size: 20px;
                font-weight: 900;
                color: #C8AEFF;
            }
            QLabel#listMeta, QLabel#fieldLabel {
                color: #B8B0C9;
                font-size: 13px;
            }
            QLabel#fileCaption {
                color: #B8B0C9;
                padding-top: 2px;
            }
            QLabel[fileTag="true"] {
                padding: 5px 12px;
                border-radius: 11px;
                font-size: 12px;
                font-weight: 900;
                border: 1px solid transparent;
            }
            QLabel[fileTag="true"][state="missing"] {
                background: rgba(255, 255, 255, 0.06);
                color: #B8B0C9;
            }
            QLabel[fileTag="true"][variant="pdf"][state="ready"] {
                background: rgba(255, 78, 108, 0.32);
                border: 1px solid rgba(255, 118, 146, 0.48);
                color: #FFF1F5;
            }
            QLabel[fileTag="true"][variant="doc"][state="ready"] {
                background: rgba(79, 151, 255, 0.34);
                border: 1px solid rgba(126, 181, 255, 0.48);
                color: #F3F8FF;
            }
            QLabel[fileTag="true"][variant="excel"][state="ready"] {
                background: rgba(44, 201, 118, 0.34);
                border: 1px solid rgba(86, 228, 145, 0.46);
                color: #F0FFF7;
            }
            QLabel#emptyPanelTitle {
                font-size: 20px;
                font-weight: 850;
                color: #F8F4FF;
            }
            QLabel#summaryPill {
                background: rgba(200, 174, 255, 0.10);
                color: #F3EEFF;
                border-radius: 14px;
                padding: 8px 12px;
                font-weight: 700;
            }
            QLineEdit, QComboBox, QPlainTextEdit, QListWidget#simpleList, QListWidget#checkList {
                background: rgba(255, 255, 255, 0.04);
                border: 1px solid rgba(200, 174, 255, 0.10);
                border-radius: 14px;
                padding: 0 14px;
                min-height: 44px;
                color: #F3EEFF;
                selection-background-color: rgba(200, 174, 255, 0.30);
            }
            QLineEdit:hover, QComboBox:hover, QPlainTextEdit:hover {
                background: rgba(255, 255, 255, 0.06);
            }
            QLineEdit:focus, QComboBox:focus, QPlainTextEdit:focus, QListWidget#simpleList:focus, QListWidget#checkList:focus {
                background: rgba(200, 174, 255, 0.08);
                border: 1px solid rgba(200, 174, 255, 0.26);
            }
            QLineEdit:disabled, QComboBox:disabled, QPlainTextEdit:disabled {
                background: rgba(255, 255, 255, 0.03);
                color: #6E6680;
            }
            QCheckBox {
                color: #F3EEFF;
                spacing: 10px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border-radius: 6px;
                border: 1px solid rgba(200, 174, 255, 0.18);
                background: rgba(255, 255, 255, 0.05);
            }
            QCheckBox::indicator:checked {
                background: #C8AEFF;
                border: 1px solid rgba(255, 255, 255, 0.18);
            }
            QCheckBox::indicator:hover {
                background: rgba(200, 174, 255, 0.16);
            }
            QPlainTextEdit {
                padding: 10px 14px;
                min-height: 96px;
            }
            QLineEdit#searchField {
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(200, 174, 255, 0.10);
                border-radius: 14px;
                min-height: 44px;
                padding: 0 16px;
                font-size: 15px;
            }
            QLineEdit#searchField:hover {
                background: rgba(255, 255, 255, 0.08);
            }
            QLineEdit#searchField:focus {
                background: rgba(200, 174, 255, 0.10);
                border: 1px solid rgba(200, 174, 255, 0.24);
            }
            QComboBox#toolbarCombo, QComboBox#formCombo {
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(200, 174, 255, 0.10);
                border-radius: 14px;
                min-height: 44px;
                padding: 0 14px;
                padding-right: 44px;
            }
            QComboBox#toolbarCombo {
                min-width: 148px;
            }
            QComboBox#formCombo {
                min-width: 168px;
            }
            QComboBox#toolbarCombo:hover, QComboBox#formCombo:hover {
                background: rgba(255, 255, 255, 0.08);
            }
            QComboBox#toolbarCombo:focus, QComboBox#formCombo:focus {
                border: 1px solid rgba(200, 174, 255, 0.24);
                background: rgba(200, 174, 255, 0.08);
            }
            QComboBox {
                padding-right: 44px;
                min-height: 44px;
                max-height: 44px;
            }
            QComboBox::drop-down {
                border: none;
                background: transparent;
                width: 0px;
                subcontrol-origin: padding;
                subcontrol-position: top right;
                margin: 0px;
                border-radius: 0px;
            }
            QComboBox::drop-down:hover, QComboBox::drop-down:on {
                background: transparent;
            }
            QComboBox::down-arrow {
                image: none;
                width: 0px;
                height: 0px;
            }
            QListWidget#simpleList::item, QListWidget#checkList::item {
                padding: 8px 10px;
                border-radius: 12px;
                margin: 2px 0px;
            }
            QListWidget#simpleList::item:hover, QListWidget#checkList::item:hover {
                background: rgba(200, 174, 255, 0.10);
            }
            QListWidget#simpleList::item:selected, QListWidget#checkList::item:selected {
                background: rgba(200, 174, 255, 0.16);
            }
            QToolButton {
                border: 1px solid transparent;
                border-radius: 14px;
                padding: 0 16px;
                min-height: 44px;
                max-height: 44px;
                font-weight: 750;
                color: #F3EEFF;
            }
            QToolButton:hover {
                border: 1px solid rgba(255, 255, 255, 0.10);
            }
            QToolButton[menuButton="true"] {
                padding-right: 34px;
            }
            QToolButton[menuButton="true"]::menu-indicator {
                subcontrol-origin: padding;
                subcontrol-position: center right;
                right: 12px;
            }
            QToolButton:disabled {
                background: rgba(255, 255, 255, 0.03);
                color: #6E6680;
            }
            QToolButton[kind="primary"], QToolButton[kind="solid"] {
                background: #C8AEFF;
                color: #120D1C;
            }
            QToolButton[kind="primary"]:hover, QToolButton[kind="solid"]:hover {
                background: #B592FF;
            }
            QToolButton[kind="primary"]:pressed, QToolButton[kind="solid"]:pressed {
                background: #AA83FF;
            }
            QToolButton[kind="secondary"], QToolButton[kind="ghost"], QToolButton[kind="soft"] {
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(200, 174, 255, 0.10);
            }
            QToolButton[kind="secondary"]:hover, QToolButton[kind="ghost"]:hover, QToolButton[kind="soft"]:hover {
                background: rgba(200, 174, 255, 0.14);
            }
            QToolButton[kind="secondary"]:pressed, QToolButton[kind="ghost"]:pressed, QToolButton[kind="soft"]:pressed {
                background: rgba(200, 174, 255, 0.20);
            }
            QToolButton[kind="danger"] {
                background: rgba(255, 106, 136, 0.14);
                border: 1px solid rgba(255, 106, 136, 0.18);
                color: #FFD9E2;
            }
            QToolButton[kind="danger"]:hover {
                background: rgba(255, 106, 136, 0.22);
            }
            QToolButton[kind="danger"]:pressed {
                background: rgba(255, 106, 136, 0.28);
            }
            QToolButton[panelDanger="true"] {
                background: rgba(255, 126, 148, 0.16);
                border: 1px solid rgba(255, 126, 148, 0.28);
                color: #FFE5EB;
                min-height: 48px;
                max-height: 48px;
            }
            QToolButton[panelPrimary="true"] {
                background: #C8AEFF;
                border: 1px solid rgba(255, 255, 255, 0.10);
                color: #120D1C;
                min-height: 48px;
                max-height: 48px;
                font-size: 15px;
                font-weight: 850;
                padding: 0 18px;
            }
            QToolButton[panelPrimary="true"]:hover {
                background: #B592FF;
            }
            QToolButton[panelPrimary="true"]:pressed {
                background: #AA83FF;
            }
            QToolButton[panelDanger="true"]:hover {
                background: rgba(255, 126, 148, 0.24);
            }
            QToolButton[panelDanger="true"]:pressed {
                background: rgba(255, 126, 148, 0.30);
            }
            QToolButton[navButton="true"] {
                text-align: left;
                padding: 10px 16px;
                min-height: 0px;
                max-height: 16777215px;
                border-radius: 18px;
                background: transparent;
                color: #B8B0C9;
                font-size: 15px;
            }
            QToolButton[navButton="true"]:hover {
                background: rgba(200, 174, 255, 0.10);
                color: #F3EEFF;
            }
            QToolButton[navButton="true"][active="true"] {
                background: rgba(200, 174, 255, 0.18);
                color: #F8F4FF;
            }
            QFrame#documentCard, QFrame#sopdCard, QFrame#reviewCard {
                background: #120D1C;
                border: none;
                border-radius: 22px;
            }
            QFrame#documentCard[selectedCard="true"], QFrame#sopdCard[selectedCard="true"], QFrame#reviewCard[selectedCard="true"] {
                background: rgba(200, 174, 255, 0.16);
            }
            QLabel[meta="true"] {
                color: #B8B0C9;
                font-size: 12px;
            }
            QLabel[metaKey="true"] {
                color: #9F97B4;
                font-size: 12px;
                font-weight: 700;
            }
            QLabel[formKey="true"] {
                color: #B8B0C9;
                font-size: 13px;
                font-weight: 700;
            }
            QLabel[cardTitle="true"] {
                font-size: 16px;
                font-weight: 850;
                color: #F8F4FF;
            }
            QLabel[badgeRole="status"], QLabel[badgeRole="priority"], QLabel[badgeRole="file"] {
                border-radius: 12px;
                padding: 5px 10px;
                font-size: 12px;
                font-weight: 800;
                border: 1px solid transparent;
            }
            QLabel[statusKind="draft"], QLabel[fileState="present"] {
                background: rgba(200, 174, 255, 0.16);
                color: #F3EEFF;
            }
            QLabel[statusKind="approval"], QLabel[priority="mid"] {
                background: rgba(255, 193, 94, 0.18);
                color: #FFF2D2;
            }
            QLabel[statusKind="active"] {
                background: rgba(100, 226, 174, 0.18);
                color: #E5FFF4;
            }
            QLabel[statusKind="review"], QLabel[priority="low"] {
                background: rgba(255, 152, 117, 0.20);
                color: #FFE6D9;
            }
            QLabel[statusKind="archive"] {
                background: rgba(138, 168, 255, 0.18);
                color: #E2EAFF;
            }
            QLabel[priority="high"] {
                background: rgba(255, 106, 136, 0.18);
                color: #FFD9E2;
            }
            QLabel[priority="none"] {
                background: rgba(255, 255, 255, 0.06);
                color: #D7D0E8;
            }
            QLabel[fileState="missing"] {
                background: rgba(255, 255, 255, 0.05);
                color: #B8B0C9;
            }
            QLabel[badgeRole="file"][fileState="present"][fileVariant="pdf"] {
                background: rgba(255, 78, 108, 0.28);
                border: 1px solid rgba(255, 118, 146, 0.42);
                color: #FFF1F5;
            }
            QLabel[badgeRole="file"][fileState="present"][fileVariant="doc"] {
                background: rgba(79, 151, 255, 0.30);
                border: 1px solid rgba(126, 181, 255, 0.42);
                color: #F3F8FF;
            }
            QLabel[badgeRole="file"][fileState="present"][fileVariant="excel"] {
                background: rgba(44, 201, 118, 0.30);
                border: 1px solid rgba(86, 228, 145, 0.40);
                color: #F0FFF7;
            }
            QProgressBar#analyticsBar {
                background: rgba(255, 255, 255, 0.05);
                border: none;
                border-radius: 6px;
                min-height: 12px;
                max-height: 12px;
                text-align: center;
            }
            QProgressBar#analyticsBar::chunk {
                background: #C8AEFF;
                border-radius: 6px;
            }
            QListWidget#simpleList::viewport, QListWidget#checkList::viewport {
                background: transparent;
                border: none;
            }
            QScrollArea {
                background: transparent;
                border: none;
            }
            QScrollBar:vertical {
                background: transparent;
                width: 10px;
                margin: 8px 2px 8px 0px;
            }
            QScrollBar::handle:vertical {
                background: rgba(200, 174, 255, 0.24);
                border-radius: 5px;
                min-height: 24px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical,
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical,
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal,
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                background: transparent;
                width: 0px;
                height: 0px;
            }
        """)

    def _status_style_key(self, status: str) -> str:
        mapping = {
            "Черновик": "draft",
            "На согласовании": "approval",
            "Действует": "active",
            "На пересмотре": "review",
            "Архив": "archive",
        }
        return mapping.get(self._normalize_status(status), "draft")

    def _meta_label(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setProperty("meta", True)
        label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        label.setWordWrap(True)
        return label

    def _meta_key_label(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setProperty("metaKey", True)
        label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        label.setMinimumWidth(88)
        return label

    def _form_key_label(self, text: str, width: int = 118) -> QLabel:
        label = QLabel(text)
        label.setProperty("formKey", True)
        label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        label.setMinimumWidth(width)
        return label

    def _card_title_label(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setProperty("cardTitle", True)
        label.setWordWrap(True)
        return label

    def set_active_page(self, page: str):
        if page not in self.PAGE_TITLES:
            page = "dashboard"
        if page == self._page_key:
            self.refresh_current_page()
            return
        if not self._maybe_discard_sopd_changes():
            return
        if not self._maybe_discard_document_changes():
            return
        self._page_key = page
        page_index = {
            "dashboard": 0,
            "attention": 1,
            "documents": 2,
            "sopd": 3,
            "settings": 4,
        }[page]
        self.page_stack.setCurrentIndex(page_index)
        self.page_title_label.setText(self.PAGE_TITLES[page])
        self.page_subtitle_label.setText(self.PAGE_SUBTITLES[page])
        self.page_subtitle_label.setVisible(bool(self.PAGE_SUBTITLES[page].strip()))
        self._update_nav_state()
        self._update_toolbar_state()
        self.clear_right_panel_state(page, preferred_company_id=self.current_filter_company_id())
        self.refresh_current_page()

    def _update_nav_state(self):
        for key, button in self.nav_buttons.items():
            button.setProperty("active", key == self._page_key)
            _refresh_widget_style(button)

    def _update_toolbar_state(self):
        document_scope_pages = {"dashboard", "attention", "documents"}
        show_search = self._page_key in {"dashboard", "attention", "documents", "sopd"}
        show_company = self._page_key in {"dashboard", "attention", "documents", "sopd"}
        show_status = self._page_key in document_scope_pages
        show_section = self._page_key in document_scope_pages
        show_problem = self._page_key in document_scope_pages

        self.search_input.setVisible(show_search)
        self.company_filter_combo.setVisible(show_company)
        self.status_filter_combo.setVisible(show_status)
        self.section_filter_combo.setVisible(show_section)
        self.problem_filter_combo.setVisible(show_problem)
        self.clear_filters_button.setVisible(show_search or show_company or show_status or show_section or show_problem)
        self.filters_top_row.setVisible(show_search or show_company or self.clear_filters_button.isVisible())
        self.filters_bottom_row.setVisible(show_status or show_section or show_problem)
        self.right_panel_frame.setVisible(self._page_key != "dashboard")

        if self._page_key == "dashboard":
            self.primary_action_button.setVisible(False)
        elif self._page_key == "attention":
            self.primary_action_button.setVisible(True)
            self.primary_action_button.setText("Добавить документ")
        elif self._page_key == "documents":
            self.primary_action_button.setVisible(True)
            self.primary_action_button.setText("Добавить документ")
        elif self._page_key == "sopd":
            self.primary_action_button.setVisible(True)
            self.primary_action_button.setText("Добавить карточку")
        elif self._page_key == "settings":
            self.primary_action_button.setVisible(True)
            self.primary_action_button.setText("Добавить компанию")
        else:
            self.primary_action_button.setVisible(False)

        self.search_input.setPlaceholderText(self._search_placeholder_for_page())
        self._populate_status_filter_options()
        self._populate_section_filter()
        self._populate_problem_filter_options()

    def _search_placeholder_for_page(self) -> str:
        return {
            "dashboard": "Найти документ, компанию или раздел",
            "attention": "Найти документ, проблему, компанию или раздел",
            "documents": "Найти документ, компанию или раздел",
            "sopd": "Найти карточку СОПД",
        }.get(self._page_key, "Поиск")

    def _show_catalog_loading_state(self):
        page = (self._page_key or "dashboard").strip().lower()
        if page == "dashboard":
            self.dashboard_scope_label.setText("Загружаем сводку по документам…")
            self.dashboard_scope_label.show()
            self.dashboard_scope_note.setText("Главный обзор появится сразу после загрузки каталога документов.")
            self.dashboard_scope_note.show()
            for value_label, note_label in self.metric_cards.values():
                value_label.setText("—")
                note_label.setText("Загрузка данных…")
                note_label.show()
            self._clear_layout_widgets(self.dashboard_attention_layout)
            self._add_dashboard_hint(self.dashboard_attention_layout, "Подбираем документы, которым нужно внимание.")
            self.dashboard_attention_layout.addStretch(1)
            self._clear_layout_widgets(self.dashboard_recent_layout)
            self._add_dashboard_hint(self.dashboard_recent_layout, "Собираем последние изменения по документам.")
            self.dashboard_recent_layout.addStretch(1)
            return

        if page == "attention":
            self.attention_urgent_summary.setText("Срочно: …")
            self.attention_issue_summary.setText("Нужно поправить: …")
            self.attention_recent_summary.setText("Недавно изменены: …")
            self._render_empty_list_state(
                self.attention_urgent_layout,
                "Загружаем документы",
                "Сейчас соберем список срочных проблем по документам.",
            )
            self._render_empty_list_state(
                self.attention_issue_layout,
                "Подготавливаем замечания",
                "Проверяем документы без дат, разделов и обязательных файлов.",
            )
            self._render_empty_list_state(
                self.attention_recent_layout,
                "Собираем изменения",
                "Через пару секунд здесь появятся последние обновления.",
            )
            return

        if page == "documents":
            self.documents_count_label.setText("Загружаем документы…")
            self._render_empty_list_state(
                self.documents_layout,
                "Загружаем список документов",
                "Каталог документов появится сразу после завершения стартовой загрузки.",
            )

    def _warm_initial_document_catalog(self):
        if not self.db or self._startup_catalog_loading or self._document_catalog_cache is not None:
            self._startup_catalog_pending = False if self._document_catalog_cache is not None else self._startup_catalog_pending
            return

        self._startup_catalog_loading = True

        def work():
            return [dict(row) for row in self.db.list_document_catalog()]

        def on_success(rows: Any):
            self._startup_catalog_loading = False
            self._startup_catalog_pending = False
            catalog_rows = list(rows or [])
            self._startup_catalog_rows = catalog_rows or None
            self._document_catalog_cache = list(catalog_rows) if catalog_rows else []
            self.refresh_current_page()

        def on_failure(msg: str):
            self._startup_catalog_loading = False
            self._startup_catalog_pending = False
            self.warn(f"Не удалось загрузить список документов:\n{msg}")

        self._run_background_task(
            label="Загрузка списка документов",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не удалось загрузить список документов:",
            show_busy_state=False,
            reject_if_busy=False,
            on_failure=on_failure,
        )

    def _populate_status_filter_options(self):
        current = self.status_filter_combo.currentData() if self.status_filter_combo.count() else None
        self.status_filter_combo.blockSignals(True)
        self.status_filter_combo.clear()
        self.status_filter_combo.addItem("Все статусы", None)
        for status in DOC_STATUSES:
            if status == "Архив":
                continue
            self.status_filter_combo.addItem(status, status)
        index = 0
        if current is not None:
            for i in range(self.status_filter_combo.count()):
                if self.status_filter_combo.itemData(i) == current:
                    index = i
                    break
        self.status_filter_combo.setCurrentIndex(index)
        self._set_combo_tooltips(self.status_filter_combo)
        self.status_filter_combo.blockSignals(False)

    def _populate_problem_filter_options(self):
        current = self.problem_filter_combo.currentData() if self.problem_filter_combo.count() else None
        self.problem_filter_combo.blockSignals(True)
        self.problem_filter_combo.clear()
        self.problem_filter_combo.addItem("Все проблемы", None)
        problem_items = [
            ("Без PDF", "missing_pdf"),
            ("Нужен доп. файл, но его нет", "missing_required_office"),
            ("Просрочено", "overdue_review"),
            ("Скоро пересмотр", "upcoming_review"),
            ("Без даты пересмотра", "missing_review"),
            ("Без раздела", "missing_section"),
            ("Недавно изменённые", "recent_update"),
        ]
        for text, value in problem_items:
            self.problem_filter_combo.addItem(text, value)
        index = 0
        if current is not None:
            for i in range(self.problem_filter_combo.count()):
                if self.problem_filter_combo.itemData(i) == current:
                    index = i
                    break
        self.problem_filter_combo.setCurrentIndex(index)
        self._set_combo_tooltips(self.problem_filter_combo)
        self.problem_filter_combo.blockSignals(False)

    def _on_status_filter_changed(self):
        if not self._maybe_discard_sopd_changes() or not self._maybe_discard_document_changes():
            self.status_filter_combo.blockSignals(True)
            self._set_combo_value(self.status_filter_combo, self._last_status_filter_value)
            self.status_filter_combo.blockSignals(False)
            return
        self._last_status_filter_value = self.status_filter_combo.currentData()
        self.clear_right_panel_state(self._page_key, preferred_company_id=self.current_filter_company_id())
        self.refresh_current_page()

    def _on_company_filter_changed(self):
        if not self._maybe_discard_sopd_changes() or not self._maybe_discard_document_changes():
            self.company_filter_combo.blockSignals(True)
            self._set_combo_value(self.company_filter_combo, self._last_company_filter_value)
            self.company_filter_combo.blockSignals(False)
            self._populate_section_filter()
            self.section_filter_combo.blockSignals(True)
            self._set_combo_value(self.section_filter_combo, self._last_section_filter_value)
            self.section_filter_combo.blockSignals(False)
            return
        self._last_company_filter_value = self.current_filter_company_id()
        self._populate_section_filter()
        self.clear_right_panel_state(self._page_key, preferred_company_id=self.current_filter_company_id())
        self.refresh_current_page()

    def _on_section_filter_changed(self):
        if not self._maybe_discard_sopd_changes() or not self._maybe_discard_document_changes():
            self.section_filter_combo.blockSignals(True)
            self._set_combo_value(self.section_filter_combo, self._last_section_filter_value)
            self.section_filter_combo.blockSignals(False)
            return
        current_section = self.section_filter_combo.currentData() if self.section_filter_combo.count() else None
        self._last_section_filter_value = int(current_section) if current_section not in (None, "") else None
        self.clear_right_panel_state(self._page_key, preferred_company_id=self.current_filter_company_id())
        self.refresh_current_page()

    def _on_problem_filter_changed(self):
        if not self._maybe_discard_sopd_changes() or not self._maybe_discard_document_changes():
            self.problem_filter_combo.blockSignals(True)
            self._set_combo_value(self.problem_filter_combo, self._last_problem_filter_value)
            self.problem_filter_combo.blockSignals(False)
            return
        self._last_problem_filter_value = self.problem_filter_combo.currentData()
        self.clear_right_panel_state(self._page_key, preferred_company_id=self.current_filter_company_id())
        self.refresh_current_page()

    def current_company_id_name(self):
        company_id = self.current_filter_company_id()
        if company_id is None:
            return None, "Все компании"
        for i in range(self.company_filter_combo.count()):
            if self.company_filter_combo.itemData(i) == company_id:
                return company_id, self.company_filter_combo.itemText(i).split("  ·  ")[0]
        return company_id, None

    def current_filter_company_id(self) -> Optional[int]:
        data = self.company_filter_combo.currentData() if hasattr(self, "company_filter_combo") else None
        return int(data) if data not in (None, "") else None

    def reset_filters(self):
        self.search_input.blockSignals(True)
        self.search_input.setText("")
        self.search_input.blockSignals(False)
        if self.company_filter_combo.count():
            self.company_filter_combo.blockSignals(True)
            self.company_filter_combo.setCurrentIndex(0)
            self.company_filter_combo.blockSignals(False)
        if self.status_filter_combo.count():
            self.status_filter_combo.blockSignals(True)
            self.status_filter_combo.setCurrentIndex(0)
            self.status_filter_combo.blockSignals(False)
        self._populate_section_filter()
        if self.problem_filter_combo.count():
            self.problem_filter_combo.blockSignals(True)
            self.problem_filter_combo.setCurrentIndex(0)
            self.problem_filter_combo.blockSignals(False)
        self._last_company_filter_value = self.current_filter_company_id()
        self._last_status_filter_value = self.status_filter_combo.currentData() if self.status_filter_combo.count() else None
        self._last_section_filter_value = self.section_filter_combo.currentData() if self.section_filter_combo.count() else None
        self._last_problem_filter_value = self.problem_filter_combo.currentData() if self.problem_filter_combo.count() else None
        self.clear_right_panel_state(self._page_key, preferred_company_id=self.current_filter_company_id())
        self.refresh_current_page()

    def _handle_primary_action(self):
        if self._page_key in {"attention", "documents"}:
            self.start_new_document()
        elif self._page_key == "sopd":
            self.start_new_sopd_record()
        elif self._page_key == "settings":
            self.settings_add_company()

    def _default_document_company_id(self) -> Optional[int]:
        company_id = self.current_filter_company_id()
        if company_id is not None:
            return company_id
        if hasattr(self, "doc_company_combo") and self.doc_company_combo.count():
            data = self.doc_company_combo.currentData()
            if data not in (None, ""):
                return int(data)
            first = self.doc_company_combo.itemData(0)
            if first not in (None, ""):
                return int(first)
        return None

    def _build_bootstrap_payload(self) -> Dict[str, Any]:
        db = Db(self.db_path, self.current_account)
        company_rows = [dict(row) for row in db.list_companies_with_stats()]
        due_reviews_count = db.count_due_reviews(today_ymd())
        return {
            "db": db,
            "company_rows": company_rows,
            "due_reviews_count": due_reviews_count,
        }

    def _finish_bootstrap(self, payload: Dict[str, Any]):
        self.db = payload["db"]
        self._startup_catalog_rows = None
        self._document_catalog_cache = None
        self._update_statusbar()
        company_rows = list(payload.get("company_rows") or [])
        self._startup_catalog_pending = bool(company_rows)
        self.load_companies(
            initial=True,
            rows=company_rows,
            refresh_page=not self._startup_catalog_pending,
        )
        self.refresh_notifications(payload.get("due_reviews_count"))
        if self._startup_catalog_pending:
            self._show_catalog_loading_state()
            self._warm_initial_document_catalog()

    def bootstrap(self):
        try:
            self.shared_root = self._load_or_pick_shared_root()
            self.db_path = os.path.join(self.shared_root, "pd_docs.db")
            self.storage_dir = os.path.join(self.shared_root, ".pd_docs")
            ensure_dir(self.storage_dir)
            self._try_hide_windows_dir(self.storage_dir)
            self._due_reviews_count = 0
            self.switch_app_page(0)
            self._update_statusbar()
            self._run_background_task(
                label="Загрузка рабочего пространства",
                work_fn=self._build_bootstrap_payload,
                on_success=self._finish_bootstrap,
                on_error_prefix="Не удалось запуститься:",
            )
        except Exception as e:
            self.warn(f"Не удалось запуститься:\n{e}")

    def load_companies(
        self,
        initial: bool = False,
        preferred_company_id: Optional[int] = None,
        rows: Optional[List[Dict[str, Any]]] = None,
        refresh_page: bool = True,
    ):
        if not self.db and rows is None:
            return
        if not initial:
            self._document_catalog_cache = None
        rows = list(rows) if rows is not None else [dict(row) for row in self.db.list_companies_with_stats()]
        if not rows:
            self.switch_app_page(0)
            self.current_document_id = None
            self.current_sopd_id = None
            self._show_right_placeholder("Пусто", "Создайте первую компанию, чтобы начать работу.")
            return

        self.switch_app_page(1)
        previous_company_id = self.current_filter_company_id()
        target_company_id = preferred_company_id if preferred_company_id is not None else previous_company_id

        self.company_filter_combo.blockSignals(True)
        self.company_filter_combo.clear()
        self.company_filter_combo.addItem("Все компании", None)
        selected_index = 0
        for idx, row in enumerate(rows, start=1):
            company_id = int(row["id"])
            self.company_filter_combo.addItem(str(row["name"]), company_id)
            if target_company_id == company_id:
                selected_index = idx
        self.company_filter_combo.setCurrentIndex(selected_index)
        self._set_combo_tooltips(self.company_filter_combo)
        self.company_filter_combo.blockSignals(False)
        self._last_company_filter_value = self.current_filter_company_id()

        self.doc_company_combo.blockSignals(True)
        self.doc_company_combo.clear()
        self.sopd_company_combo.blockSignals(True)
        self.sopd_company_combo.clear()
        for row in rows:
            company_id = int(row["id"])
            name = str(row["name"])
            self.doc_company_combo.addItem(name, company_id)
            self.sopd_company_combo.addItem(name, company_id)
        self._set_combo_tooltips(self.doc_company_combo)
        self._set_combo_tooltips(self.sopd_company_combo)
        self.doc_company_combo.blockSignals(False)
        self.sopd_company_combo.blockSignals(False)

        self.settings_company_list.blockSignals(True)
        self.settings_company_list.clear()
        self._settings_company_ids = []
        for row in rows:
            company_id = int(row["id"])
            item = QListWidgetItem(f"{row['name']}  ·  {int(row['docs_count'] or 0)} док. / {int(row['sections_count'] or 0)} разд.")
            item.setData(Qt.UserRole, company_id)
            item.setData(Qt.UserRole + 1, str(row["name"]))
            self.settings_company_list.addItem(item)
            self._settings_company_ids.append(company_id)
        if hasattr(self, "settings_company_empty_label"):
            self.settings_company_empty_label.setVisible(not bool(rows))
        if self.settings_company_list.count():
            target_settings = target_company_id if target_company_id in self._settings_company_ids else self._settings_company_ids[0]
            for i in range(self.settings_company_list.count()):
                if self.settings_company_list.item(i).data(Qt.UserRole) == target_settings:
                    self.settings_company_list.setCurrentRow(i)
                    break
        self.settings_company_list.blockSignals(False)

        self._populate_section_filter()
        section_value = self.section_filter_combo.currentData() if self.section_filter_combo.count() else None
        self._last_section_filter_value = int(section_value) if section_value not in (None, "") else None
        self._last_status_filter_value = self.status_filter_combo.currentData() if self.status_filter_combo.count() else None
        self._last_problem_filter_value = self.problem_filter_combo.currentData() if self.problem_filter_combo.count() else None
        self._update_settings_summary()
        if refresh_page:
            self.refresh_current_page()
        if initial:
            self._fade_in(self.page_main, 220)

    def _populate_section_filter(self):
        current_section_id = self.section_filter_combo.currentData() if self.section_filter_combo.count() else None
        company_id = self.current_filter_company_id()
        self.section_filter_combo.blockSignals(True)
        self.section_filter_combo.clear()
        self.section_filter_combo.addItem("Все разделы", None)
        self.section_filter_combo.setEnabled(company_id is not None)
        if company_id is not None and self.db:
            self.section_filter_combo.addItem("Без раздела", -1)
            for row in self.db.list_sections(company_id):
                self.section_filter_combo.addItem(str(row["name"]), int(row["id"]))
        index = 0
        if current_section_id is not None:
            for i in range(self.section_filter_combo.count()):
                if self.section_filter_combo.itemData(i) == current_section_id:
                    index = i
                    break
        self.section_filter_combo.setCurrentIndex(index)
        self._set_combo_tooltips(self.section_filter_combo)
        self.section_filter_combo.blockSignals(False)

    def clear_right_panel_state(self, target_page: Optional[str] = None, preferred_company_id: Optional[int] = None):
        page = (target_page or self._page_key or "documents").strip().lower()
        target_company_id = preferred_company_id
        if target_company_id is None:
            target_company_id = self.current_filter_company_id()

        self._clear_document_editor(preferred_company_id=target_company_id, show_editor=False)
        self._clear_sopd_editor(preferred_company_id=target_company_id, show_editor=False)

        if page == "settings":
            self._set_right_actions_mode("none")
            self.right_stack.setCurrentWidget(self.right_settings_page)
            self.right_panel_title.setText("Рабочее пространство")
            self.right_panel_subtitle.setText("")
            return

        placeholder_map = {
            "dashboard": ("Карточка документа", ""),
            "attention": ("Карточка документа", ""),
            "documents": ("Карточка документа", ""),
            "sopd": ("Карточка СОПД", ""),
        }
        title, text = placeholder_map.get(page, ("Детали", ""))
        self._show_right_placeholder(title, text)

    def refresh_notifications(self, due_reviews_count: Optional[int] = None):
        if not self.db and due_reviews_count is None:
            return
        if due_reviews_count is None:
            self._due_reviews_count = self.db.count_due_reviews(today_ymd())
        else:
            self._due_reviews_count = max(0, int(due_reviews_count))
        self._update_statusbar()

    def refresh_current_page(self):
        if (
            self._startup_catalog_pending
            and self._document_catalog_cache is None
            and self._page_key in {"dashboard", "attention", "documents"}
        ):
            self._show_catalog_loading_state()
            if not self._startup_catalog_loading:
                self._warm_initial_document_catalog()
            return
        if self._page_key == "dashboard":
            self.refresh_dashboard_page()
        elif self._page_key == "attention":
            self.refresh_attention_page()
        elif self._page_key == "documents":
            self.refresh_documents_page()
        elif self._page_key == "sopd":
            self.refresh_sopd_page()
        elif self._page_key == "settings":
            self.refresh_settings_page()

    def _path_exists(self, stored_path: Optional[str]) -> bool:
        abs_path = self._abs_storage_path(stored_path)
        return bool(abs_path and os.path.exists(abs_path))

    def _document_has_files(self, row: sqlite3.Row) -> bool:
        return self._path_exists(row["pdf_path"]) or self._path_exists(row["office_path"])

    def _row_has_section(self, row: sqlite3.Row, section_id: Optional[int]) -> bool:
        if section_id is None:
            return True
        raw = str(row["section_ids"] or "").strip()
        if section_id == -1:
            return not raw
        if not raw:
            return False
        values = {part.strip() for part in raw.split(",") if part.strip()}
        return str(section_id) in values

    def _document_requires_office(self, row: sqlite3.Row) -> bool:
        try:
            return bool(int(row["needs_office"] or 0))
        except Exception:
            return bool(row.get("needs_office")) if isinstance(row, dict) else False

    def _document_has_office_file(self, row: sqlite3.Row) -> bool:
        return self._path_exists(row["office_path"])

    def _document_missing_required_office(self, row: sqlite3.Row) -> bool:
        return self._document_requires_office(row) and not self._document_has_office_file(row)

    def _document_missing_status(self, row: sqlite3.Row) -> bool:
        return not str(row["status"] or "").strip()

    def _document_missing_review_date(self, row: sqlite3.Row) -> bool:
        return not parse_date_to_ymd(str(row["review_due"] or "").strip())

    def _document_missing_section(self, row: sqlite3.Row) -> bool:
        return not str(row["section_ids"] or "").strip()

    def _row_matches_problem_filter(self, row: sqlite3.Row, problem_key: Optional[str]) -> bool:
        if not problem_key:
            return True
        if problem_key == "missing_pdf":
            return not self._document_has_required_pdf(row)
        if problem_key == "missing_required_office":
            return self._document_missing_required_office(row)
        if problem_key == "overdue_review":
            return self._review_priority(row) == "high"
        if problem_key == "upcoming_review":
            return self._review_priority(row) in {"mid", "low"}
        if problem_key == "missing_review":
            return self._document_missing_review_date(row)
        if problem_key == "missing_section":
            return self._document_missing_section(row)
        if problem_key == "recent_update":
            return self._is_recent_update(row["updated_at"])
        return True

    def _review_priority(self, row: sqlite3.Row) -> Optional[str]:
        due_raw = (row["review_due"] or "").strip()
        due = parse_date_to_ymd(due_raw)
        if not due:
            return None
        today = today_ymd()
        if due < today:
            return "high"
        try:
            delta_days = (datetime.fromisoformat(due) - datetime.fromisoformat(today)).days
        except Exception:
            return None
        if delta_days <= 7:
            return "mid"
        if delta_days <= 30:
            return "low"
        return None

    def _is_recent_update(self, updated_at: Optional[str], days: int = 7) -> bool:
        raw = (updated_at or "").strip()
        if not raw:
            return False
        try:
            return datetime.fromisoformat(raw) >= datetime.now() - timedelta(days=days)
        except Exception:
            return False

    def _document_rows_for_page(self, page: Optional[str] = None) -> List[sqlite3.Row]:
        page = page or self._page_key
        if not self.db:
            return []
        query = self.search_input.text().strip().lower()
        company_id = self.current_filter_company_id()
        status_value = self.status_filter_combo.currentData() if self.status_filter_combo.isVisible() else None
        section_id = self.section_filter_combo.currentData() if self.section_filter_combo.isVisible() else None
        problem_value = self.problem_filter_combo.currentData() if self.problem_filter_combo.isVisible() else None

        rows = self._all_document_catalog_rows()
        filtered: List[sqlite3.Row] = []
        for row in rows:
            status = self._normalize_status(row["status"])
            if page == "dashboard" and status == "Архив":
                continue
            if page == "attention" and status == "Архив":
                continue
            if page == "documents" and status == "Архив" and status_value != "Архив":
                continue
            if page == "attention":
                severity, _ = self._document_attention_details(row)
                if severity <= 0 and not self._is_recent_update(row["updated_at"]):
                    continue
            if company_id is not None and int(row["company_id"]) != int(company_id):
                continue
            if status_value and status != status_value:
                continue
            if section_id is not None and not self._row_has_section(row, int(section_id)):
                continue
            if not self._row_matches_problem_filter(row, problem_value):
                continue
            haystack = " ".join([
                str(row["doc_title"] or ""),
                str(row["company_name"] or ""),
                str(row["section_names"] or ""),
                status,
                str(row["comment"] or ""),
            ]).lower()
            if query and query not in haystack:
                continue
            filtered.append(row)

        if page == "attention":
            filtered.sort(
                key=lambda row: (
                    -self._document_attention_details(row)[0],
                    0 if self._is_recent_update(row["updated_at"]) else 1,
                    {"high": 0, "mid": 1, "low": 2}.get(self._review_priority(row) or "", 3),
                    row["review_due"] or "",
                    str(row["company_name"] or "").lower(),
                    str(row["doc_title"] or "").lower(),
                )
            )
        elif page == "review":
            filtered.sort(
                key=lambda row: (
                    {"high": 0, "mid": 1, "low": 2}.get(self._review_priority(row) or "low", 3),
                    (row["review_due"] or ""),
                    str(row["company_name"] or ""),
                    int(row["sort_order"] or 0),
                )
            )
        else:
            filtered.sort(key=lambda row: (str(row["company_name"] or "").lower(), int(row["sort_order"] or 0), int(row["id"])))
        return filtered

    def _all_document_catalog_rows(self) -> List[Any]:
        if self._document_catalog_cache is not None:
            return list(self._document_catalog_cache)
        if self._startup_catalog_rows is not None:
            rows = list(self._startup_catalog_rows)
            self._startup_catalog_rows = None
            self._document_catalog_cache = list(rows)
            return rows
        if not self.db:
            return []
        rows = [dict(row) for row in self.db.list_document_catalog()]
        self._document_catalog_cache = list(rows)
        return rows

    def _render_empty_list_state(
        self,
        layout: QVBoxLayout,
        title: str,
        text: str,
        action_text: Optional[str] = None,
        action_callback=None,
    ):
        self._clear_layout_widgets(layout)
        empty = QFrame()
        empty.setObjectName("contentCard")
        empty_l = QVBoxLayout(empty)
        empty_l.setContentsMargins(24, 24, 24, 24)
        empty_l.setSpacing(8)
        empty_title = QLabel(title)
        empty_title.setObjectName("emptyPanelTitle")
        empty_title.setAlignment(Qt.AlignCenter)
        empty_l.addWidget(empty_title)
        empty_text = QLabel(text)
        empty_text.setObjectName("emptyPanelText")
        empty_text.setWordWrap(True)
        empty_text.setAlignment(Qt.AlignCenter)
        empty_l.addWidget(empty_text)
        if action_text and action_callback is not None:
            action_btn = self._create_button(action_text, kind="soft")
            action_btn.clicked.connect(action_callback)
            empty_l.addWidget(action_btn, 0, Qt.AlignCenter)
        layout.addWidget(empty)
        layout.addStretch(1)

    def _set_combo_value(self, combo: Any, value: Any, fallback_index: int = 0) -> bool:
        for i in range(combo.count()):
            if combo.itemData(i) == value:
                combo.setCurrentIndex(i)
                return True
        if combo.count():
            combo.setCurrentIndex(min(fallback_index, combo.count() - 1))
        return False

    def _set_combo_tooltips(self, combo: Any):
        for i in range(combo.count()):
            combo.setItemData(i, combo.itemText(i), Qt.ToolTipRole)
        combo.setToolTip(combo.currentText().strip())

    def _current_section_id_name(self) -> Tuple[Optional[int], str]:
        section_id = self.section_filter_combo.currentData() if self.section_filter_combo.count() else None
        if section_id is None:
            return None, "все разделы"
        return int(section_id), self.section_filter_combo.currentText()

    def _dashboard_sopd_rows(self, company_id: Optional[int]) -> List[sqlite3.Row]:
        if not self.db:
            return []
        if company_id is not None:
            return list(self.db.list_sopd_records(company_id))
        rows: List[sqlite3.Row] = []
        for company in self.db.list_companies():
            rows.extend(self.db.list_sopd_records(int(company["id"])))
        return rows

    def _add_dashboard_hint(self, layout: QVBoxLayout, text: str):
        hint = QLabel(text)
        hint.setObjectName("summaryText")
        hint.setWordWrap(True)
        layout.addWidget(hint)

    def _build_dashboard_breakdown_row(
        self,
        title: str,
        value: int,
        total: int,
        note: str,
        action_text: Optional[str] = None,
        action_callback=None,
    ) -> QFrame:
        row = QFrame()
        row.setObjectName("analyticsRow")
        lay = QVBoxLayout(row)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(8)

        head = QHBoxLayout()
        head.setSpacing(10)
        title_lbl = QLabel(title)
        title_lbl.setObjectName("analyticsTitle")
        title_lbl.setWordWrap(True)
        head.addWidget(title_lbl, 1)
        value_lbl = QLabel(str(value))
        value_lbl.setObjectName("analyticsCount")
        head.addWidget(value_lbl, 0, Qt.AlignTop)
        lay.addLayout(head)

        bar = QProgressBar()
        bar.setObjectName("analyticsBar")
        bar.setRange(0, max(total, 1))
        bar.setValue(max(0, min(value, max(total, 1))))
        bar.setTextVisible(False)
        lay.addWidget(bar)

        note_lbl = QLabel(note)
        note_lbl.setObjectName("analyticsHint")
        note_lbl.setWordWrap(True)
        lay.addWidget(note_lbl)

        if action_text and action_callback is not None:
            actions = QHBoxLayout()
            actions.setSpacing(8)
            actions.addStretch(1)
            action_btn = self._create_button(action_text, kind="ghost")
            action_btn.clicked.connect(action_callback)
            actions.addWidget(action_btn)
            lay.addLayout(actions)
        return row

    def _chart_color_for_status(self, status: str) -> str:
        normalized = self._normalize_status(status)
        if normalized == "Действует":
            return "#58D695"
        if normalized == "На согласовании":
            return "#F6C46F"
        if normalized == "На пересмотре":
            return "#FF9B74"
        return THEME_ACCENT

    def _focus_dashboard_company(self, company_id: int):
        if self.company_filter_combo.count():
            self._set_combo_value(self.company_filter_combo, company_id)

    def _focus_dashboard_section(self, company_id: Optional[int], section_id: Optional[int]):
        if company_id is not None and self.company_filter_combo.count():
            self._set_combo_value(self.company_filter_combo, company_id)
        self._populate_section_filter()
        if section_id is not None and self.section_filter_combo.count():
            self._set_combo_value(self.section_filter_combo, section_id)
        self.refresh_current_page()

    def _open_documents_for_status(self, status: str):
        self.set_active_page("documents")
        self._set_combo_value(self.status_filter_combo, status)
        self.refresh_current_page()

    def _open_sopd_company_view(self, company_id: Optional[int]):
        if company_id is not None and self.company_filter_combo.count():
            self._set_combo_value(self.company_filter_combo, company_id)
        self.set_active_page("sopd")

    def refresh_attention_page(self):
        if not self.db:
            return
        self.attention_page.setUpdatesEnabled(False)
        try:
            rows = self._document_rows_for_page("attention")
            urgent_payload: List[Tuple[int, List[str], sqlite3.Row]] = []
            issue_payload: List[Tuple[int, List[str], sqlite3.Row]] = []
            recent_rows = [row for row in rows if self._is_recent_update(row["updated_at"])]
            recent_rows.sort(key=lambda row: (row["updated_at"] or ""), reverse=True)

            for row in rows:
                severity, issues = self._document_attention_details(row)
                if severity <= 0 or not issues:
                    continue
                bucket = urgent_payload if severity >= 4 else issue_payload
                bucket.append((severity, issues, row))

            for payload in (urgent_payload, issue_payload):
                payload.sort(
                    key=lambda item: (
                        -item[0],
                        {"high": 0, "mid": 1, "low": 2}.get(self._review_priority(item[2]) or "", 3),
                        item[2]["review_due"] or "",
                        str(item[2]["company_name"] or "").lower(),
                        str(item[2]["doc_title"] or "").lower(),
                    )
                )

            self.attention_urgent_summary.setText(f"Срочно: {len(urgent_payload)}")
            self.attention_issue_summary.setText(f"Нужно поправить: {len(issue_payload)}")
            self.attention_recent_summary.setText(f"Недавно изменены: {len(recent_rows)}")

            self._review_cards = {}
            self._clear_layout_widgets(self.attention_urgent_layout)
            if urgent_payload:
                for severity, issues, row in urgent_payload[:8]:
                    card = self._build_attention_card(row, issues)
                    self._review_cards[int(row["id"])] = card
                    self.attention_urgent_layout.addWidget(card)
            else:
                self._add_dashboard_hint(self.attention_urgent_layout, "Сейчас нет срочных документов.")
            self.attention_urgent_layout.addStretch(1)

            self._clear_layout_widgets(self.attention_issue_layout)
            if issue_payload:
                for severity, issues, row in issue_payload[:10]:
                    card = self._build_attention_card(row, issues)
                    self._review_cards[int(row["id"])] = card
                    self.attention_issue_layout.addWidget(card)
            else:
                self._add_dashboard_hint(self.attention_issue_layout, "Здесь появятся документы без раздела, даты или обязательных файлов.")
            self.attention_issue_layout.addStretch(1)

            self._clear_layout_widgets(self.attention_recent_layout)
            if recent_rows:
                for row in recent_rows[:8]:
                    self.attention_recent_layout.addWidget(self._build_recent_change_row(row))
            else:
                self._add_dashboard_hint(self.attention_recent_layout, "Недавних изменений пока нет.")
            self.attention_recent_layout.addStretch(1)

            visible_ids = {int(row["id"]) for row in rows}
            if self.current_document_id not in visible_ids and self._page_key == "attention":
                self.clear_right_panel_state("attention", preferred_company_id=self.current_filter_company_id())
            else:
                self._sync_document_card_selection()
        finally:
            self.attention_page.setUpdatesEnabled(True)
            self.attention_page.update()

    def refresh_dashboard_page(self):
        if not self.db:
            return
        self.dashboard_page.setUpdatesEnabled(False)
        try:
            rows = self._document_rows_for_page("dashboard")
            _, company_name = self.current_company_id_name()
            _, section_name = self._current_section_id_name()
            query = self.search_input.text().strip()

            scope_text = f"Сейчас показываем: {company_name or 'все компании'} · {section_name}"
            if query:
                scope_text += f" · поиск: {query}"
            self.dashboard_scope_label.setText(scope_text)
            self.dashboard_scope_label.show()
            self.dashboard_scope_note.hide()

            total_docs = len(rows)
            missing_pdf_count = sum(1 for row in rows if not self._document_has_required_pdf(row))
            overdue_rows = [row for row in rows if self._review_priority(row) == "high"]
            upcoming_rows = [row for row in rows if self._review_priority(row) in {"mid", "low"}]
            recent_rows = sorted(rows, key=lambda row: (row["updated_at"] or ""), reverse=True)[:6]

            metrics = {
                "docs_total": (str(total_docs), "Во всех выбранных компаниях и разделах."),
                "docs_missing": (str(missing_pdf_count), "Документы без основного подписанного PDF."),
                "docs_overdue": (str(len(overdue_rows)), "Их уже пора пересмотреть."),
                "docs_upcoming": (str(len(upcoming_rows)), "Срок пересмотра наступит в ближайшие 30 дней."),
            }
            for key, (value, note) in metrics.items():
                if key in self.metric_cards:
                    self.metric_cards[key][0].setText(value)
                    self.metric_cards[key][1].setText(note)
                    self.metric_cards[key][1].setVisible(bool(note.strip()))

            bucket_order = ["Просрочено", "Без PDF", "Скоро пересмотр", "Без даты", "В порядке"]
            bucket_colors = {
                "Просрочено": "#FF7A88",
                "Без PDF": "#F66C7B",
                "Скоро пересмотр": "#FFB86C",
                "Без даты": "#7E7891",
                "В порядке": "#58D695",
            }
            bucket_counts = {key: 0 for key in bucket_order}
            for row in rows:
                bucket_counts[self._dashboard_priority_bucket(row)] += 1
            self.dashboard_main_chart.set_series([
                (label, bucket_counts[label], bucket_colors[label])
                for label in bucket_order
            ])

            attention_payload: List[Tuple[int, List[str], sqlite3.Row]] = []
            for row in rows:
                severity, issues = self._document_attention_details(row)
                if severity > 0 and issues:
                    attention_payload.append((severity, issues, row))
            attention_payload.sort(
                key=lambda item: (
                    -item[0],
                    {"high": 0, "mid": 1, "low": 2}.get(self._review_priority(item[2]) or "", 3),
                    item[2]["review_due"] or "",
                    str(item[2]["company_name"] or "").lower(),
                    str(item[2]["doc_title"] or "").lower(),
                )
            )

            self._review_cards = {}
            self._clear_layout_widgets(self.dashboard_attention_layout)
            if attention_payload:
                for severity, issues, row in attention_payload[:6]:
                    card = self._build_attention_card(row, issues)
                    self._review_cards[int(row["id"])] = card
                    self.dashboard_attention_layout.addWidget(card)
            else:
                self._add_dashboard_hint(self.dashboard_attention_layout, "Сейчас нет документов с явными проблемами.")
            self.dashboard_attention_layout.addStretch(1)

            self._clear_layout_widgets(self.dashboard_recent_layout)
            if recent_rows:
                for row in recent_rows:
                    self.dashboard_recent_layout.addWidget(self._build_recent_change_row(row))
            else:
                self._add_dashboard_hint(self.dashboard_recent_layout, "Недавних изменений пока нет.")
            self.dashboard_recent_layout.addStretch(1)

            self._sync_document_card_selection()
        finally:
            self.dashboard_page.setUpdatesEnabled(True)
            self.dashboard_page.update()

    def _open_document_from_dashboard(self, doc_id: int):
        if not self._maybe_discard_document_changes():
            return
        self.set_active_page("documents")
        self.select_document(doc_id)

    def _start_review_update(self, doc_id: int):
        if self._page_key != "documents":
            self.set_active_page("documents")
        self.select_document(doc_id)
        self.doc_review_date_edit.setFocus()
        self.doc_review_date_edit.selectAll()

    def _quick_upload_document_file(self, doc_id: int, kind: str):
        if not self._maybe_discard_document_changes():
            return
        if self._page_key != "documents":
            self.set_active_page("documents")
        self.select_document(doc_id)
        self.upload_file_dialog(doc_id, kind)

    def _build_status_badge(self, status: str) -> QLabel:
        label = QLabel(status)
        label.setProperty("badgeRole", "status")
        label.setProperty("statusKind", self._status_style_key(status))
        label.setAlignment(Qt.AlignCenter)
        label.setMinimumHeight(28)
        return label

    def _stored_path_ext(self, stored_path: Optional[str]) -> str:
        if not stored_path:
            return ""
        _, ext = os.path.splitext(str(stored_path))
        return ext.lower()

    def _document_has_required_pdf(self, row: sqlite3.Row) -> bool:
        return self._path_exists(row["pdf_path"])

    def _office_file_label(self, stored_path: Optional[str], fallback: str = "Word или Excel") -> str:
        ext = self._stored_path_ext(stored_path)
        labels = {
            ".doc": "DOC",
            ".docx": "DOCX",
            ".xls": "XLS",
            ".xlsx": "XLSX",
        }
        return labels.get(ext, fallback)

    def _office_file_variant(self, stored_path: Optional[str]) -> str:
        ext = self._stored_path_ext(stored_path)
        if ext in {".xls", ".xlsx"}:
            return "excel"
        return "doc"

    def _set_file_tag_state(self, label: QLabel, text: str, variant: str, ready: bool):
        label.setText(text)
        label.setProperty("fileTag", "true")
        label.setProperty("variant", variant)
        label.setProperty("state", "ready" if ready else "missing")
        _refresh_widget_style(label)

    def _set_status_badge_state(self, label: QLabel, status: str):
        label.setText(self._normalize_status(status))
        label.setProperty("badgeRole", "status")
        label.setProperty("statusKind", self._status_style_key(status))
        _refresh_widget_style(label)

    def _set_priority_badge_state(self, label: QLabel, text: str, priority: Optional[str]):
        label.setText(text)
        label.setProperty("badgeRole", "priority")
        label.setProperty("priority", priority or "none")
        _refresh_widget_style(label)

    def _build_document_history_row(self, event_text: str, created_at: str, created_by: str) -> QFrame:
        card = QFrame()
        card.setObjectName("analyticsRow")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(14, 12, 14, 12)
        lay.setSpacing(6)
        title = QLabel(event_text)
        title.setObjectName("analyticsTitle")
        title.setWordWrap(True)
        lay.addWidget(title)
        try:
            created_label = datetime.fromisoformat(created_at).strftime("%d.%m.%Y %H:%M")
        except Exception:
            created_label = fmt_date_ddmmyyyy(created_at) or created_at or "без даты"
        meta = QLabel(
            f"{created_label} • {created_by or 'без автора'}"
        )
        meta.setObjectName("analyticsHint")
        meta.setWordWrap(True)
        lay.addWidget(meta)
        return card

    def _render_document_history(self, doc_id: Optional[int]):
        if not hasattr(self, "doc_history_layout"):
            return
        self._clear_layout_widgets(self.doc_history_layout)
        if not doc_id or not self.db:
            self._add_dashboard_hint(self.doc_history_layout, "История появится после первого сохранения документа.")
            self.doc_history_layout.addStretch(1)
            return
        rows = list(self.db.list_document_history(int(doc_id), limit=10))
        if not rows:
            self._add_dashboard_hint(self.doc_history_layout, "Пока нет записей об изменениях.")
            self.doc_history_layout.addStretch(1)
            return
        for row in rows:
            self.doc_history_layout.addWidget(
                self._build_document_history_row(
                    str(row["event_text"] or ""),
                    str(row["created_at"] or ""),
                    str(row["created_by"] or ""),
                )
            )
        self.doc_history_layout.addStretch(1)

    def _update_document_sections_state(self, company_id: Optional[int]):
        if not hasattr(self, "doc_sections_hint"):
            return
        has_sections = self.doc_sections_list.count() > 0
        self.doc_sections_list.setEnabled(has_sections)
        self.doc_sections_hint.hide()

    def _sync_document_summary(self, row: Optional[sqlite3.Row] = None):
        if not hasattr(self, "doc_summary_meta"):
            return

        if row is None and self.current_document_id is not None and self.db:
            row = self.db.get_document_record(self.current_document_id)

        if row is not None:
            company_name = str(row["company_name"] or "Не указана")
            section_ids = self.db.get_doc_section_ids(int(row["id"])) if self.db else []
            section_names = self.db._section_names_for_history(self.db.conn.cursor(), section_ids) if self.db else ""
            status = self._normalize_status(row["status"])
            accept_text = fmt_date_ddmmyyyy(row["accept_date"] or "") or (row["accept_date"] or "не указана")
            updated_text = fmt_date_ddmmyyyy(row["updated_at"] or "") or (row["updated_at"] or "не указано")
            pdf_ready = self._document_has_required_pdf(row)
            office_ready = self._document_has_office_file(row)
            office_label = self._office_file_label(row["office_path"], fallback="Нет файла")
            needs_office = self._document_requires_office(row)
            priority = self._review_priority(row)
            if priority == "high":
                review_text = f"Просрочено: {self._review_due_text(row)}"
            elif priority == "mid":
                review_text = f"Скоро пересмотр: {self._review_due_text(row)}"
            elif priority == "low":
                review_text = f"Пересмотр в 30 дней: {self._review_due_text(row)}"
            else:
                review_text = f"Пересмотр: {self._review_due_text(row)}"
            severity, issues = self._document_attention_details(row)
            issues_text = "Требует внимания: " + " • ".join(issues) if issues else "Без замечаний."
            meta_text = (
                f"Компания: {company_name}\n"
                f"Разделы: {section_names or 'не выбраны'}\n"
                f"Дата утверждения: {accept_text}\n"
                f"Последнее изменение: {updated_text}"
            )
        else:
            company_name = self.doc_company_combo.currentText().strip() or "Не указана"
            section_names = ", ".join(
                self.doc_sections_list.item(i).text()
                for i in range(self.doc_sections_list.count())
                if self.doc_sections_list.item(i).checkState() == Qt.Checked
            )
            status = self.doc_status_combo.currentText().strip() or DOC_STATUS_DEFAULT
            review_text_raw = self.doc_review_date_edit.text().strip()
            review_due = parse_date_to_ymd(review_text_raw) if review_text_raw else None
            if review_text_raw and not review_due:
                review_text = "Проверьте дату пересмотра"
                priority = "high"
            elif review_due:
                fake_row = {"review_due": review_due}
                if review_due < today_ymd():
                    review_text = f"Просрочено: {fmt_date_ddmmyyyy(review_due)}"
                    priority = "high"
                elif (datetime.fromisoformat(review_due) - datetime.fromisoformat(today_ymd())).days <= 7:
                    review_text = f"Скоро пересмотр: {fmt_date_ddmmyyyy(review_due)}"
                    priority = "mid"
                elif (datetime.fromisoformat(review_due) - datetime.fromisoformat(today_ymd())).days <= 30:
                    review_text = f"Пересмотр в 30 дней: {fmt_date_ddmmyyyy(review_due)}"
                    priority = "low"
                else:
                    review_text = f"Пересмотр: {fmt_date_ddmmyyyy(review_due)}"
                    priority = None
            else:
                review_text = "Дата пересмотра не указана"
                priority = None
            pdf_pending = self._pending_document_file_path("pdf")
            office_pending = self._pending_document_file_path("office")
            pdf_ready = bool(pdf_pending and os.path.exists(pdf_pending))
            office_ready = bool(office_pending and os.path.exists(office_pending))
            office_label = self._office_file_label(office_pending, fallback="Нет файла")
            needs_office = self.doc_needs_office_checkbox.isChecked()
            issues: List[str] = []
            if not pdf_ready:
                issues.append("нет PDF")
            if needs_office and not office_ready:
                issues.append("нет Word / Excel")
            if not section_names:
                issues.append("не выбран раздел")
            if not review_due:
                issues.append("нет даты пересмотра")
            if not self.doc_title_edit.text().strip():
                issues.append("нет названия")
            issues_text = "Требует внимания: " + " • ".join(issues) if issues else "Без замечаний."
            meta_text = (
                f"Компания: {company_name}\n"
                f"Разделы: {section_names or 'не выбраны'}\n"
                f"Дата утверждения: {self.doc_accept_date_edit.text().strip() or 'не указана'}\n"
                f"Статус: {status}"
            )

        self._set_status_badge_state(self.doc_summary_status_badge, status)
        self._set_priority_badge_state(self.doc_summary_review_badge, review_text, priority)
        self._set_file_tag_state(self.doc_summary_pdf_badge, "PDF", "pdf", pdf_ready)
        self._set_file_tag_state(
            self.doc_summary_office_badge,
            office_label if (needs_office or office_ready) else "Файл не загружен",
            self._office_file_variant(row["office_path"] if row is not None else self._pending_document_file_path("office")),
            office_ready,
        )
        self.doc_summary_meta.setText(meta_text)
        self.doc_attention_label.setText(issues_text)

    def _review_due_text(self, row: sqlite3.Row, empty_text: str = "не назначена") -> str:
        return fmt_date_ddmmyyyy(row["review_due"] or "") or (row["review_due"] or empty_text)

    def _document_attention_details(self, row: sqlite3.Row) -> Tuple[int, List[str]]:
        issues: List[str] = []
        severity = 0

        pdf_stored = bool((row["pdf_path"] or "").strip())
        pdf_exists = self._document_has_required_pdf(row)
        if not pdf_exists:
            issues.append("PDF не загружен" if not pdf_stored else "PDF не найден")
            severity = max(severity, 5)

        office_stored = bool((row["office_path"] or "").strip())
        office_exists = self._path_exists(row["office_path"])
        if self._document_missing_required_office(row):
            issues.append("Нужен дополнительный файл Word / Excel")
            severity = max(severity, 3)
        elif office_stored and not office_exists:
            issues.append("Файл Word или Excel не найден")
            severity = max(severity, 3)

        if self._document_missing_status(row):
            issues.append("Статус не заполнен")
            severity = max(severity, 3)

        if self._document_missing_section(row):
            issues.append("Документ без раздела")
            severity = max(severity, 2)

        priority = self._review_priority(row)
        due_text = self._review_due_text(row, empty_text="не указана")
        if priority == "high":
            issues.append(f"Просрочен пересмотр: {due_text}")
            severity = max(severity, 5)
        elif priority == "mid":
            issues.append(f"Пересмотр в ближайшие 7 дней: {due_text}")
            severity = max(severity, 3)
        elif priority == "low":
            issues.append(f"Пересмотр в ближайшие 30 дней: {due_text}")
            severity = max(severity, 2)
        elif self._document_missing_review_date(row):
            issues.append("Не указана дата пересмотра")
            severity = max(severity, 1)

        return severity, issues

    def _dashboard_priority_bucket(self, row: sqlite3.Row) -> str:
        priority = self._review_priority(row)
        if priority == "high":
            return "Просрочено"
        if not self._document_has_required_pdf(row):
            return "Без PDF"
        if priority in {"mid", "low"}:
            return "Скоро пересмотр"
        if not parse_date_to_ymd(str(row["review_due"] or "").strip()):
            return "Без даты"
        return "В порядке"

    def _build_file_badge(self, title: str, has_file: bool, variant: str) -> QLabel:
        label = QLabel(title)
        label.setProperty("badgeRole", "file")
        label.setProperty("fileState", "present" if has_file else "missing")
        label.setProperty("fileVariant", variant)
        return label

    def _build_attention_badge(self, text: str, severity: int) -> QLabel:
        label = QLabel(text)
        label.setProperty("badgeRole", "priority")
        if severity >= 4:
            label.setProperty("priority", "high")
        elif severity >= 2:
            label.setProperty("priority", "mid")
        else:
            label.setProperty("priority", "low")
        return label

    def _build_attention_card(self, row: sqlite3.Row, issues: List[str]) -> ClickableCard:
        card = ClickableCard(int(row["id"]))
        card.setObjectName("reviewCard")
        card.clicked.connect(self.select_document)

        lay = QVBoxLayout(card)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(10)

        severity, _ = self._document_attention_details(row)
        top = QHBoxLayout()
        top.setSpacing(10)
        top.addWidget(self._card_title_label(str(row["doc_title"])), 1)
        top.addWidget(self._build_attention_badge(issues[0], severity), 0, Qt.AlignTop)
        lay.addLayout(top)

        lay.addWidget(self._meta_label(f"Компания: {row['company_name']}"))
        if len(issues) > 1:
            lay.addWidget(self._meta_label(" • ".join(issues[1:])))

        bottom = QHBoxLayout()
        bottom.setSpacing(8)
        bottom.addWidget(self._meta_label(f"Пересмотр: {self._review_due_text(row)}"), 1)
        bottom.addWidget(self._build_file_badge("PDF", self._document_has_required_pdf(row), "pdf"))
        office_exists = self._path_exists(row["office_path"])
        office_label = self._office_file_label(row["office_path"], fallback="Нет файла")
        bottom.addWidget(
            self._build_file_badge(
                office_label,
                office_exists,
                self._office_file_variant(row["office_path"]),
            )
        )
        lay.addLayout(bottom)

        actions = QHBoxLayout()
        actions.setSpacing(8)
        actions.addStretch(1)
        open_btn = self._create_button("Перейти к карточке", kind="ghost")
        open_btn.clicked.connect(lambda checked=False, doc_id=int(row["id"]): self._open_document_from_dashboard(doc_id))
        actions.addWidget(open_btn)

        if not self._document_has_required_pdf(row):
            upload_btn = self._create_button("Загрузить PDF", kind="soft")
            upload_btn.clicked.connect(lambda checked=False, doc_id=int(row["id"]): self._quick_upload_document_file(doc_id, "pdf"))
            actions.addWidget(upload_btn)
        elif self._document_missing_required_office(row):
            upload_btn = self._create_button("Загрузить файл", kind="soft")
            upload_btn.clicked.connect(lambda checked=False, doc_id=int(row["id"]): self._quick_upload_document_file(doc_id, "office"))
            actions.addWidget(upload_btn)
        elif self._review_priority(row) is not None:
            review_btn = self._create_button("Обновить дату", kind="soft")
            review_btn.clicked.connect(lambda checked=False, doc_id=int(row["id"]): self._start_review_update(doc_id))
            actions.addWidget(review_btn)
        lay.addLayout(actions)
        return card

    def _build_recent_change_row(self, row: sqlite3.Row) -> QFrame:
        card = QFrame()
        card.setObjectName("analyticsRow")

        lay = QVBoxLayout(card)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(8)

        title = QLabel(str(row["doc_title"]))
        title.setObjectName("analyticsTitle")
        title.setWordWrap(True)
        lay.addWidget(title)

        updated_text = fmt_date_ddmmyyyy(row["updated_at"] or "") or (row["updated_at"] or "без даты")
        updated_by = str(row["updated_by"] or "").strip() or "без автора"
        meta = QLabel(f"{row['company_name']} • {updated_text} • {updated_by}")
        meta.setObjectName("analyticsHint")
        meta.setWordWrap(True)
        lay.addWidget(meta)

        actions = QHBoxLayout()
        actions.setSpacing(8)
        actions.addStretch(1)
        action_btn = self._create_button("Перейти к карточке", kind="secondary")
        action_btn.clicked.connect(lambda checked=False, doc_id=int(row["id"]): self._open_document_from_dashboard(doc_id))
        actions.addWidget(action_btn)
        lay.addLayout(actions)
        return card

    def _priority_badge(self, priority: str, due_text: str) -> QLabel:
        text_map = {
            "high": f"Просрочен с {due_text}",
            "mid": f"Срок близко: {due_text}",
            "low": f"Пересмотр скоро: {due_text}",
        }
        label = QLabel(text_map.get(priority, due_text))
        label.setProperty("badgeRole", "priority")
        label.setProperty("priority", priority)
        return label

    def _build_document_card(self, row: sqlite3.Row, card_kind: str) -> ClickableCard:
        card = ClickableCard(int(row["id"]))
        card.setObjectName("reviewCard" if card_kind == "review" else "documentCard")
        card.clicked.connect(self.select_document)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(10)
        top.addWidget(self._card_title_label(str(row["doc_title"])), 1)
        if card_kind == "review":
            priority = self._review_priority(row) or "low"
            due_text = fmt_date_ddmmyyyy(row["review_due"] or "") or (row["review_due"] or "")
            top.addWidget(self._priority_badge(priority, due_text), 0, Qt.AlignTop)
        lay.addLayout(top)

        details = QGridLayout()
        details.setContentsMargins(0, 0, 0, 0)
        details.setHorizontalSpacing(12)
        details.setVerticalSpacing(6)
        details.setColumnMinimumWidth(0, 96)
        details.setColumnStretch(1, 1)
        detail_row = 0

        details.addWidget(self._meta_key_label("Компания"), detail_row, 0)
        details.addWidget(self._meta_label(str(row["company_name"] or "Не указана")), detail_row, 1)
        detail_row += 1

        if card_kind != "review":
            details.addWidget(self._meta_key_label("Статус"), detail_row, 0)
            details.addWidget(self._build_status_badge(self._normalize_status(row["status"])), detail_row, 1, 1, 1, Qt.AlignLeft | Qt.AlignTop)
            detail_row += 1

        section_names = str(row["section_names"] or "").strip()
        if section_names:
            details.addWidget(self._meta_key_label("Раздел"), detail_row, 0)
            details.addWidget(self._meta_label(section_names), detail_row, 1)
            detail_row += 1

        due_text = self._review_due_text(row)
        details.addWidget(self._meta_key_label("Пересмотр"), detail_row, 0)
        details.addWidget(self._meta_label(due_text), detail_row, 1)
        detail_row += 1

        has_office = self._path_exists(row["office_path"])
        file_row = QWidget()
        file_row_l = QHBoxLayout(file_row)
        file_row_l.setContentsMargins(0, 0, 0, 0)
        file_row_l.setSpacing(8)
        file_row_l.addWidget(self._build_file_badge("PDF", self._document_has_required_pdf(row), "pdf"), 0)
        if has_office:
            file_row_l.addWidget(
                self._build_file_badge(
                    self._office_file_label(row["office_path"]),
                    True,
                    self._office_file_variant(row["office_path"]),
                ),
                0,
            )
        file_row_l.addStretch(1)
        details.addWidget(self._meta_key_label("Файлы"), detail_row, 0)
        details.addWidget(file_row, detail_row, 1)
        lay.addLayout(details)

        if card_kind == "review":
            actions = QHBoxLayout()
            actions.setSpacing(8)
            actions.addStretch(1)
            update_btn = self._create_button("Обновить дату", kind="soft")
            update_btn.clicked.connect(lambda checked=False, doc_id=int(row["id"]): self._start_review_update(doc_id))
            actions.addWidget(update_btn)
            lay.addLayout(actions)
        return card

    def _sync_document_card_selection(self):
        all_cards: Dict[int, ClickableCard] = {}
        all_cards.update(self._document_cards)
        all_cards.update(self._review_cards)
        for doc_id, card in all_cards.items():
            card.set_selected(doc_id == self.current_document_id)

    def refresh_documents_page(self):
        self.documents_page.setUpdatesEnabled(False)
        try:
            rows = self._document_rows_for_page("documents")
            without_files = sum(1 for row in rows if not self._document_has_required_pdf(row))
            overdue = sum(1 for row in rows if self._review_priority(row) == "high")
            without_section = sum(1 for row in rows if self._document_missing_section(row))
            self.documents_count_label.setText(
                f"{len(rows)} документов  ·  без PDF: {without_files}  ·  просрочено: {overdue}  ·  без раздела: {without_section}"
            )
            self._document_cards = {}
            if not rows:
                self._render_empty_list_state(
                    self.documents_layout,
                    "Документы не найдены",
                    "Сбросьте фильтры или сразу добавьте новый документ.",
                    action_text="Добавить документ",
                    action_callback=self.start_new_document,
                )
            else:
                self._clear_layout_widgets(self.documents_layout)
                for row in rows:
                    card = self._build_document_card(row, "documents")
                    self._document_cards[int(row["id"])] = card
                    self.documents_layout.addWidget(card)
                self.documents_layout.addStretch(1)

            visible_ids = {int(row["id"]) for row in rows}
            if self.current_document_id not in visible_ids and self._page_key == "documents":
                self.clear_right_panel_state("documents", preferred_company_id=self.current_filter_company_id())
            else:
                self._sync_document_card_selection()
        finally:
            self.documents_page.setUpdatesEnabled(True)
            self.documents_page.update()

    def refresh_archive_page(self):
        rows = self._document_rows_for_page("archive")
        self.archive_count_label.setText(f"{len(rows)} документов в архиве")
        self._archive_cards = {}
        if not rows:
            self._render_empty_list_state(
                self.archive_layout,
                "Архив пока пуст",
                "Когда документы будут переведены в архив, они появятся здесь.",
            )
        else:
            self._clear_layout_widgets(self.archive_layout)
            for row in rows:
                card = self._build_document_card(row, "archive")
                self._archive_cards[int(row["id"])] = card
                self.archive_layout.addWidget(card)
            self.archive_layout.addStretch(1)

        visible_ids = {int(row["id"]) for row in rows}
        if self.current_document_id not in visible_ids and self._page_key == "archive":
            self.current_document_id = None
            self._show_right_placeholder("Архивный документ", "Выберите документ из архива, чтобы посмотреть детали и вернуть его в работу.")
        else:
            self._sync_document_card_selection()

    def refresh_review_page(self):
        rows = self._document_rows_for_page("review")
        overdue_rows = [row for row in rows if self._review_priority(row) == "high"]
        upcoming_rows = [row for row in rows if self._review_priority(row) in {"mid", "low"}]
        self.review_overdue_summary.setText(f"Просрочено: {len(overdue_rows)}")
        self.review_upcoming_summary.setText(f"Скоро пересмотр: {len(upcoming_rows)}")

        self._review_cards = {}
        if overdue_rows:
            self._clear_layout_widgets(self.review_overdue_layout)
            for row in overdue_rows:
                card = self._build_document_card(row, "review")
                self._review_cards[int(row["id"])] = card
                self.review_overdue_layout.addWidget(card)
            self.review_overdue_layout.addStretch(1)
        else:
            self._render_empty_list_state(
                self.review_overdue_layout,
                "Просроченных документов нет",
                "Здесь будут только документы, которые уже пора обновить.",
            )

        if upcoming_rows:
            self._clear_layout_widgets(self.review_upcoming_layout)
            for row in upcoming_rows:
                card = self._build_document_card(row, "review")
                self._review_cards[int(row["id"])] = card
                self.review_upcoming_layout.addWidget(card)
            self.review_upcoming_layout.addStretch(1)
        else:
            self._render_empty_list_state(
                self.review_upcoming_layout,
                "Скорых пересмотров нет",
                "Когда срок начнёт приближаться, документы появятся здесь.",
            )

        visible_ids = {int(row["id"]) for row in rows}
        if self.current_document_id not in visible_ids and self._page_key == "review":
            self.current_document_id = None
            self._show_right_placeholder("Пересмотр документа", "Выберите документ слева и обновите дату пересмотра.")
        else:
            self._sync_document_card_selection()

    def _set_document_form_dirty(self):
        if not self._doc_form_loading:
            self._doc_form_dirty = True
        self._update_document_guidance()
        self._sync_document_summary()

    def _maybe_discard_document_changes(self) -> bool:
        if not self._doc_form_dirty:
            return True
        if not self.confirm("Изменения в карточке документа не сохранены. Продолжить без сохранения?"):
            return False
        self._doc_form_dirty = False
        return True

    def _on_document_company_changed(self):
        if self._doc_form_loading:
            return
        company_id = self.doc_company_combo.currentData()
        self._populate_document_sections(int(company_id) if company_id not in (None, "") else None, checked_ids=[])
        self._set_document_form_dirty()

    def _populate_document_sections(self, company_id: Optional[int], checked_ids: Optional[List[int]] = None):
        checked_ids = checked_ids or []
        self.doc_sections_list.blockSignals(True)
        self.doc_sections_list.clear()
        if company_id and self.db:
            for row in self.db.list_sections(company_id):
                item = QListWidgetItem(str(row["name"]))
                item.setData(Qt.UserRole, int(row["id"]))
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Checked if int(row["id"]) in checked_ids else Qt.Unchecked)
                self.doc_sections_list.addItem(item)
        self.doc_sections_list.blockSignals(False)
        self._update_document_sections_state(company_id)

    def _checked_document_section_ids(self) -> List[int]:
        ids: List[int] = []
        for i in range(self.doc_sections_list.count()):
            item = self.doc_sections_list.item(i)
            if item.checkState() == Qt.Checked:
                ids.append(int(item.data(Qt.UserRole)))
        return ids

    def _reset_pending_document_files(self):
        self._pending_document_files = {"pdf": None, "office": None}

    def _pending_document_file_path(self, kind: str) -> Optional[str]:
        return self._pending_document_files.get(kind)

    def _document_preview_file_path(self, kind: str) -> Optional[str]:
        if self.current_document_id is None:
            path = self._pending_document_file_path(kind)
            return path if path and os.path.exists(path) else None
        if not self.db:
            return None
        row = self.db.get_document_record(self.current_document_id)
        if not row:
            return None
        field = "pdf_path" if kind == "pdf" else "office_path"
        abs_path = self._abs_storage_path(row[field])
        return abs_path if abs_path and os.path.exists(abs_path) else None

    def _document_has_any_preview_file(self) -> bool:
        return bool(self._document_preview_file_path("pdf") or self._document_preview_file_path("office"))

    def _sync_document_editor_actions(self, status: Optional[str] = None):
        self.doc_save_btn.setText("Сохранить документ")
        self.doc_save_btn.setEnabled(True)
        if self.current_document_id is None:
            self.doc_delete_btn.setEnabled(False)
            return
        self.doc_delete_btn.setEnabled(True)

    def _document_saved_message(self, with_files: bool = False) -> str:
        return "Документ сохранён. Файлы прикреплены." if with_files else "Документ сохранён."

    def _update_document_guidance(self):
        self.right_panel_subtitle.setText("")

    def select_document(self, doc_id: int, force: bool = False):
        if not force and self.current_document_id == doc_id and self.right_stack.currentWidget() == self.right_document_page:
            self._sync_document_card_selection()
            return
        if not self._maybe_discard_document_changes():
            return
        self.current_document_id = int(doc_id)
        self._load_document_into_editor(self.current_document_id)
        self._sync_document_card_selection()

    def _clear_document_editor(self, preferred_company_id: Optional[int] = None, show_editor: bool = True):
        self._doc_form_loading = True
        self.current_document_id = None
        self._reset_pending_document_files()
        self.doc_title_edit.setText("")
        self.doc_comment_edit.setPlainText("")
        target_company_id = preferred_company_id
        if target_company_id is None and self.doc_company_combo.count():
            current = self.current_filter_company_id()
            if current is not None:
                target_company_id = current
            else:
                target_company_id = self.doc_company_combo.itemData(0)
        if target_company_id is not None:
            for i in range(self.doc_company_combo.count()):
                if self.doc_company_combo.itemData(i) == target_company_id:
                    self.doc_company_combo.setCurrentIndex(i)
                    break
        self.doc_status_combo.setCurrentText(DOC_STATUS_DEFAULT)
        self.doc_accept_date_edit.setText("")
        self.doc_review_date_edit.setText("")
        self.doc_needs_office_checkbox.setChecked(False)
        selected_company_id = self.doc_company_combo.currentData()
        self._populate_document_sections(int(selected_company_id) if selected_company_id not in (None, "") else None, [])
        self._doc_form_loading = False
        self._doc_form_dirty = False
        self._update_document_file_controls()
        self._render_document_history(None)
        self._sync_document_editor_actions()
        self._sync_document_summary()
        if show_editor:
            self._set_right_actions_mode("document")
            self.right_stack.setCurrentWidget(self.right_document_page)
            self.right_panel_title.setText("Новый документ")
            self.right_panel_subtitle.setText("")
            self._update_document_guidance()
        else:
            self._set_right_actions_mode("none")

    def start_new_document(self):
        if not self.db:
            return
        if not self._maybe_discard_document_changes():
            return
        if self._page_key != "documents":
            self.set_active_page("documents")
            if self._page_key != "documents":
                return

        company_id = self._default_document_company_id()
        if company_id is None:
            self.warn("Сначала добавьте компанию.")
            return

        self._search_timer.stop()
        self.search_input.blockSignals(True)
        self.search_input.setText("")
        self.search_input.blockSignals(False)
        self.status_filter_combo.blockSignals(True)
        if self.status_filter_combo.count():
            self.status_filter_combo.setCurrentIndex(0)
        self.status_filter_combo.blockSignals(False)
        self.problem_filter_combo.blockSignals(True)
        if self.problem_filter_combo.count():
            self.problem_filter_combo.setCurrentIndex(0)
        self.problem_filter_combo.blockSignals(False)
        self.company_filter_combo.blockSignals(True)
        if self.company_filter_combo.count():
            self._set_combo_value(self.company_filter_combo, int(company_id))
        self.company_filter_combo.blockSignals(False)
        self._populate_section_filter()
        self.section_filter_combo.blockSignals(True)
        if self.section_filter_combo.count():
            self.section_filter_combo.setCurrentIndex(0)
        self.section_filter_combo.blockSignals(False)
        self._last_company_filter_value = self.current_filter_company_id()
        self._last_section_filter_value = None
        self._last_status_filter_value = None
        self._last_problem_filter_value = None
        self.refresh_documents_page()

        self._clear_document_editor(preferred_company_id=int(company_id), show_editor=True)
        self.right_panel_title.setText("Новый документ")
        self.right_panel_subtitle.setText("")
        self.doc_title_edit.setFocus()
        self.doc_title_edit.selectAll()

    def _load_document_into_editor(self, doc_id: int):
        if not self.db:
            return
        row = self.db.get_document_record(doc_id)
        if not row:
            self.warn("Документ не найден.")
            return
        checked_sections = self.db.get_doc_section_ids(doc_id)
        self._doc_form_loading = True
        self._reset_pending_document_files()
        self.doc_title_edit.setText(str(row["doc_title"] or ""))
        for i in range(self.doc_company_combo.count()):
            if self.doc_company_combo.itemData(i) == int(row["company_id"]):
                self.doc_company_combo.setCurrentIndex(i)
                break
        self.doc_status_combo.setCurrentText(self._normalize_status(row["status"]))
        self.doc_accept_date_edit.setText(fmt_date_ddmmyyyy(row["accept_date"] or "") or "")
        self.doc_review_date_edit.setText(fmt_date_ddmmyyyy(row["review_due"] or "") or "")
        self.doc_comment_edit.setPlainText(str(row["comment"] or ""))
        self.doc_needs_office_checkbox.setChecked(bool(row["needs_office"]))
        self._populate_document_sections(int(row["company_id"]), checked_sections)
        self._doc_form_loading = False
        self._doc_form_dirty = False
        self._update_document_file_controls()
        self._render_document_history(doc_id)
        self._set_right_actions_mode("document")
        self.right_stack.setCurrentWidget(self.right_document_page)
        self.right_panel_title.setText(str(row["doc_title"]))
        self.right_panel_subtitle.setText("")
        self._sync_document_editor_actions(str(row["status"] or ""))
        self._sync_document_summary(row)
        self._update_document_guidance()

    def _update_document_file_controls(self):
        pdf_pending = self._pending_document_file_path("pdf")
        office_pending = self._pending_document_file_path("office")

        if not self.db or self.current_document_id is None:
            pdf_ready = bool(pdf_pending and os.path.exists(pdf_pending))
            office_ready = bool(office_pending and os.path.exists(office_pending))
            office_variant = self._office_file_variant(office_pending)
            office_label = self._office_file_label(office_pending, fallback="Файл не загружен")
            self.doc_pdf_caption.setText(
                f"Выбран файл: {os.path.basename(pdf_pending)}" if pdf_ready else "Файл не загружен"
            )
            self.doc_office_caption.setText(
                f"Выбран файл: {os.path.basename(office_pending)}" if office_ready else "Файл не загружен"
            )
            self._set_file_tag_state(self.doc_pdf_tag, "Основной PDF", "pdf", pdf_ready)
            self._set_file_tag_state(
                self.doc_office_tag,
                office_label if office_ready else "Нет файла",
                office_variant,
                office_ready,
            )
            self.doc_pdf_upload_btn.setText("Заменить" if pdf_ready else "Загрузить")
            self.doc_office_upload_btn.setText("Заменить" if office_ready else "Загрузить")
            self.doc_pdf_upload_btn.setEnabled(True)
            self.doc_office_upload_btn.setEnabled(True)
            self.doc_pdf_open_btn.setEnabled(pdf_ready)
            self.doc_pdf_delete_btn.setEnabled(pdf_ready)
            self.doc_office_open_btn.setEnabled(office_ready)
            self.doc_office_delete_btn.setEnabled(office_ready)
            self._sync_document_summary()
            self._update_document_guidance()
            return

        row = self.db.get_document_record(self.current_document_id)
        if not row:
            return
        pdf_abs = self._abs_storage_path(row["pdf_path"])
        office_abs = self._abs_storage_path(row["office_path"])
        pdf_exists = bool(pdf_abs and os.path.exists(pdf_abs))
        office_exists = bool(office_abs and os.path.exists(office_abs))
        pdf_name = os.path.basename(pdf_abs) if pdf_abs else ""
        office_name = os.path.basename(office_abs) if office_abs else ""
        office_variant = self._office_file_variant(row["office_path"])
        office_label = self._office_file_label(row["office_path"], fallback="Файл не загружен")
        self.doc_pdf_caption.setText(
            f"Файл: {pdf_name}" if pdf_exists else (f"Файл не найден: {pdf_name}" if pdf_name else "Файл не загружен")
        )
        self.doc_office_caption.setText(
            f"Файл: {office_name}" if office_exists else (f"Файл не найден: {office_name}" if office_name else "Файл не загружен")
        )
        self._set_file_tag_state(self.doc_pdf_tag, "Основной PDF", "pdf", pdf_exists)
        self._set_file_tag_state(
            self.doc_office_tag,
            office_label if row["office_path"] else "Нет файла",
            office_variant,
            office_exists,
        )
        self.doc_pdf_upload_btn.setText("Заменить" if bool(row["pdf_path"]) else "Загрузить")
        self.doc_office_upload_btn.setText("Заменить" if bool(row["office_path"]) else "Загрузить")
        self.doc_pdf_upload_btn.setEnabled(True)
        self.doc_office_upload_btn.setEnabled(True)
        self.doc_pdf_open_btn.setEnabled(pdf_exists)
        self.doc_pdf_delete_btn.setEnabled(bool(row["pdf_path"]))
        self.doc_office_open_btn.setEnabled(office_exists)
        self.doc_office_delete_btn.setEnabled(bool(row["office_path"]))
        self._sync_document_summary(row)
        self._update_document_guidance()

    def _unique_destination_path(self, path: str) -> str:
        if not os.path.exists(path):
            return path
        base, ext = os.path.splitext(path)
        index = 2
        while True:
            candidate = f"{base}_{index}{ext}"
            if not os.path.exists(candidate):
                return candidate
            index += 1

    def _move_document_files_if_needed(self, row: sqlite3.Row, new_company_id: int) -> Tuple[Optional[str], Optional[str]]:
        old_company_id = int(row["company_id"])
        pdf_rel = row["pdf_path"]
        office_rel = row["office_path"]
        if old_company_id == new_company_id:
            return pdf_rel, office_rel

        moved_values = {"pdf_path": pdf_rel, "office_path": office_rel}
        target_dir = self._doc_dir(new_company_id, int(row["id"]))
        for field in ("pdf_path", "office_path"):
            stored = row[field]
            abs_path = self._abs_storage_path(stored)
            if not stored or not abs_path or not os.path.exists(abs_path):
                continue
            dst = os.path.join(target_dir, os.path.basename(abs_path))
            if os.path.abspath(abs_path) != os.path.abspath(dst):
                dst = self._unique_destination_path(dst)
                ensure_dir(os.path.dirname(dst))
                shutil.move(abs_path, dst)
            moved_values[field] = self._rel_storage_path(dst)
        return moved_values["pdf_path"], moved_values["office_path"]

    def _prepare_document_file_target(self, company_id: int, doc_id: int, kind: str, file_path: str) -> Tuple[str, str, str]:
        ext = os.path.splitext(file_path)[1].lower()
        base_name = self._safe_filename(os.path.splitext(os.path.basename(file_path))[0])
        doc_dir = self._doc_dir(company_id, doc_id)
        if kind == "pdf":
            if ext != ".pdf":
                raise ValueError("Для PDF нужен файл с расширением .pdf.")
            field = "pdf_path"
            dst = os.path.join(doc_dir, f"{base_name}.pdf")
        else:
            if ext not in (".doc", ".docx", ".xls", ".xlsx"):
                raise ValueError("Для Office нужен файл doc, docx, xls или xlsx.")
            field = "office_path"
            dst = os.path.join(doc_dir, f"{base_name}{ext}")
        return field, dst, self._rel_storage_path(dst)

    def _pick_document_file(self, kind: str) -> Optional[str]:
        if kind == "pdf":
            file_path, _ = QFileDialog.getOpenFileName(self, "Загрузить PDF", "", "PDF (*.pdf)")
        else:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Загрузить дополнительный файл",
                "",
                "Файлы Word и Excel (*.doc *.docx *.xls *.xlsx)",
            )
        return file_path or None

    def _stage_pending_document_file(self, kind: str, file_path: str):
        if not file_path or not os.path.isfile(file_path):
            self.warn("Файл не найден.")
            return
        try:
            self._prepare_document_file_target(0, 0, kind, file_path)
        except Exception as e:
            self.warn(str(e))
            return
        self._pending_document_files[kind] = file_path
        self._update_document_file_controls()
        self._set_document_form_dirty()

    def _persist_pending_document_files(self, doc_id: int, company_id: int) -> bool:
        pending = {kind: path for kind, path in self._pending_document_files.items() if path and os.path.exists(path)}
        if not pending:
            return False

        current = self.db.get_document_record(doc_id) if self.db else None
        previous_abs_paths = {
            "pdf_path": self._abs_storage_path(current["pdf_path"]) if current and current["pdf_path"] else None,
            "office_path": self._abs_storage_path(current["office_path"]) if current and current["office_path"] else None,
        }
        self._reset_pending_document_files()

        def work():
            results = []
            with self._doc_lock(company_id, doc_id):
                for kind, source_path in pending.items():
                    field, dst, rel_dst = self._prepare_document_file_target(company_id, doc_id, kind, source_path)
                    old_abs = previous_abs_paths.get(field)
                    if old_abs and os.path.exists(old_abs) and os.path.abspath(old_abs) != os.path.abspath(dst):
                        try:
                            os.remove(old_abs)
                        except Exception:
                            pass
                    atomic_copy(source_path, dst)
                    results.append((field, rel_dst))
            return results

        def on_success(results: Any):
            for field, rel_dst in results:
                self.db.update_file_path(doc_id, field, rel_dst)
            self.current_document_id = doc_id
            self.load_companies(preferred_company_id=company_id)
            self.select_document(doc_id, force=True)
            self.info(self._document_saved_message(with_files=True))

        self._run_background_task(
            label="Сохранение файлов",
            work_fn=work,
            on_success=on_success,
            on_error_prefix=f"{self._document_saved_message()} Но файлы прикрепить не получилось:",
        )
        return True

    def save_document_form(self):
        if not self.db:
            return
        title = self.doc_title_edit.text().strip()
        company_id = self.doc_company_combo.currentData()
        status = self.doc_status_combo.currentText().strip()
        accept_text = self.doc_accept_date_edit.text().strip()
        review_text = self.doc_review_date_edit.text().strip()
        comment = self.doc_comment_edit.toPlainText().strip()
        needs_office = self.doc_needs_office_checkbox.isChecked()
        accept_date = parse_date_to_ymd(accept_text) if accept_text else None
        review_due = parse_date_to_ymd(review_text) if review_text else None
        section_ids = self._checked_document_section_ids()

        if not title:
            self.warn("Укажите название документа.")
            return
        if company_id in (None, ""):
            self.warn("Выберите компанию.")
            return
        if accept_text and not accept_date:
            self.warn("Дата принятия заполнена неверно. Используйте формат дд.мм.гггг.")
            return
        if review_text and not review_due:
            self.warn("Дата пересмотра заполнена неверно. Используйте формат дд.мм.гггг.")
            return
        if not self._document_preview_file_path("pdf") and not self.confirm("У документа пока нет основного PDF. Сохранить карточку без него?"):
            return
        company_id = int(company_id)
        try:
            target_doc_id: Optional[int] = self.current_document_id
            if self.current_document_id is None:
                new_doc_id = self.db.add_document(company_id, title)
                current = self.db.get_document_record(new_doc_id)
                self.db.update_document_record(
                    new_doc_id,
                    company_id,
                    title,
                    status,
                    accept_date,
                    review_due,
                    current["pdf_path"] if current else None,
                    current["office_path"] if current else None,
                    comment,
                    needs_office,
                )
                self.db.set_doc_sections(new_doc_id, section_ids)
                self.current_document_id = new_doc_id
                target_doc_id = new_doc_id
            else:
                current = self.db.get_document_record(self.current_document_id)
                if not current:
                    self.warn("Документ не найден.")
                    return
                pdf_rel, office_rel = self._move_document_files_if_needed(current, company_id)
                self.db.update_document_record(
                    self.current_document_id,
                    company_id,
                    title,
                    status,
                    accept_date,
                    review_due,
                    pdf_rel,
                    office_rel,
                    comment,
                    needs_office,
                )
                self.db.set_doc_sections(self.current_document_id, section_ids)
                target_doc_id = self.current_document_id
            self._doc_form_dirty = False
            self.refresh_notifications()
            if target_doc_id is not None and self._persist_pending_document_files(target_doc_id, company_id):
                self.current_document_id = target_doc_id
                return
            self.load_companies(preferred_company_id=company_id)
            if target_doc_id is not None:
                self.current_document_id = target_doc_id
                self.select_document(target_doc_id, force=True)
            self.info(self._document_saved_message())
        except Exception as e:
            self.warn(str(e))

    def archive_selected_document(self):
        if self.current_document_id is None or not self.db:
            return
        row = self.db.get_document_record(self.current_document_id)
        if not row:
            return
        current_status = self._normalize_status(row["status"])
        target_status = DOC_STATUS_DEFAULT if current_status == "Архив" else "Архив"
        self.doc_status_combo.setCurrentText(target_status)
        self.save_document_form()

    def delete_selected_document(self):
        if self.current_document_id is not None:
            self.delete_document_row(self.current_document_id)

    def open_selected_document_file(self, kind: str):
        if self.current_document_id is None:
            path = self._pending_document_file_path(kind)
            if path and os.path.exists(path):
                QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(path)))
                return
            self.warn("Сначала выберите файл или сохраните документ.")
            return
        self.open_file(self.current_document_id, kind)

    def upload_selected_document_file(self, kind: str):
        if self.current_document_id is None:
            file_path = self._pick_document_file(kind)
            if file_path:
                self._stage_pending_document_file(kind, file_path)
            return
        self.upload_file_dialog(self.current_document_id, kind)

    def delete_selected_document_file(self, kind: str):
        if self.current_document_id is None:
            if self._pending_document_file_path(kind):
                self._pending_document_files[kind] = None
                self._update_document_file_controls()
                self._set_document_form_dirty()
            return
        self.delete_file(self.current_document_id, kind)

    def upload_file_dialog(self, doc_id: int, kind: str):
        file_path = self._pick_document_file(kind)
        if file_path:
            self.upload_file_from_path(doc_id, kind, file_path)

    def upload_file_from_path(self, doc_id: int, kind: str, file_path: str):
        if not self.db:
            return
        row = self.db.get_document_record(doc_id)
        if not row:
            self.warn("Документ не найден.")
            return

        company_id = int(row["company_id"])
        try:
            field, dst, rel_dst = self._prepare_document_file_target(company_id, doc_id, kind, file_path)
        except Exception as e:
            self.warn(str(e))
            return

        old_rel = row[field]
        old_abs = self._abs_storage_path(old_rel) if old_rel else None

        def work():
            with self._doc_lock(company_id, doc_id):
                if old_abs and os.path.exists(old_abs) and os.path.abspath(old_abs) != os.path.abspath(dst):
                    try:
                        os.remove(old_abs)
                    except Exception:
                        pass
                atomic_copy(file_path, dst)
            return {"doc_id": doc_id, "field": field, "rel_path": rel_dst}

        def on_success(result: Any):
            self.db.update_file_path(result["doc_id"], result["field"], result["rel_path"])
            self.load_companies(preferred_company_id=company_id)
            if self.current_document_id == doc_id:
                self._update_document_file_controls()
                self._render_document_history(doc_id)
            self.info("Файл загружен.")

        self._run_background_task(
            label="Загрузка файла",
            work_fn=work,
            on_success=on_success,
            on_error_prefix="Не получилось загрузить файл:",
        )

    def delete_file(self, doc_id: int, kind: str):
        if not self.db:
            return
        row = self.db.get_document_record(doc_id)
        if not row:
            self.warn("Документ не найден.")
            return

        company_id = int(row["company_id"])
        field = "pdf_path" if kind == "pdf" else "office_path"
        stored = row[field]
        abs_path = self._abs_storage_path(stored)
        if not stored:
            return
        if not self.confirm("Удалить файл из документа и с диска?"):
            return

        try:
            with self._doc_lock(company_id, doc_id):
                if abs_path and os.path.exists(abs_path):
                    try:
                        os.remove(abs_path)
                    except Exception:
                        pass
                self.db.update_file_path(doc_id, field, None)
            self.load_companies(preferred_company_id=company_id)
            if self.current_document_id == doc_id:
                self._update_document_file_controls()
                self._render_document_history(doc_id)
            self.info("Файл удалён.")
        except TimeoutError:
            self.warn("Документ сейчас редактирует другой пользователь. Попробуйте позже.")
        except Exception as e:
            self.warn(str(e))

    def delete_document_row(self, doc_id: int):
        if not self.db:
            return
        row = self.db.get_document_record(doc_id)
        if not row:
            self.warn("Документ не найден.")
            return
        if not self.confirm(f"Удалить документ «{row['doc_title']}»?"):
            return

        company_id = int(row["company_id"])
        try:
            with self._doc_lock(company_id, doc_id):
                for field in ("pdf_path", "office_path"):
                    abs_path = self._abs_storage_path(row[field])
                    if abs_path and os.path.exists(abs_path):
                        try:
                            os.remove(abs_path)
                        except Exception:
                            pass
                self.db.delete_document(doc_id)
            if self.current_document_id == doc_id:
                self.clear_right_panel_state("documents", preferred_company_id=company_id)
            self.load_companies(preferred_company_id=company_id)
        except TimeoutError:
            self.warn("Документ сейчас редактирует другой пользователь. Попробуйте позже.")
        except Exception as e:
            self.warn(str(e))

    def _set_sopd_form_dirty(self):
        if not self._sopd_form_loading:
            self._sopd_form_dirty = True

    def _maybe_discard_sopd_changes(self) -> bool:
        if not self._sopd_form_dirty:
            return True
        if not self.confirm("Изменения в карточке СОПД не сохранены. Продолжить без сохранения?"):
            return False
        self._sopd_form_dirty = False
        return True

    def _company_name_by_id(self, company_id: int) -> str:
        for i in range(self.company_filter_combo.count()):
            if self.company_filter_combo.itemData(i) == company_id:
                return self.company_filter_combo.itemText(i).split("  ·  ")[0]
        if self.db:
            for row in self.db.list_companies():
                if int(row["id"]) == company_id:
                    return str(row["name"])
        return "Компания"

    def _build_sopd_card(self, row: sqlite3.Row) -> ClickableCard:
        card = ClickableCard(int(row["id"]))
        card.setObjectName("sopdCard")
        card.clicked.connect(self.select_sopd)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(10)
        top.addWidget(self._card_title_label(str(row["consent_type"] or "Карточка СОПД")), 1)
        transfer_badge = QLabel(f"Передача: {row['third_party_transfer'] or 'Не указано'}")
        transfer_badge.setProperty("badgeRole", "status")
        transfer_badge.setProperty("statusKind", "review" if (row["third_party_transfer"] or "").strip() == "Да" else "draft")
        top.addWidget(transfer_badge, 0, Qt.AlignTop)
        lay.addLayout(top)

        details = QGridLayout()
        details.setContentsMargins(0, 0, 0, 0)
        details.setHorizontalSpacing(12)
        details.setVerticalSpacing(6)
        details.setColumnMinimumWidth(0, 96)
        details.setColumnStretch(1, 1)

        details.addWidget(self._meta_key_label("Компания"), 0, 0)
        details.addWidget(self._meta_label(self._company_name_by_id(int(row["company_id"]))), 0, 1)
        details.addWidget(self._meta_key_label("Цель"), 1, 0)
        details.addWidget(self._meta_label((row["purpose"] or "").strip() or "Не заполнена"), 1, 1)
        details.addWidget(self._meta_key_label("Категории"), 2, 0)
        details.addWidget(self._meta_label((row["pd_categories"] or "").strip() or "Не заполнены"), 2, 1)
        details.addWidget(self._meta_key_label("Срок"), 3, 0)
        details.addWidget(self._meta_label((row["validity_period"] or "").strip() or "Не указан"), 3, 1)

        file_row = QWidget()
        file_row_l = QHBoxLayout(file_row)
        file_row_l.setContentsMargins(0, 0, 0, 0)
        file_row_l.setSpacing(8)
        file_row_l.addWidget(self._build_file_badge("DOC", self._path_exists(row["attachment_path"]), "doc"), 0)
        file_row_l.addStretch(1)
        details.addWidget(self._meta_key_label("Файл"), 4, 0)
        details.addWidget(file_row, 4, 1)
        lay.addLayout(details)
        return card

    def _sync_sopd_card_selection(self):
        for record_id, card in self._sopd_cards.items():
            card.set_selected(record_id == self.current_sopd_id)

    def refresh_sopd_page(self):
        if not self.db:
            return
        self.sopd_page.setUpdatesEnabled(False)
        try:
            query = self.search_input.text().strip().lower()
            company_id = self.current_filter_company_id()
            rows = list(self.db.list_sopd_records(company_id)) if company_id else []
            if company_id is None:
                rows = []
                for company in self.db.list_companies():
                    rows.extend(self.db.list_sopd_records(int(company["id"])))
            filtered = []
            for row in rows:
                haystack = " ".join([
                    str(row["consent_type"] or ""),
                    str(row["purpose"] or ""),
                    str(row["pd_categories"] or ""),
                ]).lower()
                if query and query not in haystack:
                    continue
                filtered.append(row)
            filtered.sort(key=lambda row: (self._company_name_by_id(int(row["company_id"])).lower(), int(row["sort_order"] or 0), int(row["id"])))

            self.sopd_count_label.setText(f"{len(filtered)} карточек")
            self._sopd_cards = {}
            if not filtered:
                self._render_empty_list_state(
                    self.sopd_layout,
                    "Карточек СОПД пока нет",
                    "Создайте первую карточку и храните всё в одном понятном месте.",
                    action_text="Добавить карточку",
                    action_callback=self.start_new_sopd_record,
                )
            else:
                self._clear_layout_widgets(self.sopd_layout)
                for row in filtered:
                    card = self._build_sopd_card(row)
                    self._sopd_cards[int(row["id"])] = card
                    self.sopd_layout.addWidget(card)
                self.sopd_layout.addStretch(1)

            visible_ids = {int(row["id"]) for row in filtered}
            if self.current_sopd_id not in visible_ids and self._page_key == "sopd":
                self.clear_right_panel_state("sopd", preferred_company_id=self.current_filter_company_id())
            else:
                self._sync_sopd_card_selection()
        finally:
            self.sopd_page.setUpdatesEnabled(True)
            self.sopd_page.update()

    def load_sopd_records(self):
        self.refresh_sopd_page()

    def _clear_sopd_editor(self, preferred_company_id: Optional[int] = None, show_editor: bool = True):
        self._sopd_form_loading = True
        self.current_sopd_id = None
        target_company_id = preferred_company_id
        if target_company_id is None and self.sopd_company_combo.count():
            current = self.current_filter_company_id()
            target_company_id = current if current is not None else self.sopd_company_combo.itemData(0)
        if target_company_id is not None:
            for i in range(self.sopd_company_combo.count()):
                if self.sopd_company_combo.itemData(i) == target_company_id:
                    self.sopd_company_combo.setCurrentIndex(i)
                    break
        self.sopd_title_edit.setPlainText("")
        self.sopd_purpose_edit.setPlainText("")
        self.sopd_legal_basis_edit.setPlainText("")
        self.sopd_categories_edit.setPlainText("")
        self.sopd_pd_list_edit.setPlainText("")
        self.sopd_subjects_edit.setPlainText("")
        self.sopd_operations_edit.setPlainText("")
        self.sopd_method_edit.setPlainText("")
        self.sopd_transfer_combo.setCurrentText("Не указано")
        self.sopd_transfer_to_edit.setPlainText("")
        self.sopd_validity_edit.setText("")
        self.sopd_description_edit.setPlainText("")
        self._sopd_form_loading = False
        self._sopd_form_dirty = False
        self._update_sopd_file_controls()
        self.sopd_delete_btn.setEnabled(False)
        if show_editor:
            self._set_right_actions_mode("sopd")
            self.right_stack.setCurrentWidget(self.right_sopd_page)
            self.right_panel_title.setText("Новая карточка СОПД")
            self.right_panel_subtitle.setText("Поля и файл карточки.")
        else:
            self._set_right_actions_mode("none")

    def start_new_sopd_record(self):
        if not self._maybe_discard_sopd_changes():
            return
        if self._page_key != "sopd":
            self.set_active_page("sopd")
        self._clear_sopd_editor(preferred_company_id=self.current_filter_company_id())

    def select_sopd(self, record_id: int):
        if self.current_sopd_id == record_id and self.right_stack.currentWidget() == self.right_sopd_page:
            self._sync_sopd_card_selection()
            return
        if not self._maybe_discard_sopd_changes():
            return
        self.current_sopd_id = int(record_id)
        self._load_sopd_into_editor(self.current_sopd_id)
        self._sync_sopd_card_selection()

    def _load_sopd_into_editor(self, record_id: int):
        if not self.db:
            return
        row = self.db.get_sopd_record(record_id)
        if not row:
            self.warn("Карточка СОПД не найдена.")
            return
        self._sopd_form_loading = True
        for i in range(self.sopd_company_combo.count()):
            if self.sopd_company_combo.itemData(i) == int(row["company_id"]):
                self.sopd_company_combo.setCurrentIndex(i)
                break
        self.sopd_title_edit.setPlainText(str(row["consent_type"] or ""))
        self.sopd_purpose_edit.setPlainText(str(row["purpose"] or ""))
        self.sopd_legal_basis_edit.setPlainText(str(row["legal_basis"] or ""))
        self.sopd_categories_edit.setPlainText(str(row["pd_categories"] or ""))
        self.sopd_pd_list_edit.setPlainText(str(row["pd_list"] or ""))
        self.sopd_subjects_edit.setPlainText(str(row["data_subjects"] or ""))
        self.sopd_operations_edit.setPlainText(str(row["processing_operations"] or ""))
        self.sopd_method_edit.setPlainText(str(row["processing_method"] or ""))
        self.sopd_transfer_combo.setCurrentText(str(row["third_party_transfer"] or "Не указано"))
        self.sopd_transfer_to_edit.setPlainText(str(row["transfer_to"] or ""))
        self.sopd_validity_edit.setText(str(row["validity_period"] or ""))
        self.sopd_description_edit.setPlainText(str(row["sopd_description"] or ""))
        self._sopd_form_loading = False
        self._sopd_form_dirty = False
        self._update_sopd_file_controls()
        self._set_right_actions_mode("sopd")
        self.right_stack.setCurrentWidget(self.right_sopd_page)
        self.right_panel_title.setText(str(row["consent_type"] or "Карточка СОПД"))
        self.right_panel_subtitle.setText("Поля и файл карточки.")
        self.sopd_delete_btn.setEnabled(True)

    def _update_sopd_file_controls(self):
        if not self.db or self.current_sopd_id is None:
            self.sopd_file_caption.setText("Файл не загружен")
            self.sopd_file_upload_btn.setEnabled(False)
            self.sopd_file_open_btn.setEnabled(False)
            self.sopd_file_delete_btn.setEnabled(False)
            return
        row = self.db.get_sopd_record(self.current_sopd_id)
        if not row:
            return
        file_abs = self._abs_storage_path(row["attachment_path"])
        file_exists = bool(file_abs and os.path.exists(file_abs))
        self.sopd_file_caption.setText(
            f"Файл: {os.path.basename(file_abs)}" if file_exists else "Файл не загружен"
        )
        self.sopd_file_upload_btn.setText("Заменить файл" if bool(row["attachment_path"]) else "Загрузить файл")
        self.sopd_file_upload_btn.setEnabled(True)
        self.sopd_file_open_btn.setEnabled(file_exists)
        self.sopd_file_delete_btn.setEnabled(bool(row["attachment_path"]))

    def save_sopd_form(self):
        if not self.db:
            return
        company_id = self.sopd_company_combo.currentData()
        if company_id in (None, ""):
            self.warn("Выберите компанию для карточки СОПД.")
            return
        values = {
            "consent_type": self.sopd_title_edit.toPlainText().strip(),
            "purpose": self.sopd_purpose_edit.toPlainText().strip(),
            "legal_basis": self.sopd_legal_basis_edit.toPlainText().strip(),
            "pd_categories": self.sopd_categories_edit.toPlainText().strip(),
            "data_subjects": self.sopd_subjects_edit.toPlainText().strip(),
            "pd_list": self.sopd_pd_list_edit.toPlainText().strip(),
            "processing_operations": self.sopd_operations_edit.toPlainText().strip(),
            "processing_method": self.sopd_method_edit.toPlainText().strip(),
            "third_party_transfer": self.sopd_transfer_combo.currentText().strip(),
            "transfer_to": self.sopd_transfer_to_edit.toPlainText().strip(),
            "sopd_description": self.sopd_description_edit.toPlainText().strip(),
            "validity_period": self.sopd_validity_edit.text().strip(),
        }
        try:
            if self.current_sopd_id is None:
                self.current_sopd_id = self.db.add_sopd_record(int(company_id), values)
            else:
                self.db.update_sopd_record(self.current_sopd_id, values)
            self._sopd_form_dirty = False
            self.load_companies(preferred_company_id=int(company_id))
            if self.current_sopd_id is not None:
                self.select_sopd(self.current_sopd_id)
            self.info("Карточка СОПД сохранена.")
        except Exception as e:
            self.warn(str(e))

    def delete_selected_sopd(self):
        if self.current_sopd_id is None or not self.db:
            return
        row = self.db.get_sopd_record(self.current_sopd_id)
        if not row:
            return
        if not self.confirm(f"Удалить карточку СОПД «{row['consent_type']}»?"):
            return
        try:
            self.delete_sopd_record_file(self.current_sopd_id, ask_confirm=False, silent=True, refresh_after=False)
            self.db.delete_sopd_record(self.current_sopd_id)
            company_id = int(row["company_id"])
            self.clear_right_panel_state("sopd", preferred_company_id=company_id)
            self.load_companies(preferred_company_id=company_id)
        except Exception as e:
            self.warn(str(e))

    def upload_selected_sopd_file(self):
        if self.current_sopd_id is None:
            self.warn("Сначала сохраните карточку СОПД.")
            return
        self.upload_sopd_record_file_dialog(self.current_sopd_id)

    def open_selected_sopd_file(self):
        if self.current_sopd_id is not None:
            self.open_sopd_record_file(self.current_sopd_id)

    def delete_selected_sopd_file(self):
        if self.current_sopd_id is None:
            return
        if self.delete_sopd_record_file(self.current_sopd_id, ask_confirm=True, silent=False, refresh_after=False):
            self.load_companies(preferred_company_id=self.sopd_company_combo.currentData())
            if self.current_sopd_id is not None:
                self.select_sopd(self.current_sopd_id)

    def _selected_settings_company(self) -> Tuple[Optional[int], Optional[str]]:
        item = self.settings_company_list.currentItem()
        if not item:
            return None, None
        return int(item.data(Qt.UserRole)), str(item.data(Qt.UserRole + 1))

    def _refresh_settings_action_controls(self):
        company_id, _ = self._selected_settings_company()
        has_company = company_id is not None
        has_section = has_company and self.settings_section_list.currentItem() is not None

        if hasattr(self, "settings_company_menu_btn"):
            self.settings_company_menu_btn.setEnabled(has_company)
        if hasattr(self, "settings_company_rename_action"):
            self.settings_company_rename_action.setEnabled(has_company)
        if hasattr(self, "settings_company_copy_action"):
            self.settings_company_copy_action.setEnabled(has_company)
        if hasattr(self, "settings_company_delete_action"):
            self.settings_company_delete_action.setEnabled(has_company)

        if hasattr(self, "settings_add_section_btn"):
            self.settings_add_section_btn.setEnabled(has_company)
        if hasattr(self, "settings_section_menu_btn"):
            self.settings_section_menu_btn.setEnabled(has_section)
        if hasattr(self, "settings_section_rename_action"):
            self.settings_section_rename_action.setEnabled(has_section)
        if hasattr(self, "settings_section_delete_action"):
            self.settings_section_delete_action.setEnabled(has_section)

    def _on_settings_company_changed(self):
        company_id, _ = self._selected_settings_company()
        self.settings_section_list.clear()
        self._settings_section_ids = []
        if not company_id or not self.db:
            if hasattr(self, "settings_section_stack"):
                self.settings_section_stack.setCurrentWidget(self.settings_section_empty_page)
            self._refresh_settings_action_controls()
            self._update_settings_summary()
            return
        rows = list(self.db.list_sections_with_doc_counts(company_id))
        for row in rows:
            section_id = int(row["id"])
            item = QListWidgetItem(f"{row['name']}  ·  {int(row['docs_count'] or 0)} док.")
            item.setData(Qt.UserRole, section_id)
            item.setData(Qt.UserRole + 1, str(row["name"]))
            self.settings_section_list.addItem(item)
            self._settings_section_ids.append(section_id)
        if hasattr(self, "settings_section_stack"):
            self.settings_section_stack.setCurrentWidget(
                self.settings_section_list_page if rows else self.settings_section_empty_page
            )
        if self.settings_section_list.count():
            self.settings_section_list.setCurrentRow(0)
        self._refresh_settings_action_controls()
        self._update_settings_summary()

    def refresh_settings_page(self):
        if hasattr(self, "settings_company_empty_label"):
            self.settings_company_empty_label.setVisible(self.settings_company_list.count() == 0)
        if hasattr(self, "settings_section_stack"):
            self.settings_section_stack.setCurrentWidget(
                self.settings_section_list_page if self.settings_section_list.count() else self.settings_section_empty_page
            )
        self._refresh_settings_action_controls()
        self._update_settings_summary()

    def _update_settings_summary(self):
        self.settings_workspace_label.setText(self.shared_root or "Папка ещё не выбрана.")
        self.settings_db_label.setText(self.db_path or "")
        company_id, company_name = self._selected_settings_company()
        if company_id and self.db:
            docs = self.db.count_documents_by_company(company_id)
            sections = len(self.db.list_sections(company_id))
            sopd = self.db.count_sopd_records(company_id)
            self.settings_summary_label.setText(
                f"Аккаунт: {self.current_account}\n"
                f"Компания: {company_name}\n"
                f"Документов: {docs}\n"
                f"Разделов: {sections}\n"
                f"Карточек СОПД: {sopd}"
            )
        else:
            self.settings_summary_label.setText(f"Аккаунт: {self.current_account}")

    def _open_workspace_folder(self):
        if not self.shared_root:
            self.warn("Рабочая папка ещё не выбрана.")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(self.shared_root)))

    def settings_add_company(self):
        self.add_company()

    def settings_rename_company(self):
        company_id, company_name = self._selected_settings_company()
        if not company_id or not company_name:
            self.warn("Сначала выберите компанию.")
            return
        dlg = InputDialog(self, "Переименовать компанию", "Новое название:", default=company_name)
        if dlg.exec() != QDialog.Accepted:
            return
        try:
            self.db.rename_company(company_id, dlg.value())
            self.load_companies(preferred_company_id=company_id)
        except sqlite3.IntegrityError:
            self.warn("Компания с таким названием уже существует.")
        except Exception as e:
            self.warn(str(e))

    def settings_copy_company(self):
        company_id, company_name = self._selected_settings_company()
        if not company_id or not company_name:
            self.warn("Сначала выберите компанию.")
            return
        default_name = self._unique_copy_name(company_name)
        dlg = InputDialog(self, "Копировать компанию", "Название копии:", default=default_name)
        if dlg.exec() != QDialog.Accepted:
            return
        self.copy_company(company_id, company_name, dlg.value())

    def settings_delete_company(self):
        company_id, company_name = self._selected_settings_company()
        if not company_id or not company_name:
            self.warn("Сначала выберите компанию.")
            return
        if not self.confirm(f"Удалить компанию «{company_name}» и все её документы?"):
            return
        try:
            self.db.delete_company(company_id)
            try:
                shutil.rmtree(os.path.join(self.storage_dir, f"company_{company_id}"), ignore_errors=True)
            except Exception:
                pass
            self.load_companies()
        except Exception as e:
            self.warn(str(e))

    def settings_add_section(self):
        company_id, company_name = self._selected_settings_company()
        if not company_id:
            self.warn("Сначала выберите компанию для нового раздела.")
            return
        dlg = InputDialog(self, "Новый раздел", f"Название раздела для «{company_name}»:")
        if dlg.exec() != QDialog.Accepted:
            return
        try:
            self.db.add_section(company_id, dlg.value())
            self.load_companies(preferred_company_id=company_id)
        except sqlite3.IntegrityError:
            self.warn("Раздел с таким названием уже существует.")
        except Exception as e:
            self.warn(str(e))

    def settings_rename_section(self):
        company_id, _ = self._selected_settings_company()
        item = self.settings_section_list.currentItem()
        if not company_id or not item:
            self.warn("Сначала выберите раздел.")
            return
        section_id = int(item.data(Qt.UserRole))
        section_name = str(item.data(Qt.UserRole + 1))
        dlg = InputDialog(self, "Переименовать раздел", "Новое название:", default=section_name)
        if dlg.exec() != QDialog.Accepted:
            return
        try:
            self.db.rename_section(section_id, dlg.value())
            self.load_companies(preferred_company_id=company_id)
        except sqlite3.IntegrityError:
            self.warn("Раздел с таким названием уже существует.")
        except Exception as e:
            self.warn(str(e))

    def settings_delete_section(self):
        company_id, _ = self._selected_settings_company()
        item = self.settings_section_list.currentItem()
        if not company_id or not item:
            self.warn("Сначала выберите раздел.")
            return
        section_id = int(item.data(Qt.UserRole))
        section_name = str(item.data(Qt.UserRole + 1))
        if not self.confirm(f"Удалить раздел «{section_name}»? Документы останутся, исчезнет только привязка к разделу."):
            return
        try:
            self.db.delete_section(section_id)
            self.load_companies(preferred_company_id=company_id)
        except Exception as e:
            self.warn(str(e))

    def _build_document_card(self, row: sqlite3.Row, card_kind: str) -> ClickableCard:
        card = ClickableCard(int(row["id"]))
        card.setObjectName("reviewCard" if card_kind == "review" else "documentCard")
        card.clicked.connect(self.select_document)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(10)
        top.addWidget(self._card_title_label(str(row["doc_title"])), 1)
        if card_kind == "review":
            priority = self._review_priority(row) or "low"
            due_text = fmt_date_ddmmyyyy(row["review_due"] or "") or (row["review_due"] or "")
            top.addWidget(self._priority_badge(priority, due_text), 0, Qt.AlignTop)
        lay.addLayout(top)

        due_text = self._review_due_text(row)
        section_names = str(row["section_names"] or "").strip()
        has_office = self._path_exists(row["office_path"])
        _, issues = self._document_attention_details(row)

        file_row = QWidget()
        file_row_l = QHBoxLayout(file_row)
        file_row_l.setContentsMargins(0, 0, 0, 0)
        file_row_l.setSpacing(8)
        file_row_l.addWidget(self._build_file_badge("PDF", self._document_has_required_pdf(row), "pdf"), 0)
        office_badge_text = self._office_file_label(
            row["office_path"],
            fallback="Нужен файл" if self._document_requires_office(row) else "Нет файла",
        )
        file_row_l.addWidget(
            self._build_file_badge(
                office_badge_text,
                has_office,
                self._office_file_variant(row["office_path"]),
            ),
            0,
        )
        file_row_l.addStretch(1)

        details = QGridLayout()
        details.setContentsMargins(0, 0, 0, 0)
        details.setHorizontalSpacing(16)
        details.setVerticalSpacing(8)
        details.setColumnMinimumWidth(0, 84)
        details.setColumnStretch(1, 2)
        details.setColumnMinimumWidth(2, 80)
        details.setColumnStretch(3, 2)

        details.addWidget(self._meta_key_label("Компания"), 0, 0)
        details.addWidget(self._meta_label(str(row["company_name"] or "Не указана")), 0, 1)

        if card_kind == "review":
            details.addWidget(self._meta_key_label("Пересмотр"), 0, 2)
            details.addWidget(self._meta_label(due_text), 0, 3)
            details.addWidget(self._meta_key_label("Файлы"), 1, 0)
            details.addWidget(file_row, 1, 1)
            if section_names:
                details.addWidget(self._meta_key_label("Раздел"), 1, 2)
                details.addWidget(self._meta_label(section_names), 1, 3)
        else:
            status_key = self._meta_key_label("Статус")
            status_key.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            details.addWidget(status_key, 0, 2, 1, 1, Qt.AlignLeft | Qt.AlignVCenter)
            details.addWidget(
                self._build_status_badge(self._normalize_status(row["status"])),
                0,
                3,
                1,
                1,
                Qt.AlignLeft | Qt.AlignVCenter,
            )
            details.addWidget(self._meta_key_label("Пересмотр"), 1, 0)
            details.addWidget(self._meta_label(due_text), 1, 1)
            details.addWidget(self._meta_key_label("Файлы"), 1, 2)
            details.addWidget(file_row, 1, 3)
            if section_names:
                details.addWidget(self._meta_key_label("Раздел"), 2, 0)
                details.addWidget(self._meta_label(section_names), 2, 1, 1, 3)

        lay.addLayout(details)

        if card_kind != "review" and issues:
            attention = QLabel(" • ".join(issues[:2]))
            attention.setObjectName("analyticsHint")
            attention.setWordWrap(True)
            lay.addWidget(attention)

        if card_kind == "review":
            actions = QHBoxLayout()
            actions.setSpacing(8)
            actions.addStretch(1)
            update_btn = self._create_button("Обновить дату", kind="soft")
            update_btn.clicked.connect(lambda checked=False, doc_id=int(row["id"]): self._start_review_update(doc_id))
            actions.addWidget(update_btn)
            lay.addLayout(actions)
        card.enable_child_click_proxy()
        return card

    def _build_sopd_card(self, row: sqlite3.Row) -> ClickableCard:
        card = ClickableCard(int(row["id"]))
        card.setObjectName("sopdCard")
        card.clicked.connect(self.select_sopd)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(10)
        top.addWidget(self._card_title_label(str(row["consent_type"] or "Карточка СОПД")), 1)
        transfer_badge = QLabel(f"Передача: {row['third_party_transfer'] or 'Не указано'}")
        transfer_badge.setProperty("badgeRole", "status")
        transfer_badge.setProperty("statusKind", "review" if (row["third_party_transfer"] or "").strip() == "Да" else "draft")
        top.addWidget(transfer_badge, 0, Qt.AlignTop)
        lay.addLayout(top)

        file_row = QWidget()
        file_row_l = QHBoxLayout(file_row)
        file_row_l.setContentsMargins(0, 0, 0, 0)
        file_row_l.setSpacing(8)
        file_row_l.addWidget(self._build_file_badge("DOC", self._path_exists(row["attachment_path"]), "doc"), 0)
        file_row_l.addStretch(1)

        details = QGridLayout()
        details.setContentsMargins(0, 0, 0, 0)
        details.setHorizontalSpacing(16)
        details.setVerticalSpacing(8)
        details.setColumnMinimumWidth(0, 84)
        details.setColumnStretch(1, 2)
        details.setColumnMinimumWidth(2, 70)
        details.setColumnStretch(3, 2)

        details.addWidget(self._meta_key_label("Компания"), 0, 0)
        details.addWidget(self._meta_label(self._company_name_by_id(int(row["company_id"]))), 0, 1)
        details.addWidget(self._meta_key_label("Файл"), 0, 2)
        details.addWidget(file_row, 0, 3)
        details.addWidget(self._meta_key_label("Цель"), 1, 0)
        details.addWidget(self._meta_label((row["purpose"] or "").strip() or "Не заполнена"), 1, 1)
        details.addWidget(self._meta_key_label("Срок"), 1, 2)
        details.addWidget(self._meta_label((row["validity_period"] or "").strip() or "Не указан"), 1, 3)
        details.addWidget(self._meta_key_label("Категории"), 2, 0)
        details.addWidget(self._meta_label((row["pd_categories"] or "").strip() or "Не заполнены"), 2, 1, 1, 3)
        lay.addLayout(details)
        card.enable_child_click_proxy()
        return card

    def _show_right_placeholder(self, title: str, text: str):
        self._set_right_actions_mode("none")
        self.right_panel_title.setText(title)
        self.right_panel_subtitle.setText(text)
        self.right_placeholder_title.setText(title)
        self.right_placeholder_text.setText(text)
        self.right_placeholder_text.setVisible(bool(text.strip()))
        self.right_stack.setCurrentWidget(self.right_placeholder_page)

def main():
    app = QApplication([])
    app.setStyle("Fusion")
    w = MainWindow()
    w.show()
    app.exec()


if __name__ == "__main__":
    main()
